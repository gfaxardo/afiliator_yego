"""E2E test with real DB data - Phase 4.5."""
import openpyxl
import io
import json
import requests

# Get real scout from DB
r = requests.get('http://localhost:8000/scout-liq/scouts', timeout=10)
scouts = r.json()
if not scouts:
    print("No scouts in DB. Creating one...")
    r2 = requests.post('http://localhost:8000/scout-liq/scouts', json={
        'scout_name': 'JUAN PEREZ', 'scout_type': 'cabinet', 'status': 'active'
    }, timeout=10)
    print(f"Created: {r2.status_code} - {r2.json().get('id', 'err')}")

# Get a real driver license from source - find one with a license
r3 = requests.get('http://localhost:8000/scout-liq/source/drivers?limit=20', timeout=10)
drivers = r3.json()
real_driver_id = None
real_license = None
for d in drivers.get('drivers', []):
    if d.get('license') and d.get('driver_id'):
        real_driver_id = d['driver_id']
        real_license = d['license']
        break
print(f"Real driver: id={real_driver_id}, license={real_license}")

# Create test XLSX with real scout name
wb = openpyxl.Workbook()
wb.remove(wb.active)
ws = wb.create_sheet('01_PAGOS_HISTORICOS')

headers = [
    'external_payment_id', 'source_file', 'source_sheet', 'source_row',
    'cutoff_external_id', 'cutoff_name', 'cutoff_window_from', 'cutoff_window_to',
    'fecha_pago', 'estado_pago', 'scout_name_raw', 'supervisor_name_raw',
    'scout_type_raw', 'origin_raw', 'driver_license_raw', 'driver_name_raw',
    'driver_phone_raw', 'driver_id_resolved', 'payment_scheme_name',
    'payment_scheme_type', 'payment_rule', 'milestone',
    'trips_0_7_count_reported', 'trips_8_14_count_reported',
    'trips_0_14_count_reported', 'amount_paid', 'currency',
    'payment_component', 'payment_reference', 'paid_by', 'notes'
]
for col, h in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=h)

# Use real scout name from DB if exists, otherwise "JUAN PEREZ" we just created
scout_name = scouts[0]['scout_name'] if scouts else 'JUAN PEREZ'

# Test row with real or semi-real data
test_rows = []
if real_driver_id and real_license:
    test_rows.append([
        'E2E-001', 'e2e_test.xlsx', '01_PAGOS_HISTORICOS', '1',
        'CORTE-2024-06', 'Corte Junio 2024',
        '2024-06-01', '2024-06-30', '2024-06-30',
        'PAGADO', scout_name, 'SUPERVISOR TEST',
        'cabinet', 'app', real_license, 'Driver Test',
        '', real_driver_id, 'Esquema Calidad',
        'quality_conversion', '5_viajes_7_dias', '5_VIAJES',
        '5', '2', '7', '150.00', 'PEN',
        'scout_driver_payment', 'REF-E2E', 'Admin', 'E2E test payment'
    ])

# Also add row with driver_id_resolved pre-filled
if real_driver_id:
    test_rows.append([
        'E2E-002', 'e2e_test.xlsx', '01_PAGOS_HISTORICOS', '2',
        'CORTE-2024-07', 'Corte Julio 2024',
        '2024-07-01', '2024-07-31', '2024-07-31',
        'PAID', scout_name, '',
        'cabinet', '', '', '',
        '', real_driver_id, 'Esquema Calidad',
        'quality_conversion', '1_viaje', '1_VIAJE',
        '1', '0', '1', '50.00', 'PEN',
        'scout_driver_payment', 'REF-E2E-2', 'Admin', 'E2E test 2'
    ])

if not test_rows:
    print("Warning: no real driver data available. Using generic test data.")
    test_rows = [
        ['E2E-001', 'e2e_test.xlsx', '01_PAGOS_HISTORICOS', '1',
         'CORTE-2024-06', 'Corte Junio', '2024-06-01', '2024-06-30',
         '2024-06-30', 'PAGADO', scout_name, '',
         'cabinet', 'app', 'Q12345678', 'Driver Test', '', '',
         'Esquema Calidad', 'quality_conversion', '5_viajes', '5_VIAJES',
         '5', '0', '5', '150.00', 'PEN',
         'scout_driver_payment', 'REF-E2E', 'Admin', 'test']
    ]

for r_idx, row_data in enumerate(test_rows, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws.cell(row=r_idx, column=c_idx, value=val)

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

url = 'http://localhost:8000/scout-liq/historical-imports/preview'
files = {'file': ('e2e_test.xlsx', buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
r = requests.post(url, files=files, params={'sheet': '01_PAGOS_HISTORICOS'}, timeout=30)

if r.status_code != 200:
    print(f"ERROR HTTP {r.status_code}: {r.text}")
    exit(1)

result = r.json()

summary = {k: v for k, v in result.items() if k != 'lines'}
print(json.dumps(summary, indent=2, default=str))

print(f"\nLines:")
for l in result.get('lines', []):
    status_icon = "OK" if l.get('import_status') == 'ready_to_import' else l.get('import_status', '?')
    print(f"  Row={l.get('source_row')} [{status_icon}] scout_id={l.get('scout_id_resolved')} driver_id={l.get('driver_id_resolved')} amt={l.get('amount_paid')} reason={l.get('import_reason','')}")

print(f"\n=== SUMMARY ===")
print(f"Ready to import: {summary.get('ready_to_import')}")
print(f"Rejected: {summary.get('rejected')}")
print(f"Manual review: {summary.get('manual_review')}")
print(f"Duplicate: {summary.get('duplicate')}")
print(f"Amount ready: S/ {summary.get('amount_ready')}")

# COMMIT if there are importable rows
if summary.get('ready_to_import', 0) > 0 and summary.get('batch_id'):
    print(f"\n=== COMMITTING batch #{summary['batch_id']} ===")
    r2 = requests.post(
        f'http://localhost:8000/scout-liq/historical-imports/commit',
        params={'batch_id': summary['batch_id'], 'uploaded_by': 'e2e_test'},
        timeout=30)
    if r2.status_code == 200:
        cr = r2.json()
        print(f"Commit result: imported={cr.get('imported')} rejected={cr.get('rejected')} "
              f"manual_review={cr.get('manual_review')} duplicate={cr.get('duplicate')} "
              f"amount=S/ {cr.get('amount_imported')}")
    else:
        print(f"Commit failed: {r2.status_code} {r2.text}")

    # Check paid_history
    r3 = requests.get(
        'http://localhost:8000/scout-liq/paid-history',
        params={'import_source': 'historical_upload', 'limit': 10},
        timeout=10)
    ph = r3.json()
    print(f"\nPaid History (historical_upload): {ph.get('total')} records")
    for item in ph.get('items', [])[:5]:
        print(f"  PH#{item['id']} scout={item['scout_id']} driver={item.get('driver_id')} "
              f"amt=S/ {item['amount_paid']} component={item.get('payment_component')} "
              f"hash={item.get('unique_hash','')[:16]}...")

    # Download errors CSV
    errors_url = f"http://localhost:8000/scout-liq/historical-imports/{summary['batch_id']}/errors.csv"
    r4 = requests.get(errors_url, timeout=10)
    if r4.status_code == 200:
        print(f"\nErrors CSV: {len(r4.text)} bytes")
        for line in r4.text.split('\r\n')[:5]:
            print(f"  {line}")
    else:
        print(f"Errors CSV status: {r4.status_code}")

print("\n=== E2E TEST COMPLETE ===")
