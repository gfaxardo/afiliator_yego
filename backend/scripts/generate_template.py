"""
Generate the AFILIATOR standard import template XLSX.
Run: python scripts/generate_template.py
Output: static/Plantilla_AFILIATOR_Carga_Historica_Esquemas_Manual.xlsx
"""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "static",
                           "Plantilla_AFILIATOR_Carga_Historica_Esquemas_Manual.xlsx")

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
REQUIRED_FONT = Font(name="Calibri", size=11, italic=True, color="CC7A00")
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SHEETS = {
    "00_README": {
        "rows": [
            ["PLANTILLA OFICIAL - Liquidador de Calidad Scouts Yego"],
            ["Version: 1.0"],
            [""],
            ["INSTRUCCIONES:"],
            ["1. Cada hoja tiene un proposito especifico. Lee las VALIDACIONES_IMPORT antes de llenar."],
            ["2. Columnas con fondo AMARILLO son obligatorias para importacion."],
            ["3. Columnas con fondo AZUL son encabezados. NO modificar la fila de encabezados."],
            ["4. Consulta la hoja LISTAS para valores permitidos en columnas con dropdown."],
            ["5. Sube el archivo lleno desde la UI en la tab 'Carga Historica'."],
            ["6. El sistema hara preview sin insertar datos. Revisa el resultado antes del commit."],
            ["7. Solo se insertaran en paid_history las filas validadas."],
            ["8. Las filas rechazadas o en revision manual se pueden descargar como CSV."],
            [""],
            ["HOJAS:"],
            ["  - 01_PAGOS_HISTORICOS: Pagos realizados en cortes anteriores."],
            ["  - 02_SCOUTS: Registro masivo de scouts."],
            ["  - 03_ESQUEMAS: Esquemas de pago versionados."],
            ["  - 04_PAGOS_MANUALES: Pagos manuales o ajustes."],
            ["  - 05_SUPERVISORES_BONOS: Comisiones de supervisor y bonos a scouts."],
            ["  - 06_ATRIBUCIONES_HISTORICAS: Atribuciones historicas scout->conductor."],
            ["  - LISTAS: Valores permitidos para columnas con listas."],
            ["  - VALIDACIONES_IMPORT: Reglas de validacion del sistema."],
        ]
    },
    "01_PAGOS_HISTORICOS": {
        "columns": [
            ("external_payment_id", False),
            ("source_file", False),
            ("source_sheet", False),
            ("source_row", False),
            ("cutoff_external_id", False),
            ("cutoff_name", False),
            ("cutoff_window_from", False),
            ("cutoff_window_to", False),
            ("fecha_pago", False),
            ("estado_pago", True),        # REQUIRED
            ("scout_name_raw", True),     # REQUIRED
            ("supervisor_name_raw", False),
            ("scout_type_raw", False),
            ("origin_raw", False),
            ("driver_license_raw", True),  # REQUIRED
            ("driver_name_raw", False),
            ("driver_phone_raw", False),
            ("driver_id_resolved", False),
            ("payment_scheme_name", False),
            ("payment_scheme_type", False),
            ("payment_rule", True),        # REQUIRED
            ("milestone", False),
            ("trips_0_7_count_reported", False),
            ("trips_8_14_count_reported", False),
            ("trips_0_14_count_reported", False),
            ("amount_paid", True),         # REQUIRED
            ("currency", True),            # REQUIRED
            ("payment_component", False),
            ("payment_reference", False),
            ("paid_by", False),
            ("notes", False),
        ],
        "example": [
            "EXT-001", "", "", "1", "CORTE-2024-01", "Corte Enero 2024",
            "2024-01-01", "2024-01-31", "2024-01-31", "PAGADO",
            "JUAN PEREZ", "MARIA GOMEZ", "cabinet", "app",
            "Q12345678", "Carlos Lopez", "999888777", "",
            "Esquema Cabinet", "legacy_milestone", "conexion", "CONEXION",
            "1", "0", "1", "50.00", "PEN", "scout_driver_payment",
            "REF-001", "Admin", "Pago conexion",
        ],
    },
    "02_SCOUTS": {
        "columns": [
            ("scout_name", True),
            ("document_number", False),
            ("phone", False),
            ("email", False),
            ("country", False),
            ("city", False),
            ("scout_type", False),
            ("supervisor_name_raw", False),
            ("active_from", False),
            ("active_to", False),
            ("status", False),
            ("external_key", False),
            ("notes", False),
            ("source_sheet", False),
            ("source_row", False),
        ],
        "example": [
            "JUAN PEREZ", "12345678", "999888777", "juan@email.com",
            "PE", "Lima", "cabinet", "MARIA GOMEZ",
            "2024-01-01", "", "active", "SCT-001", "Scout de prueba",
            "PADRON_SCOUTS", "1",
        ],
    },
    "03_ESQUEMAS": {
        "columns": [
            ("scheme_name", True),
            ("scheme_type", True),
            ("origin", False),
            ("scout_type", False),
            ("valid_from", False),
            ("valid_to", False),
            ("active", False),
            ("metric_code", False),
            ("base_pago", False),
            ("min_affiliations", False),
            ("min_conversion_rate", False),
            ("payment_per_converted_driver", False),
            ("milestone", False),
            ("required_trips", False),
            ("window_start_day", False),
            ("window_end_day", False),
            ("amount", False),
            ("currency", False),
            ("commission_rate", False),
            ("config_notes", False),
        ],
        "example": [
            "Esquema Calidad 5v7d", "quality_conversion", "cabinet", "cabinet",
            "2024-01-01", "2025-12-31", "Si", "5plus_0_7", "por_convertido",
            "1", "0.05", "150.00", "5_VIAJES_7_DIAS", "5", "0", "7",
            "", "PEN", "", "Pago por calidad de conversion",
        ],
    },
    "04_PAGOS_MANUALES": {
        "columns": [
            ("payment_date", False),
            ("scout_name", True),
            ("driver_license_raw", False),
            ("driver_id", False),
            ("payment_scheme_name", False),
            ("payment_rule", False),
            ("milestone", False),
            ("amount", True),
            ("currency", True),
            ("reason", True),
            ("payment_component", False),
            ("supervisor_name", False),
            ("cutoff_external_id", False),
            ("payment_reference", False),
            ("status", False),
            ("notes", False),
        ],
        "example": [
            "2024-06-15", "JUAN PEREZ", "Q12345678", "DRV-001",
            "Manual", "ajuste", "", "100.00", "PEN",
            "Ajuste por bono excepcional", "manual_adjustment",
            "", "", "", "draft", "Aprobado por gerencia",
        ],
    },
    "05_SUPERVISORES_BONOS": {
        "columns": [
            ("cutoff_external_id", False),
            ("supervisor_name", False),
            ("scout_name", True),
            ("payment_component", True),
            ("commission_rate", False),
            ("base_amount", False),
            ("bonus_type", False),
            ("amount", True),
            ("currency", True),
            ("reason", True),
            ("status", False),
            ("payment_reference", False),
            ("notes", False),
        ],
        "example": [
            "CORTE-2024-06", "MARIA GOMEZ", "JUAN PEREZ",
            "supervisor_commission", "0.10", "500.00", "",
            "50.00", "PEN", "Comision 10% corte Junio 2024",
            "draft", "", "",
        ],
    },
    "06_ATRIBUCIONES_HISTORICAS": {
        "columns": [
            ("external_attribution_id", False),
            ("source_file", False),
            ("source_sheet", False),
            ("source_row", False),
            ("cutoff_external_id", False),
            ("assignment_date", False),
            ("hire_date", False),
            ("fecha_atribucion", False),
            ("tipo_evento", False),
            ("scout_name_raw", True),
            ("supervisor_name_raw", False),
            ("scout_type_raw", False),
            ("origin_raw", False),
            ("driver_license_raw", True),
            ("driver_name_raw", False),
            ("driver_phone_raw", False),
            ("driver_id_resolved", False),
            ("payment_status_raw", False),
            ("payment_amount_raw", False),
            ("payment_rule_raw", False),
            ("ok_1_viaje_raw", False),
            ("ok_5_viajes_raw", False),
            ("ok_25_viajes_raw", False),
            ("ok_50_viajes_raw", False),
            ("notes", False),
        ],
        "example": [
            "ATTR-001", "import.xlsx", "Registros-conductores", "1",
            "CORTE-2024-01", "2024-01-15", "2024-01-01",
            "2024-01-01", "new",
            "JUAN PEREZ", "MARIA GOMEZ", "cabinet", "app",
            "Q12345678", "Carlos Lopez", "999888777", "",
            "PAGADO", "50.00", "conexion",
            "SI", "NO", "NO", "NO",
            "Atribucion historica de prueba",
        ],
    },
    "LISTAS": {
        "rows": [
            ["LISTA DE VALORES PERMITIDOS"],
            [""],
            ["estado_pago (01_PAGOS_HISTORICOS):"],
            ["  PAGADO"],
            ["  APROBADO"],
            ["  PAID"],
            ["  EXCLUIDO (sera rechazado)"],
            ["  NO ELEGIBLE (sera rechazado)"],
            ["  NO ALCANZADO (sera rechazado)"],
            [""],
            ["payment_scheme_type:"],
            ["  legacy_milestone"],
            ["  quality_conversion"],
            ["  manual"],
            [""],
            ["payment_component:"],
            ["  scout_driver_payment"],
            ["  supervisor_commission"],
            ["  scout_bonus"],
            ["  manual_adjustment"],
            [""],
            ["scheme_type (03_ESQUEMAS):"],
            ["  legacy_milestone"],
            ["  quality_conversion"],
            ["  manual"],
            ["  supervisor_commission"],
            ["  bonus"],
            [""],
            ["currency:"],
            ["  PEN"],
            ["  USD"],
            [""],
            ["bonus_type:"],
            ["  best_scout"],
            [""],
            ["milestone (legacy):"],
            ["  CONEXION"],
            ["  1_VIAJE"],
            ["  5_VIAJES"],
            ["  25_VIAJES"],
            ["  50_VIAJES"],
            [""],
            ["status (scouts):"],
            ["  active"],
            ["  inactive"],
        ],
    },
    "VALIDACIONES_IMPORT": {
        "rows": [
            ["REGLAS DE VALIDACION DEL SISTEMA"],
            [""],
            ["1. PAGOS HISTORICOS (01_PAGOS_HISTORICOS):"],
            ["   - estado_pago en (PAGADO, APROBADO, PAID) => intenta importar"],
            ["   - estado_pago en (EXCLUIDO, NO ELEGIBLE, NO ALCANZADO) => RECHAZADO (rejected_not_paid)"],
            ["   - amount_paid <= 0 => RECHAZADO (rejected_invalid_amount)"],
            ["   - driver_license_raw no resuelve a driver_id => MANUAL_REVIEW (manual_review_no_driver_match)"],
            ["   - scout_name_raw no resuelve a scout => MANUAL_REVIEW (manual_review_no_scout_match)"],
            ["   - fila duplicada por hash o logica => DUPLICATE (duplicate_skipped)"],
            [""],
            ["2. SCOUTS (02_SCOUTS):"],
            ["   - scout_name normalizado ya existe y sin cambios => SKIP"],
            ["   - scout_name normalizado existe con cambios (supervisor, tipo) => UPDATE"],
            ["   - scout_name no existe => CREATE"],
            [""],
            ["3. ESQUEMAS (03_ESQUEMAS):"],
            ["   - scheme_name + scheme_type + valid_from ya existe => SKIP"],
            ["   - nuevo => CREATE con versionado"],
            [""],
            ["4. PAGOS MANUALES (04_PAGOS_MANUALES):"],
            ["   - amount > 0 y reason no vacio => intenta crear"],
            ["   - scout debe resolverse"],
            ["   - mark-paid crea paid_history con import_source=manual_payment"],
            [""],
            ["5. SUPERVISORES Y BONOS (05_SUPERVISORES_BONOS):"],
            ["   - payment_component = supervisor_commission o scout_bonus"],
            ["   - commission_rate configurable (default 0.10)"],
            ["   - bonus_type = best_scout"],
            [""],
            ["6. GENERAL:"],
            ["   - PREVIEW NUNCA inserta en paid_history"],
            ["   - COMMIT solo inserta filas ready_to_import"],
            ["   - Toda fila importada guarda: source_file, source_sheet, source_row, unique_hash, import_source"],
            ["   - Errores y manual_review se pueden descargar como CSV"],
        ],
    },
}


def _apply_header_style(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_required_row(ws, sheet_name, num_cols, required_cols):
    req_row = ws.max_row + 1
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=req_row, column=col)
        if (col - 1) in required_cols:
            cell.fill = REQUIRED_FILL
            cell.font = REQUIRED_FONT
            cell.value = "OBLIGATORIO"
            cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _write_example_row(ws, data, num_cols):
    row = ws.max_row + 1
    for col in range(1, num_cols + 1):
        val = data[col - 1] if col - 1 < len(data) else ""
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER


def _build_sheet_with_columns(ws, sheet_name, sheet_def):
    columns = sheet_def["columns"]
    example = sheet_def.get("example", [])

    for col_idx, (col_name, required) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
    _apply_header_style(ws, len(columns))

    required_indices = [i for i, (_, r) in enumerate(columns) if r]
    _apply_required_row(ws, sheet_name, len(columns), required_indices)

    if example:
        _write_example_row(ws, example, len(columns))

    # Adjust column widths
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(columns[col_idx - 1][0])
        ws.column_dimensions[col_letter].width = max(header_len + 4, 16)

    ws.freeze_panes = "A4"


def _build_info_sheet(ws, rows):
    for i, row_data in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=row_data[0] if isinstance(row_data, list) else row_data)
        if i == 1:
            cell.font = Font(name="Calibri", size=14, bold=True)
        elif isinstance(row_data, list) and row_data[0] and row_data[0].startswith("  "):
            cell.font = Font(name="Calibri", size=11)
        else:
            cell.font = NORMAL_FONT
    ws.column_dimensions["A"].width = 80
    ws.freeze_panes = "A1"


def generate():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, sheet_def in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        if "rows" in sheet_def:
            _build_info_sheet(ws, sheet_def["rows"])
        elif "columns" in sheet_def:
            _build_sheet_with_columns(ws, sheet_name, sheet_def)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Template generated: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    generate()
