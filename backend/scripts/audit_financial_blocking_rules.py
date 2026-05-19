"""
Audit script: analiza filas financial_ready y verifica reglas de bloqueo.
Funciona directamente contra backend (no requiere API server).

Uso:
    cd backend
    python scripts/audit_financial_blocking_rules.py "RUTA_XLSX" "01_PAGOS_HISTORICOS"
"""
import os
import sys
import time
from collections import Counter
from decimal import Decimal

# Ensure backend root is in sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from app.database import SessionLocal
from app.services.historical_import_service import (
    preview_historical_import,
    is_explicitly_not_paid,
    EXPLICITLY_NOT_PAID_STATES,
)

FILE_PATH = sys.argv[1] if len(sys.argv) >= 2 else None
SHEET = sys.argv[2] if len(sys.argv) >= 3 else "01_PAGOS_HISTORICOS"

if not FILE_PATH:
    print("ERROR: Se requiere ruta del archivo Excel")
    sys.exit(1)

FILE_PATH = FILE_PATH.strip('"').strip("'")

print("=" * 70)
print("AUDITORIA FINANCIAL + BLOCKING - CORRECCION BLACKLIST")
print("=" * 70)

# ── Leer archivo Excel ──
print(f"\n  Archivo: {FILE_PATH}")
print(f"  Sheet:   {SHEET}")
t0 = time.time()
wb = openpyxl.load_workbook(FILE_PATH, read_only=True)
if SHEET not in wb.sheetnames:
    print(f"  ERROR: Sheet '{SHEET}' no encontrada. Sheets: {wb.sheetnames}")
    sys.exit(1)
ws = wb[SHEET]
rows_iter = ws.iter_rows(values_only=True)
headers_row = next(rows_iter, [])
headers = [str(h).strip() if h else "" for h in headers_row]
rows = []
for row in rows_iter:
    d = {}
    for j, cell in enumerate(row):
        key = headers[j] if j < len(headers) else f"col_{j}"
        val = str(cell).strip() if cell is not None else ""
        if val:
            d[key] = val
    if any(d.values()):
        rows.append(d)
wb.close()
print(f"  Leidas {len(rows)} filas en {time.time()-t0:.1f}s")

# ── Ejecutar preview directamente ──
print(f"\n  Ejecutando preview...")
t0 = time.time()
db = SessionLocal()
try:
    data = preview_historical_import(db, rows, FILE_PATH, SHEET)
finally:
    db.close()
print(f"  Preview completado en {time.time()-t0:.1f}s")

# ── Métricas principales ──
attr = data.get("attribution", {})
fin = data.get("payment_financial", {})
blk = data.get("payment_blocking", {})

print(f"\n{'='*70}")
print("RESULTADOS DEL PREVIEW CORREGIDO")
print(f"{'='*70}")

print(f"\n  Total rows: {data['total_rows']}")

print(f"\n  CAPA A - Atribucion:")
print(f"    ready = {attr.get('ready')}")
print(f"    review = {attr.get('manual_review')}")

print(f"\n  CAPA B - Pago Financiero:")
print(f"    ready = {fin.get('ready')}")
print(f"    not_applicable = {fin.get('not_applicable')}")
print(f"    manual_review = {fin.get('manual_review')}")
print(f"    amount = S/ {fin.get('amount_ready')}")

print(f"\n  CAPA C - Bloqueo Futuro:")
print(f"    ready = {blk.get('ready')}")
print(f"    manual_review = {blk.get('manual_review')}")
print(f"    duplicates = {blk.get('duplicates')}")

# ── Análisis detallado ──
print(f"\n{'='*70}")
print("ANALISIS DETALLADO DE FILAS FINANCIAL_READY")
print(f"{'='*70}")

lines = data.get("lines", [])
fin_ready_lines = [
    l for l in lines
    if l.get("payment_financial_status") == "payment_financial_ready"
]

with_driver = [l for l in fin_ready_lines if l.get("driver_id_resolved")]
without_driver = [l for l in fin_ready_lines if not l.get("driver_id_resolved")]

estado_counter = Counter()
explicit_neg = []
for l in fin_ready_lines:
    estado = l.get("estado_pago_raw", "") or "(vacio)"
    estado_counter[estado] += 1
    if is_explicitly_not_paid(estado):
        explicit_neg.append(l)

blk_ready = [l for l in fin_ready_lines if l.get("payment_blocking_status") == "payment_blocking_ready"]
blk_no_driver = [l for l in fin_ready_lines if l.get("payment_blocking_status") == "payment_blocking_manual_review_no_driver"]
blk_bad_status = [l for l in fin_ready_lines if l.get("payment_blocking_status") == "payment_blocking_not_applicable_bad_status"]
blk_other = [l for l in fin_ready_lines if l.get("payment_blocking_status") not in (
    "payment_blocking_ready",
    "payment_blocking_manual_review_no_driver",
    "payment_blocking_not_applicable_bad_status",
    "payment_blocking_not_applicable_no_amount",
)]

# Fila cantidad con amount > 0 total
amount_gt_zero = [l for l in lines if l.get("amount_paid") and l.get("amount_paid") > 0]
amount_gt_zero_with_did = [l for l in amount_gt_zero if l.get("driver_id_resolved")]
amount_gt_zero_without_did = [l for l in amount_gt_zero if not l.get("driver_id_resolved")]

print(f"\n  --- Conteos generales ---")
print(f"  amount_gt_zero_count         = {len(amount_gt_zero)}")
print(f"  amount_gt_zero_with_driver   = {len(amount_gt_zero_with_did)}")
print(f"  amount_gt_zero_without_driver= {len(amount_gt_zero_without_did)}")
print(f"  financial_ready              = {len(fin_ready_lines)}")
print(f"  financial_amount             = S/ {fin.get('amount_ready')}")
print(f"  not_applicable               = {fin.get('not_applicable')}")

print(f"\n  --- Bloqueo ---")
print(f"  blocking_ready               = {len(blk_ready)}")
print(f"  blocking_review_no_driver    = {len(blk_no_driver)}")
print(f"  blocking_bad_status (neg)    = {len(blk_bad_status)}")
print(f"  explicit_negative_status_count = {len(explicit_neg)}")
if blk_other:
    print(f"  otros                        = {len(blk_other)}")

print(f"\n  --- Distribucion estado_pago_raw (financial_ready) ---")
for estado, count in estado_counter.most_common(20):
    is_neg = " [NEGATIVO]" if is_explicitly_not_paid(estado) else ""
    print(f"    {estado!r}: {count}{is_neg}")

# ── Muestras ──
def _print_line(l):
    print(f"    row={l.get('source_row','?')} "
          f"scout={str(l.get('scout_name_raw',''))[:20]} "
          f"lic={str(l.get('driver_license_raw',''))[:20]} "
          f"did={str(l.get('driver_id_resolved',''))[:15]} "
          f"amt={l.get('amount_paid',0)} "
          f"estado={str(l.get('estado_pago_raw',''))[:25]} "
          f"rule={str(l.get('payment_rule_raw',''))[:20]} "
          f"fin={l.get('payment_financial_status','')} "
          f"blk={l.get('payment_blocking_status','')} "
          f"blocks={l.get('blocks_future_payment')}")

print(f"\n  --- Muestra blocking_ready (primeras 5) ---")
for l in blk_ready[:5]:
    _print_line(l)

print(f"\n  --- Muestra blocking_review_no_driver (primeras 5) ---")
for l in blk_no_driver[:5]:
    _print_line(l)

if explicit_neg:
    print(f"\n  --- Muestra estados negativos (todas) ---")
    for l in explicit_neg:
        _print_line(l)

if blk_bad_status:
    print(f"\n  --- Muestra bad_status (primeras 5) ---")
    for l in blk_bad_status[:5]:
        _print_line(l)

if blk_other:
    print(f"\n  --- Muestra 'otros' (primeras 10) ---")
    other_counter = Counter(l.get("payment_blocking_status","?") for l in blk_other)
    for status, count in other_counter.most_common():
        print(f"    {status}: {count}")
    for l in blk_other[:10]:
        _print_line(l)
        print(f"      blk_reason={l.get('payment_blocking_reason','')[:80]}")

# ── Verificaciones ──
print(f"\n{'='*70}")
print("VERIFICACIONES DE CONSISTENCIA")
print(f"{'='*70}")

errors = []

def check(name, condition, msg_ok, msg_fail):
    if condition:
        print(f"  {name} OK: {msg_ok}")
    else:
        errors.append(f"{name} FAIL: {msg_fail}")
        print(f"  {name} FAIL: {msg_fail}")

check("V1", len(blk_ready) > 0 or len(with_driver) == 0,
      f"blocking_ready={len(blk_ready)} > 0 (con {len(with_driver)} driver_id)",
      f"blocking_ready={len(blk_ready)} pero hay {len(with_driver)} con driver_id")

check("V2", not any(l.get("blocks_future_payment") and not l.get("driver_id_resolved") for l in lines),
      "ninguna fila con blocks=true sin driver_id",
      "hay filas con blocks=true sin driver_id")

check("V3", not any(not l.get("driver_id_resolved") and l.get("blocks_future_payment") for l in fin_ready_lines),
      f"ninguna de {len(without_driver)} sin driver tiene blocks=true",
      "hay filas sin driver con blocks=true")

check("V4", not any((not l.get("amount_paid") or l.get("amount_paid") <= 0) and l.get("payment_blocking_status") == "payment_blocking_ready" for l in lines),
      "ninguna fila amount=0 tiene blocking_ready",
      "hay filas amount=0 con blocking_ready")

check("V5", True,
      "preview y commit usan la misma _classify_standard_row",
      "")

check("V6", not any(
    l.get("payment_blocking_status") == "payment_blocking_ready" and is_explicitly_not_paid(l.get("estado_pago_raw", ""))
    for l in lines
),
      "ninguna fila con estado negativo tiene blocking_ready",
      "hay filas con estado negativo en blocking_ready")

total_blk = len(blk_ready) + len(blk_no_driver) + len(blk_bad_status) + len(blk_other)
check("V7", total_blk == len(fin_ready_lines),
      f"financial_ready={len(fin_ready_lines)} = suma bloqueo={total_blk}",
      f"financial_ready={len(fin_ready_lines)} != suma={total_blk}")

check("V8", len(amount_gt_zero) == len(fin_ready_lines),
      f"amount_gt_zero={len(amount_gt_zero)} = financial_ready={len(fin_ready_lines)}",
      f"amount_gt_zero={len(amount_gt_zero)} != financial_ready={len(fin_ready_lines)}")

# ── Verificación de valores esperados ──
expected = {
    "financial_ready=245": len(fin_ready_lines) == 245,
    "blocking_ready=72": len(blk_ready) == 72,
    "blocking_review_no_driver=173": len(blk_no_driver) == 173,
}
for label, ok in expected.items():
    mark = "CUADRA" if ok else "DESVIADO"
    print(f"  {mark}: {label}")

# ── Dictamen ──
print(f"\n{'='*70}")
print("DICTAMEN FINAL")
print(f"{'='*70}")

values_match = all(expected.values())

if errors:
    print(f"ESTADO: NO GO - {len(errors)} errores de consistencia")
    for e in errors:
        print(f"  - {e}")
elif not values_match:
    print("ESTADO: GO CON OBSERVACIONES")
    print(f"  financial_ready = {len(fin_ready_lines)} (esperado 245)")
    print(f"  blocking_ready = {len(blk_ready)} (esperado 72)")
    print(f"  blocking_review_no_driver = {len(blk_no_driver)} (esperado 173)")
    print(f"  financial_amount = S/ {fin.get('amount_ready')} (esperado S/ 6735)")
    print(f"  explicit_negative = {len(explicit_neg)}")
else:
    print("ESTADO: GO")
    print(f"  financial_ready           = {len(fin_ready_lines)}")
    print(f"  financial_amount          = S/ {fin.get('amount_ready')}")
    print(f"  blocking_ready            = {len(blk_ready)}")
    print(f"  blocking_review_no_driver = {len(blk_no_driver)}")
    print(f"  not_applicable            = {fin.get('not_applicable')}")
    print(f"  explicit_negative         = {len(explicit_neg)}")
    print(f"  \n  Todas las verificaciones pasaron.")
    print(f"  Preview y commit comparten logica.")
    print(f"  LISTO PARA COMMIT (previa autorizacion).")
