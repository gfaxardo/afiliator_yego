"""Quick preview test after bugfix."""
import json
from app.main import app
from fastapi.testclient import TestClient
import openpyxl, io

client = TestClient(app)

wb = openpyxl.Workbook()
wb.remove(wb.active)
ws = wb.create_sheet('01_PAGOS_HISTORICOS')
for col, h in enumerate(['estado_pago','scout_name_raw','driver_license_raw','amount_paid','currency','payment_rule'], 1):
    ws.cell(row=1, column=col, value=h)
ws.cell(row=2, column=1, value='PAGADO')
ws.cell(row=2, column=2, value='Scout Validacion F1.5')
ws.cell(row=2, column=3, value='Q25822973')
ws.cell(row=2, column=4, value='150.00')
ws.cell(row=2, column=5, value='PEN')
ws.cell(row=2, column=6, value='5_viajes')
buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

r = client.post(
    '/scout-liq/historical-imports/preview?sheet=01_PAGOS_HISTORICOS',
    files={'file': ('test2.xlsx', buf, 'application/octet-stream')})
print(f'Status: {r.status_code}')
data = r.json()
print(json.dumps({k:v for k,v in data.items() if k != 'lines'}, indent=2, default=str))
for l in data.get('lines', []):
    print(f'  Row={l.get("source_row")} status={l.get("import_status")} scout={l.get("scout_id_resolved")} driver={l.get("driver_id_resolved")} amt={l.get("amount_paid")} reason={l.get("import_reason","")}')

# Test commit if ready
if data.get('ready_to_import', 0) > 0:
    batch_id = data['batch_id']
    r2 = client.post(f'/scout-liq/historical-imports/commit?batch_id={batch_id}&uploaded_by=test')
    print(f'\nCommit: {r2.status_code}')
    print(r2.json())

    r3 = client.get('/scout-liq/paid-history?import_source=historical_upload&limit=5')
    ph = r3.json()
    print(f"\nPaid History: {ph['total']} records")
    for item in ph['items']:
        print(f"  PH#{item['id']} scout={item['scout_id']} driver={item.get('driver_id')} amt={item['amount_paid']}")

    # Verify NOT duplicated
    print("\n=== Running preview again to check duplicate detection ===")
    buf2 = io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    r4 = client.post(
        '/scout-liq/historical-imports/preview?sheet=01_PAGOS_HISTORICOS',
        files={'file': ('test2.xlsx', buf2, 'application/octet-stream')})
    data2 = r4.json()
    for l in data2.get('lines', []):
        print(f'  Row={l.get("source_row")} status={l.get("import_status")} reason={l.get("import_reason","")}')
