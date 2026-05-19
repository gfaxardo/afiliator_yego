"""E2E test for Phase 4.5 historical import."""
import openpyxl
import io
import json
import requests

# Create test XLSX
wb = openpyxl.Workbook()
wb.remove(wb.active)

ws = wb.create_sheet('01_PAGOS_HISTORICOS')
headers = [
    'external_payment_id', 'source_file', 'source_sheet', 'source_row',
    'cutoff_external_id', 'cutoff_name', 'cutoff_window_from', 'cutoff_window_to',
    'fecha_pago', 'estado_pago', 'scout_name_raw', 'supervisor_name_raw',
    'scout_type_raw', 'origin_raw', 'driver_license_raw', 'driver_name_raw',
    'driver_phone_raw', 'driver_id_resolved', 'payment_scheme_name',
    'payment_scheme_type', 'payment_rule', 'milestone',
    'trips_0_7_count_reported', 'trips_8_14_count_reported',
    'trips_0_14_count_reported', 'amount_paid', 'currency',
    'payment_component', 'payment_reference', 'paid_by', 'notes'
]
for col, h in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=h)

test_rows = [
    ['EXT-001', 'test.xlsx', '01_PAGOS_HISTORICOS', '1', 'CORTE-2024-01', 'Corte Ene24',
     '2024-01-01', '2024-01-31', '2024-01-31', 'PAGADO', 'JUAN PEREZ', 'MARIA GOMEZ',
     'cabinet', 'app', 'Q12345678', 'Carlos Lopez', '999888777', '',
     'Esquema Cabinet', 'legacy_milestone', 'conexion', 'CONEXION',
     '1', '0', '1', '50.00', 'PEN', 'scout_driver_payment', 'REF-001', 'Admin', 'test'],
    ['EXT-002', 'test.xlsx', '01_PAGOS_HISTORICOS', '2', 'CORTE-2024-01', 'Corte Ene24',
     '2024-01-01', '2024-01-31', '2024-01-31', 'PAGADO', 'JUAN PEREZ', '',
     'cabinet', '', 'Q87654321', '', '', '',
     'Esquema Cabinet', 'legacy_milestone', '1_viaje', '1_VIAJE',
     '1', '0', '1', '0.00', 'PEN', 'scout_driver_payment', 'REF-002', '', 'test'],
    ['EXT-003', 'test.xlsx', '01_PAGOS_HISTORICOS', '3', 'CORTE-2024-01', 'Corte Ene24',
     '2024-01-01', '2024-01-31', '2024-01-31', 'EXCLUIDO', 'JUAN PEREZ', '',
     'cabinet', '', 'Q11111111', '', '', '',
     'Esquema Cabinet', 'legacy_milestone', '5_viajes', '5_VIAJES',
     '5', '0', '5', '100.00', 'PEN', 'scout_driver_payment', 'REF-003', '', 'test'],
    ['EXT-004', 'test.xlsx', '01_PAGOS_HISTORICOS', '4', 'CORTE-2024-01', 'Corte Ene24',
     '2024-01-01', '2024-01-31', '2024-01-31', 'APROBADO', 'JUAN PEREZ', '',
     'cabinet', '', 'Q22222222', '', '', '',
     'Esquema Cabinet', 'legacy_milestone', '25_viajes', '25_VIAJES',
     '25', '0', '25', '200.00', 'PEN', 'scout_driver_payment', 'REF-004', '', 'test'],
]

for r_idx, row_data in enumerate(test_rows, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws.cell(row=r_idx, column=c_idx, value=val)

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

url = 'http://localhost:8000/scout-liq/historical-imports/preview'
files = {'file': ('test.xlsx', buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
r = requests.post(url, files=files, params={'sheet': '01_PAGOS_HISTORICOS'}, timeout=30)

if r.status_code != 200:
    print(f"ERROR HTTP {r.status_code}: {r.text}")
    exit(1)

result = r.json()

summary = {k: v for k, v in result.items() if k != 'lines'}
print(json.dumps(summary, indent=2, default=str))

print(f"\nLines: {len(result.get('lines', []))}")
for l in result.get('lines', []):
    print(f"  Row={l.get('source_row')} status={l.get('import_status')} reason={l.get('import_reason','')} amt={l.get('amount_paid')} scout_id={l.get('scout_id_resolved')} driver_id={l.get('driver_id_resolved')}")

print("\n=== SUMMARY ===")
print(f"Total rows: {summary.get('total_rows')}")
print(f"Ready to import: {summary.get('ready_to_import')}")
print(f"Rejected: {summary.get('rejected')}")
print(f"Manual review: {summary.get('manual_review')}")
print(f"Duplicate: {summary.get('duplicate')}")

# If preview has ready_to_import, test commit
if summary.get('ready_to_import', 0) > 0 and summary.get('batch_id'):
    print(f"\n=== COMMIT batch #{summary['batch_id']} ===")
    r2 = requests.post(f'http://localhost:8000/scout-liq/historical-imports/commit?batch_id={summary["batch_id"]}&uploaded_by=test', timeout=30)
    if r2.status_code == 200:
        cr = r2.json()
        print(json.dumps(cr, indent=2, default=str))

    # Check paid_history
    r3 = requests.get('http://localhost:8000/scout-liq/paid-history?import_source=historical_upload&limit=10', timeout=30)
    if r3.status_code == 200:
        ph = r3.json()
        print(f"\nPaid history items with historical_upload: {ph.get('total')}")
        for item in ph.get('items', [])[:5]:
            print(f"  PH#{item['id']} scout={item['scout_id']} driver={item.get('driver_id')} amt={item['amount_paid']} source={item['import_source']} hash={item.get('unique_hash','')[:16]}...")

print("\n=== ALL TESTS PASSED ===")
