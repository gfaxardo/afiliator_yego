"""E2E test for Phase 4.6 attributions."""
import openpyxl, io, json, requests

# Start backend
import subprocess, time, os, signal, sys

# Get real data
r = requests.get('http://localhost:8000/scout-liq/scouts', timeout=10)
scouts = r.json()
scout_name = scouts[0]['scout_name'] if scouts else 'Test Scout'
scout_id = scouts[0]['id'] if scouts else None

# Get drivers with licenses
r2 = requests.get('http://localhost:8000/scout-liq/source/drivers?limit=20', timeout=10)
drivers_data = r2.json()
driver1 = None
driver2 = None
for d in drivers_data.get('drivers', []):
    if d.get('license') and d.get('driver_id'):
        if not driver1:
            driver1 = d
        elif not driver2:
            driver2 = d
            break

if not driver1 or not driver2:
    print("No drivers with licenses found")
    sys.exit(1)

lic1 = driver1['license']
did1 = driver1['driver_id']
lic2 = driver2['license']
did2 = driver2['driver_id']
print(f"Driver 1: {did1} / {lic1}")
print(f"Driver 2: {did2} / {lic2}")
print(f"Scout: {scout_name} (id={scout_id})")

# First, create an assignment for driver2 with a different scout to test conflicts
# Actually, let's create a test: assign driver2 to scout 1 (or the first scout)
# Then try to import driver2 with a different scout → should get conflict

# Create test XLSX
wb = openpyxl.Workbook()
wb.remove(wb.active)
ws = wb.create_sheet('06_ATRIBUCIONES_HISTORICAS')

headers = [
    'scout_name_raw', 'driver_license_raw', 'driver_name_raw', 'origin_raw',
    'payment_status_raw', 'payment_amount_raw', 'payment_rule_raw',
    'ok_1_viaje_raw', 'ok_5_viajes_raw', 'notes'
]
for col, h in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=h)

rows_data = [
    # Row 1: valid - scout resolved, license resolved → ready_to_import
    [scout_name, lic1, driver1.get('driver_nombre', 'Driver 1'), 'app',
     'PAGADO', '150.00', '5_viajes', 'NO', 'SI', 'Test row 1 - valid'],
    # Row 2: no payment - still valid attribution
    [scout_name, lic2, driver2.get('driver_nombre', 'Driver 2'), 'cabinet',
     'NO PAGADO', '0', '1_viaje', 'SI', 'NO', 'Test row 2 - no payment'],
    # Row 3: license not resolvable → manual_review
    [scout_name, 'ZZ999999', 'Unknown Driver', 'app',
     'PAGADO', '50.00', 'conexion', 'NO', 'NO', 'Test row 3 - bad license'],
    # Row 4: good driver but scout not resolved → manual_review
    ['SCOUT INEXISTENTE XYZ', lic1, driver1.get('driver_nombre', 'Driver 1'), 'app',
     'PAGADO', '100.00', '25_viajes', 'SI', 'NO', 'Test row 4 - bad scout'],
]

for r_idx, row_data in enumerate(rows_data, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws.cell(row=r_idx, column=c_idx, value=val)

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

# PREVIEW
files = {'file': ('test_attr.xlsx', buf, 'application/octet-stream')}
r = requests.post(
    'http://localhost:8000/scout-liq/attributions/preview?sheet=06_ATRIBUCIONES_HISTORICAS',
    files=files, timeout=30)

if r.status_code != 200:
    print(f"ERROR PREVIEW {r.status_code}: {r.text[:500]}")
    sys.exit(1)

result = r.json()
print(f"\n=== PREVIEW ===")
summary = {k: v for k, v in result.items() if k != 'lines'}
print(json.dumps(summary, indent=2, default=str))
for l in result.get('lines', []):
    print(f"  Row={l.get('source_row')} status={l.get('import_status')} "
          f"scout={l.get('scout_id_resolved')} driver={l.get('driver_id_resolved')} "
          f"reason={l.get('import_reason','')}")

# COMMIT
print(f"\n=== COMMIT ===")
buf2 = io.BytesIO()
wb.save(buf2)
buf2.seek(0)
files2 = {'file': ('test_attr.xlsx', buf2, 'application/octet-stream')}
r2 = requests.post(
    'http://localhost:8000/scout-liq/attributions/commit?sheet=06_ATRIBUCIONES_HISTORICAS',
    files=files2, timeout=30)

if r2.status_code == 200:
    cr = r2.json()
    print(f"Created: {cr.get('assignments_created')} Updated: {cr.get('assignments_updated')}")
    print(f"Attributions: {cr.get('historical_attributions_created')}")
    print(f"Manual review: {cr.get('manual_review')} Conflicts: {cr.get('conflicts')}")
    print(f"Rejected: {cr.get('rejected')}")
else:
    print(f"ERROR COMMIT {r2.status_code}: {r2.text[:500]}")

# Check attributions
print(f"\n=== ATTRIBUTIONS ===")
r3 = requests.get('http://localhost:8000/scout-liq/attributions?limit=10', timeout=10)
if r3.status_code == 200:
    data = r3.json()
    print(f"Total: {data['total']}")
    for item in data['items'][:5]:
        print(f"  #{item['id']} scout={item['scout_name_raw']} lic={item['driver_license_raw']} "
              f"did={item['driver_id_resolved']} status={item['import_status']} "
              f"asg={item['linked_assignment_id']}")

# Check assignments
print(f"\n=== ASSIGNMENTS (recent) ===")
r4 = requests.get(f'http://localhost:8000/scout-liq/assignments?status=active', timeout=10)
if r4.status_code == 200:
    data4 = r4.json()
    print(f"Active assignments: {len(data4)}")
    for a in data4[-5:]:
        print(f"  #{a['id']} driver={a['driver_id']} scout={a['scout_id']} by={a.get('assigned_by')}")

# Test duplicate detection - run preview again
buf3 = io.BytesIO()
wb.save(buf3)
buf3.seek(0)
files3 = {'file': ('test_attr.xlsx', buf3, 'application/octet-stream')}
r5 = requests.post(
    'http://localhost:8000/scout-liq/attributions/preview?sheet=06_ATRIBUCIONES_HISTORICAS',
    files=files3, timeout=30)
if r5.status_code == 200:
    data5 = r5.json()
    print(f"\n=== RE-PREVIEW (duplicate check) ===")
    for l in data5.get('lines', []):
        print(f"  Row={l.get('source_row')} status={l.get('import_status')} reason={l.get('import_reason','')}")

print("\n=== E2E COMPLETE ===")
