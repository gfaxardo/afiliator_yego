import csv
import io
import json
import hashlib
import time
import logging
import traceback
from typing import Optional, List, Any, Dict
from datetime import date, datetime

_logger = logging.getLogger("scout_liq")

from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Request
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text
from decimal import Decimal

from app.database import get_db
from app.config import settings
from app.models.scout_liq import (
    Scout, ConversionScheme, ConversionTier, DriverAssignment, CutoffRun,
    PaidHistory, HistoricalImportBatch, HistoricalImportLine,
    SchemeVersion, SchemeChangeLog, ManualPayment,
    SupervisorCommission, ScoutBonus, HistoricalAttribution,
)
from app.adapters.source_adapter import (
    get_source_drivers as adapter_get_drivers,
    get_source_driver_by_id as adapter_get_driver,
    get_unassigned_source_drivers as adapter_get_unassigned,
    get_source_diagnostic_summary as adapter_get_summary,
    get_quality_data_contract_status as adapter_get_quality_contract,
)
from app.services.cutoff_engine import (
    create_cutoff_run, calculate_cutoff, check_already_paid,
    get_cutoff_summary, get_cutoff_lines,
    review_cutoff, approve_cutoff, mark_cutoff_paid,
    export_cutoff_financial_csv,
    create_cutoff_from_cohort,
    create_sweep_cutoff,
)
from app.services.historical_import_service import (
    preview_historical_import,
    commit_historical_import,
    get_batch_errors_csv,
)
from app.services.scheme_import_service import (
    preview_scheme_import,
    commit_scheme_import,
    get_scheme_versions,
    get_scheme_change_log,
)
from app.services.manual_payment_service import (
    create_manual_payment_preview,
    approve_manual_payment,
    mark_manual_payment_paid,
    list_manual_payments,
    calculate_supervisor_commission,
    list_commissions,
    mark_commission_paid,
    create_bonus,
    approve_bonus,
    mark_bonus_paid,
    list_bonuses,
)
from app.services.attribution_import_service import (
    preview_attributions,
    commit_attributions,
    get_attributions,
    get_attribution_batch_errors_csv,
)
from app.services.sheet_validator import (
    SHEET_TYPE_MAP, classify_sheet, validate_sheet_for_endpoint,
    get_sheets_for_endpoint, get_sheet_type_label,
    log_preview_start, log_preview_done, log_preview_error,
    log_wrong_sheet, log_commit_start, log_commit_done,
)
from app.services.workbook_import_service import (
    workbook_preview, workbook_commit, _extract_xlsx_all_sheets,
)
from app.services.operation_service import (
    get_affiliations, get_affiliation_detail, get_operation_summary,
    get_operation_filters, export_affiliations_csv,
)
from app.services.canonical_operation_service import (
    get_canonical_operation_snapshot,
    get_operation_diagnostic,
)
from app.services.cohort_service import (
    get_iso_cohorts,
    get_cohort_diagnostic,
)
from app.services.payment_scheme_resolver import (
    resolve_payment_scheme_for_cohort,
)
from app.services.payment_scheme_admin_service import (
    list_payment_schemes,
    get_payment_scheme_detail,
    create_payment_scheme,
    create_payment_scheme_version,
    activate_payment_scheme_version,
    archive_payment_scheme_version,
    get_payment_schemes_history,
)
from app.services.manual_override_service import (
    create_manual_override,
    approve_manual_override,
    apply_manual_override,
    reject_manual_override,
    list_manual_overrides,
    get_driver_overrides,
)
from app.services.dashboard_service import (
    get_dashboard_overview, get_dashboard_by_scout, get_dashboard_by_week,
    get_dashboard_quality_funnel, get_dashboard_alerts, get_cutoff_trend,
)
from app.services.reconciliation_service import (
    export_reconciliation_csv,
    compare_upload,
)
from app.services.unified_load_service import (
    unified_preview,
    unified_apply,
    unified_preview_stream,
    unified_apply_stream,
    generate_preview_audit_csv,
    generate_apply_audit_csv,
    generate_full_audit_csv,
    generate_summary_csv,
    _get_preview,
    _parse_rows_from_csv,
    _parse_rows_from_xlsx,
)
from app.services.scout_liq_health_service import (
    get_health_summary,
    get_health_summary_lite,
    get_cohort_health,
    get_source_health,
    get_jobs_health,
)
from app.services.scout_liq_health_registry_service import (
    full_refresh_cycle,
    get_registry,
    get_events,
    compute_health_score,
    compute_health_score_lite,
)
from app.services.scout_liq_payment_flow_service import (
    create_payment_draft,
    recalculate_draft,
    review_cutoff_flow,
    approve_cutoff_flow,
    mark_cutoff_paid_flow,
    cancel_cutoff,
    undo_cutoff_status,
    get_cutoff_report,
    get_scout_payment_report,
    get_cohort_payment_report,
    export_cutoff_csv_full,
    export_cutoff_xlsx,
    get_payment_history_report,
    simulate_payment,
)
from app.schemas.scout_liq import (
    ScoutCreate,
    ScoutUpdate,
    ScoutResponse,
    ConversionSchemeCreate,
    ConversionSchemeResponse,
    ConversionTierResponse,
    HealthResponse,
    SourceDriverListResponse,
    SourceDriverResponse,
    AssignmentCreate,
    AssignmentResponse,
    AssignmentUploadResult,
    SourceDiagnosticSummary,
    ScoutUploadPreviewResult,
    ScoutUploadCommitResult,
    SchemeImportPreviewResult,
    SchemeImportCommitResult,
    HistoricalImportPreviewResult,
    ManualPaymentCreate,
    ManualPaymentApprove,
    ScoutBonusCreate,
    ScoutBonusApprove,
    CanonicalOperationSnapshotResponse,
    OperationDiagnosticResponse,
    CohortListResponse,
    CohortDiagnosticResponse,
    ResolvedPaymentScheme,
    PaymentSchemeListItem,
    PaymentSchemeDetail,
    CreatePaymentSchemeRequest,
    CreateVersionRequest,
    VersionCreatedResponse,
    VersionActivatedResponse,
    SchemeCreatedResponse,
    ManualOverrideResponse,
    CreateManualOverrideRequest,
    ReconciliationCompareResponse,
    UnifiedLoadPreviewResponse,
    UnifiedLoadApplyResponse,
)

router = APIRouter(prefix="/scout-liq", tags=["scout-liq"])


def _parse_snapshot(config_snapshot: str | None) -> dict:
    """Extrae campos del config_snapshot JSON de un cutoff_run."""
    if not config_snapshot:
        return {}
    try:
        return json.loads(config_snapshot)
    except (json.JSONDecodeError, TypeError):
        return {}


# ── Health ────────────────────────────────────────────────────────────────

@router.get("/health", response_model=HealthResponse)
def health():
    return HealthResponse(
        status="ok",
        environment=settings.ENVIRONMENT,
        source_table=settings.SOURCE_TABLE,
    )


# ── Source Diagnostic ─────────────────────────────────────────────────────

@router.get("/source/diagnostic")
def source_diagnostic(db: Session = Depends(get_db)):
    table = settings.SOURCE_TABLE
    results = {"source_table": table}

    try:
        row = db.execute(
            text(
                "SELECT column_name, data_type FROM information_schema.columns "
                "WHERE table_name = :tbl ORDER BY ordinal_position"
            ),
            {"tbl": table},
        ).fetchall()
        results["columns"] = [{"name": r[0], "type": r[1]} for r in row]
        col_names = [r[0] for r in row]
    except Exception as e:
        results["columns"] = f"ERROR: {e}"
        col_names = []

    try:
        results["total_rows"] = db.execute(
            text(f"SELECT COUNT(*) FROM {table}")
        ).scalar()
    except Exception as e:
        results["total_rows"] = f"ERROR: {e}"

    def safe_count(column_name: str) -> str:
        if column_name not in col_names:
            return f"COLUMN_NOT_FOUND: {column_name}"
        try:
            return db.execute(
                text(f"SELECT COUNT(*) FROM {table} WHERE {column_name} IS NULL")
            ).scalar()
        except Exception as e:
            return f"ERROR: {e}"

    results["null_driver_id"] = safe_count("driver_id")
    results["null_hire_date"] = safe_count("hire_date")
    results["null_origen"] = safe_count("origen")

    try:
        if "hire_date" in col_names:
            row = db.execute(
                text(f"SELECT MIN(hire_date), MAX(hire_date) FROM {table}")
            ).first()
            results["hire_date_min"] = str(row[0]) if row and row[0] else None
            results["hire_date_max"] = str(row[1]) if row and row[1] else None
        else:
            results["hire_date_min"] = "COLUMN_NOT_FOUND"
            results["hire_date_max"] = "COLUMN_NOT_FOUND"
    except Exception as e:
        results["hire_date_range"] = f"ERROR: {e}"

    try:
        if "origen" in col_names:
            row = db.execute(
                text(f"SELECT origen, COUNT(*) AS cnt FROM {table} "
                     "WHERE origen IS NOT NULL GROUP BY origen ORDER BY cnt DESC")
            ).fetchall()
            results["origen_distribution"] = [
                {"value": r[0], "count": r[1]} for r in row
            ]
        else:
            results["origen_distribution"] = "COLUMN_NOT_FOUND"
    except Exception as e:
        results["origen_distribution"] = f"ERROR: {e}"

    results["column_names_available"] = col_names
    results["column_mapping"] = {
        "driver_id": "driver_id" if "driver_id" in col_names else None,
        "hire_date": "hire_date" if "hire_date" in col_names else None,
        "hire_date_type": "VARCHAR (needs CAST to DATE)",
        "origin": "origen" if "origen" in col_names else None,
        "trips_7d_flag": "viajes_0_7" if "viajes_0_7" in col_names else None,
        "trips_14d_flag": "viajes_8_14" if "viajes_8_14" in col_names else None,
        "total_orders": "orders" if "orders" in col_names else None,
    }

    return results


# ── Source Diagnostic Summary (adapter) ───────────────────────────────────

@router.get("/source/summary", response_model=SourceDiagnosticSummary)
def source_summary(db: Session = Depends(get_db)):
    return adapter_get_summary(db)


# ── Source Drivers ────────────────────────────────────────────────────────

@router.get("/source/drivers", response_model=SourceDriverListResponse)
def list_source_drivers(
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    origin: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    return adapter_get_drivers(
        db,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        origin=origin,
        limit=limit,
        offset=offset,
    )


@router.get("/source/drivers/{driver_id}", response_model=SourceDriverResponse)
def get_source_driver(driver_id: str, db: Session = Depends(get_db)):
    driver = adapter_get_driver(db, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado en fuente")
    return driver


# ── Scouts (existing) ─────────────────────────────────────────────────────

@router.get("/scouts", response_model=List[ScoutResponse])
def list_scouts(
    status: Optional[str] = Query(None),
    scout_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Scout)
    if status:
        q = q.filter(Scout.status == status)
    if scout_type:
        q = q.filter(Scout.scout_type == scout_type)
    return q.order_by(Scout.created_at.desc()).all()


@router.post("/scouts", response_model=ScoutResponse, status_code=201)
def create_scout(data: ScoutCreate, db: Session = Depends(get_db)):
    scout = Scout(**data.model_dump())
    db.add(scout)
    db.commit()
    db.refresh(scout)
    return scout


@router.put("/scouts/{scout_id}", response_model=ScoutResponse)
def update_scout(scout_id: int, data: ScoutUpdate, db: Session = Depends(get_db)):
    scout = db.query(Scout).filter(Scout.id == scout_id).first()
    if not scout:
        raise HTTPException(status_code=404, detail="Scout no encontrado")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(scout, field, value)
    db.commit()
    db.refresh(scout)
    return scout


@router.get("/scouts/{scout_id}", response_model=ScoutResponse)
def get_scout(scout_id: int, db: Session = Depends(get_db)):
    scout = db.query(Scout).filter(Scout.id == scout_id).first()
    if not scout:
        raise HTTPException(status_code=404, detail="Scout no encontrado")
    return scout


# ── Schemes (existing) ────────────────────────────────────────────────────

@router.get("/schemes", response_model=List[ConversionSchemeResponse])
def list_schemes(db: Session = Depends(get_db)):
    return db.query(ConversionScheme).order_by(ConversionScheme.created_at.desc()).all()


@router.post("/schemes", response_model=ConversionSchemeResponse, status_code=201)
def create_scheme(data: ConversionSchemeCreate, db: Session = Depends(get_db)):
    scheme = ConversionScheme(**data.model_dump())
    db.add(scheme)
    db.commit()
    db.refresh(scheme)
    return scheme


@router.get("/schemes/{scheme_id}", response_model=ConversionSchemeResponse)
def get_scheme(scheme_id: int, db: Session = Depends(get_db)):
    scheme = db.query(ConversionScheme).filter(ConversionScheme.id == scheme_id).first()
    if not scheme:
        raise HTTPException(status_code=404, detail="Esquema no encontrado")
    return scheme


@router.get("/tiers", response_model=List[ConversionTierResponse])
def list_tiers(scheme_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    q = db.query(ConversionTier)
    if scheme_id:
        q = q.filter(ConversionTier.scheme_id == scheme_id)
    return q.order_by(ConversionTier.min_conversion_rate).all()


# ── Assignments (Fase 2) ──────────────────────────────────────────────────

@router.get("/assignments", response_model=List[AssignmentResponse])
def list_assignments(
    scout_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(DriverAssignment, Scout.scout_name).join(
        Scout, DriverAssignment.scout_id == Scout.id
    )
    if scout_id:
        q = q.filter(DriverAssignment.scout_id == scout_id)
    if status:
        q = q.filter(DriverAssignment.status == status)
    rows = q.order_by(DriverAssignment.created_at.desc()).all()
    return [
        AssignmentResponse(
            id=a.id,
            driver_id=a.driver_id,
            scout_id=a.scout_id,
            origin=a.origin,
            hire_date=str(a.hire_date) if a.hire_date else None,
            notes=a.notes,
            status=a.status,
            source_hire_date_raw=a.source_hire_date_raw,
            source_origin=a.source_origin,
            assigned_by=a.assigned_by,
            assigned_at=str(a.assigned_at) if a.assigned_at else None,
            created_at=str(a.created_at) if a.created_at else None,
            updated_at=str(a.updated_at) if a.updated_at else None,
            scout_name=name,
        )
        for a, name in rows
    ]


@router.post("/assignments", status_code=201)
def create_assignment(data: AssignmentCreate, db: Session = Depends(get_db)):
    scout = db.query(Scout).filter(
        Scout.id == data.scout_id, Scout.status == "active"
    ).first()
    if not scout:
        raise HTTPException(status_code=400, detail="Scout no encontrado o inactivo")

    source_driver = adapter_get_driver(db, data.driver_id)
    if not source_driver:
        raise HTTPException(status_code=400, detail="Driver no existe en la fuente")

    origin = data.origin or source_driver.get("origin")

    existing = db.query(DriverAssignment).filter(
        DriverAssignment.driver_id == data.driver_id,
        DriverAssignment.status == "active",
    ).first()
    if existing:
        return {
            "status": "skipped_duplicate",
            "detail": f"Driver {data.driver_id} ya asignado activamente al scout {existing.scout_id}",
            "existing_assignment_id": existing.id,
        }

    warning = None
    if not source_driver.get("hire_date_parsed"):
        warning = "Driver sin hire_date valida en fuente"

    assignment = DriverAssignment(
        driver_id=data.driver_id,
        scout_id=data.scout_id,
        origin=origin,
        notes=data.notes,
        status="active",
        source_hire_date_raw=source_driver.get("hire_date_raw"),
        source_origin=origin,
        assigned_by="manual",
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)

    resp = {
        "status": "created",
        "id": assignment.id,
        "driver_id": assignment.driver_id,
        "scout_id": assignment.scout_id,
        "origin": assignment.origin,
        "source_hire_date_raw": assignment.source_hire_date_raw,
    }
    if warning:
        resp["warning"] = warning
    return resp


@router.put("/assignments/{assignment_id}")
def update_assignment(
    assignment_id: int,
    notes: Optional[str] = None,
    origin: Optional[str] = None,
    db: Session = Depends(get_db),
):
    assignment = db.query(DriverAssignment).filter(
        DriverAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Asignacion no encontrada")
    if notes is not None:
        assignment.notes = notes
    if origin is not None:
        assignment.origin = origin
        assignment.source_origin = origin
    db.commit()
    return {"status": "updated", "id": assignment.id}


@router.delete("/assignments/{assignment_id}")
def deactivate_assignment(assignment_id: int, db: Session = Depends(get_db)):
    assignment = db.query(DriverAssignment).filter(
        DriverAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Asignacion no encontrada")
    assignment.status = "inactive"
    db.commit()
    return {"status": "inactive", "id": assignment.id}


# ── Unassigned Drivers ────────────────────────────────────────────────────

@router.get("/assignments/unassigned-drivers", response_model=SourceDriverListResponse)
def unassigned_drivers(
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    return adapter_get_unassigned(db, limit=limit, offset=offset)


# ── Upload CSV/XLSX ──────────────────────────────────────────────────────

@router.post("/assignments/upload")
def upload_assignments(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()

    if filename.endswith(".csv"):
        return _process_csv(db, content)
    elif filename.endswith(".xlsx"):
        return _process_xlsx(db, content)
    else:
        raise HTTPException(
            status_code=400,
            detail="Formato no soportado. Usar .csv o .xlsx",
        )


def _process_csv(db: Session, content: bytes) -> dict:
    text_content = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text_content = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text_content is None:
        raise HTTPException(status_code=400, detail="No se pudo decodificar el archivo")

    reader = csv.DictReader(io.StringIO(text_content))
    return _process_rows(db, list(reader))


def _process_xlsx(db: Session, content: bytes) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Ejecuta: pip install openpyxl",
        )

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() if h else "" for h in next(rows_iter, [])]
    reader = []
    for row in rows_iter:
        reader.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
    wb.close()
    return _process_rows(db, reader)


def _process_rows(db: Session, rows: list) -> dict:
    result = {
        "total_rows": len(rows),
        "created": 0,
        "skipped_duplicates": 0,
        "invalid_driver": 0,
        "invalid_scout": 0,
        "missing_hire_date_warnings": 0,
        "errors": [],
        "warnings": [],
    }

    for i, row in enumerate(rows):
        line = i + 2
        driver_id = (row.get("driver_id") or "").strip()
        scout_id_raw = (row.get("scout_id") or "").strip()
        scout_name = (row.get("scout_name") or "").strip()
        origin = (row.get("origin") or "").strip() or None
        notes = (row.get("notes") or "").strip() or None

        if not driver_id:
            result["errors"].append(f"Fila {line}: driver_id vacio")
            continue

        scout_id = None
        if scout_id_raw:
            try:
                scout_id = int(scout_id_raw)
            except ValueError:
                result["errors"].append(f"Fila {line}: scout_id invalido")
                continue
        elif scout_name:
            scout = db.query(Scout).filter(
                Scout.scout_name.ilike(scout_name)
            ).first()
            if scout:
                scout_id = scout.id
            else:
                result["invalid_scout"] += 1
                result["errors"].append(f"Fila {line}: scout '{scout_name}' no encontrado")
                continue
        else:
            result["errors"].append(f"Fila {line}: se requiere scout_id o scout_name")
            continue

        source_driver = adapter_get_driver(db, driver_id)
        if not source_driver:
            result["invalid_driver"] += 1
            result["errors"].append(f"Fila {line}: driver_id {driver_id} no existe en fuente")
            continue

        existing = db.query(DriverAssignment).filter(
            DriverAssignment.driver_id == driver_id,
            DriverAssignment.status == "active",
        ).first()
        if existing:
            result["skipped_duplicates"] += 1
            continue

        if not source_driver.get("hire_date_parsed"):
            result["missing_hire_date_warnings"] += 1

        final_origin = origin or source_driver.get("origin")

        assignment = DriverAssignment(
            driver_id=driver_id,
            scout_id=scout_id,
            origin=final_origin,
            notes=notes,
            status="active",
            source_hire_date_raw=source_driver.get("hire_date_raw"),
            source_origin=final_origin,
            assigned_by="csv_upload",
        )
        db.add(assignment)
        result["created"] += 1

    db.commit()
    return result


# ── Fase 3: Quality Contract ──────────────────────────────────────────────

@router.get("/source/quality-contract")
def quality_contract(db: Session = Depends(get_db)):
    return adapter_get_quality_contract(db)


# ── Fase 3: Cutoffs ──────────────────────────────────────────────────────

@router.post("/cutoffs")
def create_cutoff(
    cutoff_name: str,
    hire_date_from: date,
    hire_date_to: date,
    scheme_id: int,
    origin_filter: Optional[str] = None,
    country_filter: Optional[str] = None,
    city_filter: Optional[str] = None,
    scout_type_filter: Optional[str] = None,
    created_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        run = create_cutoff_run(
            db, cutoff_name, hire_date_from, hire_date_to, scheme_id,
            origin_filter, country_filter, city_filter, scout_type_filter, created_by,
        )
        result = calculate_cutoff(db, run.id)
        return {
            "cutoff_run_id": run.id,
            "cutoff_name": run.cutoff_name,
            "status": run.status,
            "calculation": result,
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/cutoffs/from-cohort")
def create_cutoff_from_cohort_endpoint(
    cohort_iso_week: str,
    scheme_type: Optional[str] = Query(None, description="cabinet | fleet | custom (preferido)"),
    scheme_id: Optional[int] = Query(None, description="Legacy scheme_id (deprecado)"),
    origin_filter: Optional[str] = None,
    scout_type_filter: Optional[str] = None,
    created_by: Optional[str] = None,
    force_override: bool = False,
    db: Session = Depends(get_db),
):
    """
    Crea un cutoff desde una cohorte ISO madura, resolviendo automaticamente
    la version de esquema de pago aplicable.

    - cohort_iso_week: ej. '2026-W18'
    - scheme_type: cabinet | fleet | custom (usa el resolver versionado)
    - scheme_id: deprecado, solo para compatibilidad legacy
    - force_override: si ya existe un corte locked, forzar recalculo
    """
    try:
        result = create_cutoff_from_cohort(
            db,
            cohort_iso_week=cohort_iso_week,
            scheme_type=scheme_type or "",
            scheme_id=scheme_id,
            origin_filter=origin_filter,
            scout_type_filter=scout_type_filter,
            created_by=created_by,
            force_override=force_override,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/cutoffs/sweep")
def create_sweep_cutoff_endpoint(
    scheme_type: str = Query("cabinet", description="cabinet | fleet | custom"),
    origin_filter: Optional[str] = None,
    scout_type_filter: Optional[str] = None,
    created_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Crea un cutoff en modo PAYABLE_SWEEP (barrido pagable).

    Evalua TODOS los drivers activos que cumplen la regla del scheme vigente HOY
    y que NO tienen paid_history bloqueante. Sin filtro de hire_date.
    """
    try:
        result = create_sweep_cutoff(
            db,
            scheme_type=scheme_type,
            origin_filter=origin_filter,
            scout_type_filter=scout_type_filter,
            created_by=created_by,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/cutoffs")
def list_cutoffs(db: Session = Depends(get_db)):
    runs = db.query(CutoffRun).order_by(CutoffRun.created_at.desc()).all()
    result = []
    for r in runs:
        cfg = _parse_snapshot(r.config_snapshot)
        result.append({
            "id": r.id,
            "cutoff_name": r.cutoff_name,
            "hire_date_from": str(r.hire_date_from) if r.hire_date_from else None,
            "hire_date_to": str(r.hire_date_to) if r.hire_date_to else None,
            "origin_filter": r.origin_filter,
            "status": r.status,
            "quality_data_contract_status": r.quality_data_contract_status,
            "conversion_metric_status": r.conversion_metric_status,
            "cohort_iso_week": r.cohort_iso_week,
            "cohort_from": str(r.cohort_from) if r.cohort_from else None,
            "cohort_to": str(r.cohort_to) if r.cohort_to else None,
            "maturity_days": r.maturity_days,
            "ready_to_liquidate": r.ready_to_liquidate,
            "snapshot_locked_at": str(r.snapshot_locked_at) if r.snapshot_locked_at else None,
            "scheme_name": cfg.get("scheme_name"),
            "scheme_type": cfg.get("scheme_type"),
            "version_name": cfg.get("version_name"),
            "min_activated": cfg.get("min_activated"),
            "activation_rule": cfg.get("activation_rule"),
            "quality_rule": cfg.get("quality_rule"),
            "formula_type": cfg.get("formula_type"),
            "created_at": str(r.created_at) if r.created_at else None,
            "notes": r.notes,
            "cancelled_at": str(r.cancelled_at) if r.cancelled_at else None,
            "cancelled_reason": r.cancelled_reason,
            "config_snapshot": cfg,
        })
    return result


@router.get("/cutoffs/{cutoff_id}")
def get_cutoff(cutoff_id: int, db: Session = Depends(get_db)):
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Cutoff no encontrado")
    cfg = _parse_snapshot(run.config_snapshot)
    return {
        "id": run.id,
        "cutoff_name": run.cutoff_name,
        "hire_date_from": str(run.hire_date_from) if run.hire_date_from else None,
        "hire_date_to": str(run.hire_date_to) if run.hire_date_to else None,
        "origin_filter": run.origin_filter,
        "status": run.status,
        "quality_data_contract_status": run.quality_data_contract_status,
        "conversion_metric_status": run.conversion_metric_status,
        "total_source_drivers_count": run.total_source_drivers_count,
        "excluded_invalid_hire_date_count": run.excluded_invalid_hire_date_count,
        "excluded_missing_trip_counts_count": run.excluded_missing_trip_counts_count,
        "cohort_iso_week": run.cohort_iso_week,
        "cohort_from": str(run.cohort_from) if run.cohort_from else None,
        "cohort_to": str(run.cohort_to) if run.cohort_to else None,
        "maturity_days": run.maturity_days,
        "maturity_completed_at": str(run.maturity_completed_at) if run.maturity_completed_at else None,
        "ready_to_liquidate": run.ready_to_liquidate,
        "snapshot_locked_at": str(run.snapshot_locked_at) if run.snapshot_locked_at else None,
        "scheme_name": cfg.get("scheme_name"),
        "scheme_type": cfg.get("scheme_type"),
        "version_name": cfg.get("version_name"),
        "min_activated": cfg.get("min_activated"),
        "activation_rule": cfg.get("activation_rule"),
        "quality_rule": cfg.get("quality_rule"),
        "formula_type": cfg.get("formula_type"),
        "currency": cfg.get("currency"),
        "tiers": cfg.get("tiers", []),
        "created_at": str(run.created_at) if run.created_at else None,
        "approved_at": str(run.approved_at) if run.approved_at else None,
        "paid_at": str(run.paid_at) if run.paid_at else None,
        "cutoff_mode": run.cutoff_mode or "COHORT",
    }


@router.get("/cutoffs/{cutoff_id}/summary")
def cutoff_summary(cutoff_id: int, db: Session = Depends(get_db)):
    return get_cutoff_summary(db, cutoff_id)


    return get_cutoff_lines(db, cutoff_id, scout_id)


@router.get("/cutoffs/{cutoff_id}/lines")
def cutoff_lines(
    cutoff_id: int,
    scout_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    return get_cutoff_lines(db, cutoff_id, scout_id)


@router.get("/cutoffs/{cutoff_id}/export-financial.csv")
def export_cutoff_financial(
    cutoff_id: int,
    db: Session = Depends(get_db),
):
    """
    Export financiero auditable para un cutoff en formato CSV.
    Usa config_snapshot congelado, NO el resolver dinamico.
    """
    try:
        csv_content = export_cutoff_financial_csv(db, cutoff_id)
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=cutoff_{cutoff_id}_financial.csv"},
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
def cutoff_lines(
    cutoff_id: int,
    scout_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    return get_cutoff_lines(db, cutoff_id, scout_id)


@router.post("/cutoffs/{cutoff_id}/recalculate")
def recalculate_cutoff(cutoff_id: int, db: Session = Depends(get_db)):
    try:
        run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_id).first()
        if not run:
            raise HTTPException(status_code=404, detail="Cutoff no encontrado")
        if run.status not in ("draft", "calculated"):
            raise HTTPException(status_code=400, detail=f"No se puede recalcular en estado '{run.status}'")
        run.status = "draft"
        db.commit()
        result = calculate_cutoff(db, cutoff_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/cutoffs/{cutoff_id}/review")
def review_cutoff_endpoint(cutoff_id: int, db: Session = Depends(get_db)):
    try:
        return review_cutoff(db, cutoff_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/cutoffs/{cutoff_id}/approve")
def approve_cutoff_endpoint(
    cutoff_id: int,
    approved_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        return approve_cutoff(db, cutoff_id, approved_by)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/cutoffs/{cutoff_id}/mark-paid")
def mark_paid_endpoint(cutoff_id: int, db: Session = Depends(get_db)):
    try:
        return mark_cutoff_paid(db, cutoff_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/cutoffs/{cutoff_id}/export.csv")
def export_cutoff_csv(cutoff_id: int, db: Session = Depends(get_db)):
    summaries = get_cutoff_summary(db, cutoff_id)
    lines_rows = get_cutoff_lines(db, cutoff_id)

    csv_output = io.StringIO()
    csv_output.write("# Resumen por Scout\r\n")
    csv_output.write("scout_id,scout_name,afiliados,total_afiliaciones,1plus_0_7,5plus_0_7,1plus_8_14,5plus_0_14,no_convertidos,conversion,tramo,pago_por_convertido,calculado,estado\r\n")
    for s in summaries:
        csv_output.write(
            f"{s['scout_id']},{s['scout_name']},{s['total_affiliations']},"
            f"{s['drivers_1plus_0_7']},{s['drivers_5plus_0_7']},{s['drivers_1plus_8_14']},"
            f"{s['drivers_5plus_0_14']},{s['not_converted']},{s['conversion_rate']:.4f},"
            f"{s['tier_reached'] or ''},{s['payment_per_converted_driver']:.2f},"
            f"{s['amount_calculated']:.2f},{s['status']}\r\n"
        )
    csv_output.write("\r\n# Lineas por Driver\r\n")
    csv_output.write("driver_id,scout_id,hire_date,origen,trips_0_7,trips_8_14,trips_0_14,orders,hito,estado,motivo\r\n")
    for l in lines_rows:
        csv_output.write(
            f"{l['driver_id']},{l['scout_id']},{l['hire_date'] or ''},{l['origin'] or ''},"
            f"{l['trips_0_7_count'] or 0},{l['trips_8_14_count'] or 0},{l['trips_0_14_count'] or 0},"
            f"{l['total_orders'] or ''},{'Si' if l['is_converted_5trips_7d'] else 'No'},"
            f"{l['line_status']},{l['blocked_reason'] or ''}\r\n"
        )
    return Response(content=csv_output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=cutoff_{cutoff_id}.csv"})


# ═══════════════════════════════════════════════════════════════════════════
# PAYMENT FLOW: Flujo completo de pagos
# ═══════════════════════════════════════════════════════════════════════════


@router.post("/payments/drafts")
def create_payment_draft_endpoint(
    hire_date_from: date,
    hire_date_to: date,
    scheme_id: int,
    origin: Optional[str] = None,
    country: Optional[str] = None,
    city: Optional[str] = None,
    scout_type: Optional[str] = None,
    notes: Optional[str] = None,
    created_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        return create_payment_draft(
            db,
            hire_date_from=hire_date_from,
            hire_date_to=hire_date_to,
            scheme_id=scheme_id,
            origin=origin,
            country=country,
            city=city,
            scout_type=scout_type,
            notes=notes,
            created_by=created_by,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════════
# SIMULACION DE PAGO (sin persistir) — definida ANTES de rutas con {cutoff_run_id}
# ═══════════════════════════════════════════════════════════════════════════

class SimulationTier(BaseModel):
    min_conversion_rate: float
    payout_amount: float


class SimulationRequest(BaseModel):
    hire_date_from: date
    hire_date_to: date
    scheme_id: int
    origin: Optional[str] = None
    scout_type: Optional[str] = None
    override_tiers: Optional[List[SimulationTier]] = None
    override_min_affiliations: Optional[int] = None
    override_quality_rule: Optional[str] = None
    override_pays_on_rule: Optional[str] = None
    override_payout_formula_type: Optional[str] = None
    override_activation_rule: Optional[str] = None


@router.post("/payments/simulate")
def simulate_payment_endpoint(body: SimulationRequest, db: Session = Depends(get_db)):
    try:
        override_tiers = None
        if body.override_tiers:
            override_tiers = [
                {
                    "min_conversion_rate": t.min_conversion_rate,
                    "payout_amount": t.payout_amount,
                    "payment_per_converted_driver": t.payout_amount,
                    "currency": "PEN",
                    "sort_order": 0,
                }
                for t in body.override_tiers
            ]
        return simulate_payment(
            db,
            hire_date_from=body.hire_date_from,
            hire_date_to=body.hire_date_to,
            scheme_id=body.scheme_id,
            origin=body.origin,
            scout_type=body.scout_type,
            override_tiers=override_tiers,
            override_min_affiliations=body.override_min_affiliations,
            override_quality_rule=body.override_quality_rule,
            override_pays_on_rule=body.override_pays_on_rule,
            override_payout_formula_type=body.override_payout_formula_type,
            override_activation_rule=body.override_activation_rule,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payments/{cutoff_run_id}/recalculate")
def recalculate_draft_endpoint(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        return recalculate_draft(db, cutoff_run_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payments/{cutoff_run_id}/review")
def review_cutoff_flow_endpoint(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        return review_cutoff_flow(db, cutoff_run_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payments/{cutoff_run_id}/approve")
def approve_cutoff_flow_endpoint(
    cutoff_run_id: int,
    approved_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        return approve_cutoff_flow(db, cutoff_run_id, approved_by)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payments/{cutoff_run_id}/mark-paid")
def mark_paid_flow_endpoint(
    cutoff_run_id: int,
    paid_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        return mark_cutoff_paid_flow(db, cutoff_run_id, paid_by)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payments/{cutoff_run_id}/cancel")
def cancel_cutoff_endpoint(
    cutoff_run_id: int,
    reason: str = Query(..., description="Motivo de cancelacion"),
    db: Session = Depends(get_db),
):
    try:
        return cancel_cutoff(db, cutoff_run_id, reason)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payments/{cutoff_run_id}/undo-status")
def undo_cutoff_status_endpoint(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        return undo_cutoff_status(db, cutoff_run_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/payments/{cutoff_run_id}")
def get_payment_detail(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        return get_cutoff_report(db, cutoff_run_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/payments/{cutoff_run_id}/report")
def get_payment_report(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        return get_cutoff_report(db, cutoff_run_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/payments/{cutoff_run_id}/export.csv")
def export_payment_csv(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        csv_content = export_cutoff_csv_full(db, cutoff_run_id)
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=payment_{cutoff_run_id}.csv"},
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/payments/{cutoff_run_id}/export.xlsx")
def export_payment_xlsx(cutoff_run_id: int, db: Session = Depends(get_db)):
    try:
        xlsx_content = export_cutoff_xlsx(db, cutoff_run_id)
        return Response(
            content=xlsx_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=payment_{cutoff_run_id}.xlsx"},
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/reports/scout/{scout_id}")
def scout_payment_report_endpoint(
    scout_id: int,
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    try:
        return get_scout_payment_report(db, scout_id, date_from, date_to)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/reports/cohort/{cohort_key}")
def cohort_payment_report_endpoint(cohort_key: str, db: Session = Depends(get_db)):
    try:
        return get_cohort_payment_report(db, cohort_key)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/reports/payment-history")
def payment_history_report_endpoint(
    scout_id: Optional[int] = Query(None),
    driver_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    return get_payment_history_report(db, scout_id, driver_id, date_from, date_to, limit, offset)


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Paid History
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/paid-history")
def get_paid_history(
    cutoff_run_id: Optional[int] = Query(None),
    scout_id: Optional[int] = Query(None),
    supervisor_id: Optional[int] = Query(None),
    driver_license_raw: Optional[str] = Query(None),
    payment_component: Optional[str] = Query(None),
    import_source: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    q = db.query(PaidHistory)
    if cutoff_run_id:
        q = q.filter(PaidHistory.cutoff_run_id == cutoff_run_id)
    if scout_id:
        q = q.filter(PaidHistory.scout_id == scout_id)
    if supervisor_id:
        q = q.filter(PaidHistory.supervisor_id == supervisor_id)
    if driver_license_raw:
        q = q.filter(PaidHistory.driver_license_raw.ilike(f"%{driver_license_raw}%"))
    if payment_component:
        q = q.filter(PaidHistory.payment_component == payment_component)
    if import_source:
        q = q.filter(PaidHistory.import_source == import_source)
    total = q.count()
    rows = q.order_by(PaidHistory.created_at.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r.id,
                "cutoff_run_id": r.cutoff_run_id,
                "scout_id": r.scout_id,
                "driver_id": r.driver_id,
                "driver_license_raw": r.driver_license_raw,
                "scout_name_raw": r.scout_name_raw,
                "supervisor_id": r.supervisor_id,
                "payment_scheme_id": r.payment_scheme_id,
                "payment_scheme_name": r.payment_scheme_name,
                "payment_scheme_type": r.payment_scheme_type,
                "payment_rule": r.payment_rule,
                "amount_paid": float(r.amount_paid) if r.amount_paid else 0,
                "currency": r.currency,
                "paid_at": str(r.paid_at) if r.paid_at else None,
                "import_source": r.import_source,
                "payment_component": r.payment_component,
                "milestone": r.milestone,
                "cutoff_external_id": r.cutoff_external_id,
                "cutoff_window_from": str(r.cutoff_window_from) if r.cutoff_window_from else None,
                "cutoff_window_to": str(r.cutoff_window_to) if r.cutoff_window_to else None,
                "reason": r.reason,
                "status": r.status,
                "unique_hash": r.unique_hash,
                "created_at": str(r.created_at) if r.created_at else None,
            }
            for r in rows
        ],
    }


@router.get("/paid-history/{payment_id}")
def get_paid_history_item(payment_id: int, db: Session = Depends(get_db)):
    r = db.query(PaidHistory).filter(PaidHistory.id == payment_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    return {
        "id": r.id,
        "cutoff_run_id": r.cutoff_run_id,
        "scout_id": r.scout_id,
        "driver_id": r.driver_id,
        "driver_license_raw": r.driver_license_raw,
        "scout_name_raw": r.scout_name_raw,
        "supervisor_id": r.supervisor_id,
        "payment_scheme_id": r.payment_scheme_id,
        "payment_scheme_name": r.payment_scheme_name,
        "payment_scheme_type": r.payment_scheme_type,
        "payment_rule": r.payment_rule,
        "amount_paid": float(r.amount_paid) if r.amount_paid else 0,
        "currency": r.currency,
        "paid_at": str(r.paid_at) if r.paid_at else None,
        "import_source": r.import_source,
        "import_batch_id": r.import_batch_id,
        "source_file": r.source_file,
        "source_sheet": r.source_sheet,
        "source_row": r.source_row,
        "payment_component": r.payment_component,
        "milestone": r.milestone,
        "cutoff_external_id": r.cutoff_external_id,
        "cutoff_window_from": str(r.cutoff_window_from) if r.cutoff_window_from else None,
        "cutoff_window_to": str(r.cutoff_window_to) if r.cutoff_window_to else None,
        "unique_hash": r.unique_hash,
        "paid_by": r.paid_by,
        "reason": r.reason,
        "status": r.status,
        "resolution_status": r.resolution_status,
        "blocks_future_payment": r.blocks_future_payment,
        "financial_record_status": r.financial_record_status,
        "created_at": str(r.created_at) if r.created_at else None,
        "updated_at": str(r.updated_at) if r.updated_at else None,
    }


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Historical Import
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/historical-imports/preview")
def historical_import_preview(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)
    used_sheet = detected_sheet or "unknown"

    # Validate sheet type
    is_valid, import_type, err_msg = validate_sheet_for_endpoint(used_sheet, "historical-imports")
    if not is_valid:
        log_wrong_sheet("historical-imports", used_sheet)
        raise HTTPException(
            status_code=400,
            detail={
                "error": "wrong_sheet_for_import",
                "sheet": used_sheet,
                "sheet_type": import_type,
                "message": err_msg,
                "valid_sheets": get_sheets_for_endpoint("historical-imports"),
            },
        )

    t0 = time.time()
    log_preview_start(import_type, filename, used_sheet, len(rows))

    try:
        result = preview_historical_import(db, rows, filename, used_sheet)

        batch = HistoricalImportBatch(
            source_file=filename,
            status="previewing",
            total_rows=result["total_rows"],
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)

        _save_preview_lines(db, batch.id, result["lines"])
        result["batch_id"] = batch.id

        elapsed = (time.time() - t0) * 1000
        log_preview_done(
            import_type, batch_id=batch.id,
            total_rows=result["total_rows"],
            ready=result["ready_to_import"],
            review=result["manual_review"],
            rejected=result["rejected"],
            duplicate=result["duplicate"],
            amount_ready=float(result.get("amount_ready", 0)),
            elapsed_ms=elapsed,
            top_errors=result.get("errors_by_type"),
        )
        return result
    except Exception as e:
        elapsed = (time.time() - t0) * 1000
        log_preview_error(import_type, used_sheet, str(e))
        raise


@router.post("/historical-imports/commit")
def historical_import_commit(
    batch_id: int,
    uploaded_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    batch = db.query(HistoricalImportBatch).filter(HistoricalImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch no encontrado")

    batch.uploaded_by = uploaded_by
    db.commit()

    lines = db.query(HistoricalImportLine).filter(
        HistoricalImportLine.batch_id == batch_id
    ).all()

    if not lines:
        raise HTTPException(status_code=400, detail="No hay lineas en el preview. Ejecuta preview primero.")

    preview_lines = []
    for l in lines:
        preview_lines.append({
            "source_sheet": l.source_sheet,
            "source_row": l.source_row,
            "corte_id_raw": l.corte_id_raw,
            "scout_name_raw": l.scout_name_raw,
            "scout_id_resolved": l.scout_id_resolved,
            "driver_license_raw": l.driver_license_raw,
            "driver_id_resolved": l.driver_id_resolved,
            "driver_name_raw": l.driver_name_raw,
            "amount_paid_raw": l.amount_paid_raw,
            "amount_paid": float(l.amount_paid) if l.amount_paid else None,
            "currency": (l.currency or "PEN") if hasattr(l, 'currency') else "PEN",
            "import_status": l.import_status,
            "import_reason": l.import_reason,
            "unique_hash": l.unique_hash,
            "payment_rule_raw": l.payment_rule_raw,
            "milestone_raw": l.milestone_raw,
            "fecha_pago_raw": l.fecha_pago_raw,
            "supervisor_id_resolved": l.supervisor_id_resolved,
            "supervisor_raw": l.supervisor_raw,
            "scout_type_raw": l.scout_type_raw,
            "origin_raw": l.origin_raw,
            "hire_date_raw": l.hire_date_raw,
            "estado_pago_raw": l.estado_pago_raw,
            "payment_scheme_raw": l.payment_scheme_raw,
            "payment_scheme_type_raw": getattr(l, 'payment_scheme_type_raw', None),
            "payment_component": "scout_driver_payment",
            "notes": getattr(l, 'import_reason', None),
            # Dual-layer fields (critical for commit)
            "attribution_status": l.attribution_status,
            "attribution_reason": l.attribution_reason,
            "payment_financial_status": l.payment_financial_status,
            "payment_financial_reason": l.payment_financial_reason,
            "payment_blocking_status": l.payment_blocking_status,
            "payment_blocking_reason": l.payment_blocking_reason,
            "blocks_future_payment": l.blocks_future_payment,
            "final_status": l.final_status,
        })

    preview = {"lines": preview_lines}
    try:
        result = commit_historical_import(db, batch_id, preview)
        return result
    except Exception as e:
        import traceback, logging
        logging.getLogger("scout_liq").error(f"[COMMIT ERROR] batch_id={batch_id}: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Commit error: {str(e)}")


# ── Internal helper: save preview lines into batch ──

def _save_preview_lines(db: Session, batch_id: int, lines_data: List[dict]):
    for ld in lines_data:
        line = HistoricalImportLine(
            batch_id=batch_id,
            source_sheet=ld.get("source_sheet"),
            source_row=ld.get("source_row"),
            corte_id_raw=ld.get("corte_id_raw"),
            scout_name_raw=ld.get("scout_name_raw"),
            scout_id_resolved=ld.get("scout_id_resolved"),
            driver_license_raw=ld.get("driver_license_raw"),
            driver_id_resolved=ld.get("driver_id_resolved"),
            driver_name_raw=ld.get("driver_name_raw"),
            amount_paid_raw=ld.get("amount_paid_raw"),
            amount_paid=Decimal(str(ld.get("amount_paid", 0))) if ld.get("amount_paid") else None,
            import_status=ld.get("final_status") or ld.get("import_status"),
            import_reason=ld.get("attribution_reason") or ld.get("payment_reason") or ld.get("import_reason"),
            attribution_status=ld.get("attribution_status"),
            attribution_reason=ld.get("attribution_reason"),
            payment_status=ld.get("payment_status"),
            payment_reason=ld.get("payment_reason"),
            payment_financial_status=ld.get("payment_financial_status"),
            payment_financial_reason=ld.get("payment_financial_reason"),
            payment_blocking_status=ld.get("payment_blocking_status"),
            payment_blocking_reason=ld.get("payment_blocking_reason"),
            blocks_future_payment=ld.get("blocks_future_payment"),
            final_status=ld.get("final_status"),
            unique_hash=ld.get("unique_hash"),
            payment_rule_raw=ld.get("payment_rule_raw"),
            milestone_raw=ld.get("milestone_raw"),
            supervisor_id_resolved=ld.get("supervisor_id_resolved"),
            fecha_pago_raw=ld.get("fecha_pago_raw"),
        )
        db.add(line)
    db.commit()


@router.get("/historical-imports")
def list_historical_imports(db: Session = Depends(get_db)):
    batches = db.query(HistoricalImportBatch).order_by(
        HistoricalImportBatch.created_at.desc()
    ).all()
    return [
        {
            "id": b.id,
            "upload_batch_id": b.upload_batch_id,
            "source_file": b.source_file,
            "uploaded_by": b.uploaded_by,
            "status": b.status,
            "total_rows": b.total_rows,
            "imported_count": b.imported_count,
            "rejected_count": b.rejected_count,
            "manual_review_count": b.manual_review_count,
            "duplicate_count": b.duplicate_count,
            "amount_imported": float(b.amount_imported) if b.amount_imported else 0,
            "notes": b.notes,
            "created_at": str(b.created_at) if b.created_at else None,
        }
        for b in batches
    ]


@router.get("/historical-imports/{batch_id}")
def get_historical_import(batch_id: int, db: Session = Depends(get_db)):
    b = db.query(HistoricalImportBatch).filter(HistoricalImportBatch.id == batch_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Batch no encontrado")
    return {
        "id": b.id,
        "upload_batch_id": b.upload_batch_id,
        "source_file": b.source_file,
        "uploaded_by": b.uploaded_by,
        "status": b.status,
        "total_rows": b.total_rows,
        "imported_count": b.imported_count,
        "rejected_count": b.rejected_count,
        "manual_review_count": b.manual_review_count,
        "duplicate_count": b.duplicate_count,
        "amount_imported": float(b.amount_imported) if b.amount_imported else 0,
        "notes": b.notes,
        "created_at": str(b.created_at) if b.created_at else None,
    }


@router.get("/historical-imports/{batch_id}/lines")
def get_historical_import_lines(
    batch_id: int,
    import_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(HistoricalImportLine).filter(HistoricalImportLine.batch_id == batch_id)
    if import_status:
        q = q.filter(HistoricalImportLine.import_status == import_status)
    lines = q.order_by(HistoricalImportLine.source_row).all()
    return [
        {
            "id": l.id,
            "batch_id": l.batch_id,
            "source_sheet": l.source_sheet,
            "source_row": l.source_row,
            "corte_id_raw": l.corte_id_raw,
            "scout_name_raw": l.scout_name_raw,
            "scout_id_resolved": l.scout_id_resolved,
            "supervisor_raw": l.supervisor_raw,
            "supervisor_id_resolved": l.supervisor_id_resolved,
            "driver_license_raw": l.driver_license_raw,
            "driver_id_resolved": l.driver_id_resolved,
            "driver_name_raw": l.driver_name_raw,
            "payment_scheme_raw": l.payment_scheme_raw,
            "payment_rule_raw": l.payment_rule_raw,
            "milestone_raw": l.milestone_raw,
            "amount_paid_raw": l.amount_paid_raw,
            "amount_paid": float(l.amount_paid) if l.amount_paid else None,
            "import_status": l.import_status,
            "import_reason": l.import_reason,
            "attribution_status": l.attribution_status,
            "attribution_reason": l.attribution_reason,
            "payment_financial_status": l.payment_financial_status,
            "payment_financial_reason": l.payment_financial_reason,
            "payment_blocking_status": l.payment_blocking_status,
            "payment_blocking_reason": l.payment_blocking_reason,
            "blocks_future_payment": l.blocks_future_payment,
            "final_status": l.final_status,
            "paid_history_id": l.paid_history_id,
            "unique_hash": l.unique_hash,
            "created_at": str(l.created_at) if l.created_at else None,
        }
        for l in lines
    ]


@router.get("/historical-imports/{batch_id}/errors.csv")
def export_historical_errors(batch_id: int, db: Session = Depends(get_db)):
    csv_content = get_batch_errors_csv(db, batch_id)
    return Response(content=csv_content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=errors_batch_{batch_id}.csv"})


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Scout Bulk Upload
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/scouts/upload-preview")
def scouts_upload_preview(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)
    used_sheet = detected_sheet or ""

    # Validate sheet type
    is_valid, import_type, err_msg = validate_sheet_for_endpoint(used_sheet, "scouts")
    if not is_valid:
        log_wrong_sheet("scouts", used_sheet)
        raise HTTPException(
            status_code=400,
            detail={
                "error": "wrong_sheet_for_import",
                "sheet": used_sheet,
                "sheet_type": import_type,
                "message": err_msg,
                "valid_sheets": get_sheets_for_endpoint("scouts"),
            },
        )

    t0 = time.time()
    log_preview_start(import_type, filename, used_sheet, len(rows))

    result = {
        "sheet": used_sheet,
        "import_type": import_type,
        "total_rows": len(rows),
        "will_create": 0,
        "will_update": 0,
        "duplicate_skipped": 0,
        "manual_review": 0,
        "rejected": 0,
        "lines": [],
    }

    for i, row in enumerate(rows):
        line = _classify_scout_row(db, row, used_sheet, i + 2)
        result["lines"].append(line)
        s = line.get("status", "rejected")
        if s == "will_create":
            result["will_create"] += 1
        elif s == "will_update":
            result["will_update"] += 1
        elif s == "duplicate":
            result["duplicate_skipped"] += 1
        elif s == "manual_review":
            result["manual_review"] += 1
        else:
            result["rejected"] += 1

    elapsed = (time.time() - t0) * 1000
    log_preview_done(
        import_type, total_rows=len(rows),
        ready=result["will_create"] + result["will_update"],
        review=result["manual_review"],
        rejected=result["rejected"],
        duplicate=result["duplicate_skipped"],
        elapsed_ms=elapsed,
    )
    return result


def _classify_scout_row(db: Session, row: dict, sheet: str, row_num: int) -> dict:
    scout_name = str(row.get("SCOUT", row.get("scout", ""))).strip()
    supervisor = str(row.get("SUPERVISOR", row.get("supervisor", ""))).strip()
    modalidad = str(row.get("MODALIDAD", row.get("modalidad", ""))).strip()
    desde = str(row.get("DESDE", row.get("desde", ""))).strip()
    hasta = str(row.get("HASTA", row.get("hasta", ""))).strip()
    estado = str(row.get("ESTADO", row.get("estado", row.get("ACTIVO", "")))).strip().upper()
    regimen = str(row.get("regimen", row.get("REGIMEN", ""))).strip()

    line = {
        "source_sheet": sheet,
        "source_row": row_num,
        "scout_name_raw": scout_name or None,
        "supervisor_raw": supervisor or None,
        "scout_type_raw": modalidad or None,
        "active_from": desde or None,
        "active_to": hasta or None,
        "estado_raw": estado or None,
        "status": "rejected",
        "reason": None,
    }

    if not scout_name:
        line["reason"] = "sin nombre de scout"
        return line

    # Resolve supervisor
    supervisor_id = None
    if supervisor:
        sup = db.query(Scout).filter(Scout.scout_name.ilike(supervisor)).first()
        if sup:
            supervisor_id = sup.id
    line["supervisor_id_resolved"] = supervisor_id

    # Check if scout exists
    existing = db.query(Scout).filter(Scout.scout_name.ilike(scout_name)).first()
    if existing:
        # Check if needs update (different supervisor, modalidad, etc.)
        needs_update = False
        if supervisor_id and existing.supervisor_id != supervisor_id:
            needs_update = True
        if modalidad and existing.scout_type != modalidad:
            needs_update = True
        if needs_update:
            line["status"] = "will_update"
            line["existing_id"] = existing.id
            line["reason"] = "actualizar supervisor/modalidad"
        else:
            line["status"] = "duplicate"
            line["existing_id"] = existing.id
            line["reason"] = "sin cambios"
        return line

    is_active = estado not in ("INACTIVO", "FALSE", "DESACTIVADO", "NO", "0")
    line["status"] = "will_create"
    line["is_active"] = is_active
    return line


@router.post("/scouts/upload-commit")
def scouts_upload_commit(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)
    used_sheet = detected_sheet or ""

    result = {
        "sheet": used_sheet,
        "total_rows": len(rows),
        "created": 0,
        "updated": 0,
        "duplicate_skipped": 0,
        "manual_review": 0,
        "rejected": 0,
    }

    for i, row in enumerate(rows):
        line = _classify_scout_row(db, row, used_sheet, i + 2)
        s = line.get("status", "rejected")

        if s == "will_create":
            scout = Scout(
                scout_name=line["scout_name_raw"],
                scout_type=line.get("scout_type_raw"),
                supervisor_name_raw=line.get("supervisor_raw"),
                supervisor_id=line.get("supervisor_id_resolved"),
                imported_from="bulk_upload",
                source_sheet=used_sheet,
                source_row=line["source_row"],
                status="active" if line.get("is_active", True) else "inactive",
            )
            db.add(scout)
            result["created"] += 1
        elif s == "will_update":
            existing = db.query(Scout).filter(Scout.id == line.get("existing_id")).first()
            if existing:
                if line.get("supervisor_id_resolved"):
                    existing.supervisor_id = line["supervisor_id_resolved"]
                if line.get("scout_type_raw"):
                    existing.scout_type = line["scout_type_raw"]
                existing.imported_from = "bulk_upload"
                existing.source_sheet = used_sheet
                existing.source_row = line["source_row"]
                result["updated"] += 1
        elif s == "duplicate":
            result["duplicate_skipped"] += 1
        elif s == "manual_review":
            result["manual_review"] += 1
        else:
            result["rejected"] += 1

    db.commit()
    return result


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Scheme Import
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/schemes/import-preview")
def schemes_import_preview(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)
    used_sheet = detected_sheet or ""

    # Validate sheet type
    is_valid, import_type, err_msg = validate_sheet_for_endpoint(used_sheet, "schemes")
    if not is_valid:
        log_wrong_sheet("schemes", used_sheet)
        raise HTTPException(
            status_code=400,
            detail={
                "error": "wrong_sheet_for_import",
                "sheet": used_sheet,
                "sheet_type": import_type,
                "message": err_msg,
                "valid_sheets": get_sheets_for_endpoint("schemes"),
            },
        )

    t0 = time.time()
    log_preview_start(import_type, file.filename or "", used_sheet, len(rows))
    result = preview_scheme_import(db, rows, used_sheet)
    elapsed = (time.time() - t0) * 1000
    log_preview_done(
        import_type, total_rows=result.get("total_rows", 0),
        ready=result.get("will_import", 0),
        rejected=result.get("errors", 0),
        duplicate=result.get("will_skip", 0),
        elapsed_ms=elapsed,
    )
    return result


@router.post("/schemes/import-commit")
def schemes_import_commit(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    created_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)
    used_sheet = detected_sheet or ""
    preview = preview_scheme_import(db, rows, used_sheet)
    result = commit_scheme_import(db, preview, created_by)
    return result


@router.get("/scheme-versions")
def list_scheme_versions(
    scheme_type: Optional[str] = Query(None),
    active_only: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    return get_scheme_versions(db, scheme_type, active_only)


@router.get("/scheme-change-log")
def get_change_log(scheme_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    return get_scheme_change_log(db, scheme_id)


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Manual Payments
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/manual-payments/preview")
def manual_payment_preview(data: ManualPaymentCreate, db: Session = Depends(get_db)):
    return create_manual_payment_preview(db, data.model_dump())


@router.post("/manual-payments/approve")
def manual_payment_approve(
    body: ManualPaymentApprove,
    payment_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    if not payment_id:
        # If no payment_id in query, try to find a draft payment by scout
        mp = db.query(ManualPayment).filter(
            ManualPayment.status == "draft"
        ).order_by(ManualPayment.created_at.desc()).first()
        if not mp:
            raise HTTPException(status_code=404, detail="No hay pagos manuales en draft. Especifica payment_id.")
        payment_id = mp.id

    return approve_manual_payment(db, payment_id, body.approved_by, body.payment_reference)


@router.post("/manual-payments/mark-paid")
def manual_payment_mark_paid(
    payment_id: Optional[int] = Query(None),
    paid_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    if not payment_id:
        mp = db.query(ManualPayment).filter(
            ManualPayment.status == "approved"
        ).order_by(ManualPayment.approved_at.desc()).first()
        if not mp:
            raise HTTPException(status_code=404, detail="No hay pagos manuales aprobados. Especifica payment_id.")
        payment_id = mp.id

    return mark_manual_payment_paid(db, payment_id, paid_by)


@router.post("/manual-payments")
def create_manual_payment(data: ManualPaymentCreate, db: Session = Depends(get_db)):
    scout = db.query(Scout).filter(Scout.id == data.scout_id).first()
    if not scout:
        raise HTTPException(status_code=404, detail="Scout no encontrado")

    if not data.amount or float(data.amount) <= 0:
        raise HTTPException(status_code=400, detail="Monto debe ser mayor a 0")

    if not data.reason:
        raise HTTPException(status_code=400, detail="Motivo obligatorio")

    mp = ManualPayment(
        cutoff_run_id=data.cutoff_run_id,
        scout_id=data.scout_id,
        supervisor_id=data.supervisor_id,
        driver_id=data.driver_id,
        driver_license_raw=data.driver_license_raw,
        payment_scheme_id=data.payment_scheme_id,
        payment_rule=data.payment_rule,
        amount=Decimal(str(data.amount)),
        currency=data.currency or "PEN",
        reason=data.reason,
        status="draft",
        created_by=data.created_by,
    )
    db.add(mp)
    db.commit()
    db.refresh(mp)
    return {
        "id": mp.id,
        "scout_id": mp.scout_id,
        "amount": float(mp.amount),
        "reason": mp.reason,
        "status": mp.status,
    }


@router.get("/manual-payments")
def get_manual_payments(
    scout_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    return list_manual_payments(db, scout_id, status)


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Supervisor Commission
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/commissions/calculate")
def calculate_commission(
    cutoff_run_id: int,
    commission_rate: float = Query(0.10),
    db: Session = Depends(get_db),
):
    try:
        return calculate_supervisor_commission(db, cutoff_run_id, commission_rate)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/commissions")
def get_commissions(cutoff_run_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    return list_commissions(db, cutoff_run_id)


@router.post("/commissions/{commission_id}/mark-paid")
def mark_commission_paid_endpoint(commission_id: int, db: Session = Depends(get_db)):
    try:
        return mark_commission_paid(db, commission_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4: Scout Bonuses
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/bonuses")
def create_scout_bonus(data: ScoutBonusCreate, db: Session = Depends(get_db)):
    try:
        return create_bonus(db, data.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/bonuses/{bonus_id}/approve")
def approve_scout_bonus(
    bonus_id: int,
    body: ScoutBonusApprove,
    db: Session = Depends(get_db),
):
    try:
        return approve_bonus(db, bonus_id, body.approved_by)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/bonuses/{bonus_id}/mark-paid")
def mark_bonus_paid_endpoint(bonus_id: int, db: Session = Depends(get_db)):
    try:
        return mark_bonus_paid(db, bonus_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/bonuses")
def get_bonuses(
    cutoff_run_id: Optional[int] = Query(None),
    scout_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    return list_bonuses(db, cutoff_run_id, scout_id)


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4.6: Historical Attributions
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/attributions/preview")
def attributions_preview(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)
    used_sheet = detected_sheet or ""

    # Validate sheet type
    is_valid, import_type, err_msg = validate_sheet_for_endpoint(used_sheet, "attributions")
    if not is_valid:
        log_wrong_sheet("attributions", used_sheet)
        raise HTTPException(
            status_code=400,
            detail={
                "error": "wrong_sheet_for_import",
                "sheet": used_sheet,
                "sheet_type": import_type,
                "message": err_msg,
                "valid_sheets": get_sheets_for_endpoint("attributions"),
            },
        )

    t0 = time.time()
    log_preview_start(import_type, filename, used_sheet, len(rows))
    result = preview_attributions(db, rows, filename, used_sheet)
    result["import_type"] = import_type
    elapsed = (time.time() - t0) * 1000
    log_preview_done(
        import_type, total_rows=result.get("total_rows", 0),
        ready=result.get("ready_to_import", 0),
        review=result.get("manual_review", 0),
        rejected=result.get("rejected", 0),
        duplicate=result.get("duplicates", 0),
        elapsed_ms=elapsed,
        top_errors={"conflicts": result.get("conflicts", 0)},
    )
    return result


@router.post("/attributions/commit")
def attributions_commit(
    file: UploadFile = File(...),
    sheet: Optional[str] = Query(None),
    uploaded_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    rows, detected_sheet = _extract_xlsx_rows(content, sheet)

    # Run preview to classify
    preview = preview_attributions(db, rows, filename, detected_sheet)
    lines = preview.get("lines", [])

    result = commit_attributions(db, 0, lines)
    result["source_file"] = filename
    result["sheet"] = detected_sheet
    return result


@router.get("/attributions/imports")
def list_attribution_batches(db: Session = Depends(get_db)):
    batches = db.query(HistoricalAttribution.import_batch_id).distinct().all()
    result = []
    for (bid,) in batches:
        if bid is None:
            continue
        count = db.query(HistoricalAttribution).filter(
            HistoricalAttribution.import_batch_id == bid
        ).count()
        imported = db.query(HistoricalAttribution).filter(
            HistoricalAttribution.import_batch_id == bid,
            HistoricalAttribution.import_status == "ready_to_import",
        ).count()
        sample = db.query(HistoricalAttribution).filter(
            HistoricalAttribution.import_batch_id == bid
        ).first()
        result.append({
            "batch_id": bid,
            "source_file": sample.source_file if sample else None,
            "total_rows": count,
            "imported": imported,
            "created_at": str(sample.created_at) if sample else None,
        })
    return result


@router.get("/attributions/imports/{batch_id}/lines")
def get_attribution_batch_lines(
    batch_id: int,
    import_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(HistoricalAttribution).filter(
        HistoricalAttribution.import_batch_id == batch_id
    )
    if import_status:
        q = q.filter(HistoricalAttribution.import_status == import_status)
    rows = q.order_by(HistoricalAttribution.source_row).all()
    return [
        {
            "id": r.id,
            "source_sheet": r.source_sheet,
            "source_row": r.source_row,
            "scout_name_raw": r.scout_name_raw,
            "scout_id_resolved": r.scout_id_resolved,
            "driver_license_raw": r.driver_license_raw,
            "driver_id_resolved": r.driver_id_resolved,
            "driver_name_raw": r.driver_name_raw,
            "origin_raw": r.origin_raw,
            "payment_status_raw": r.payment_status_raw,
            "payment_amount": float(r.payment_amount) if r.payment_amount else None,
            "import_status": r.import_status,
            "import_reason": r.import_reason,
            "linked_assignment_id": r.linked_assignment_id,
            "created_at": str(r.created_at) if r.created_at else None,
        }
        for r in rows
    ]


@router.get("/attributions/imports/{batch_id}/errors.csv")
def attributions_errors_csv(batch_id: int, db: Session = Depends(get_db)):
    csv_content = get_attribution_batch_errors_csv(db, batch_id)
    return Response(content=csv_content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=attr_errors_{batch_id}.csv"})


@router.get("/attributions")
def list_attributions(
    scout_id: Optional[int] = Query(None),
    driver_id: Optional[str] = Query(None),
    license: Optional[str] = Query(None),
    source_file: Optional[str] = Query(None),
    source_sheet: Optional[str] = Query(None),
    import_status: Optional[str] = Query(None),
    origin_raw: Optional[str] = Query(None),
    cutoff_external_id: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    filters = {}
    if scout_id: filters["scout_id"] = scout_id
    if driver_id: filters["driver_id"] = driver_id
    if license: filters["license"] = license
    if source_file: filters["source_file"] = source_file
    if source_sheet: filters["source_sheet"] = source_sheet
    if import_status: filters["import_status"] = import_status
    if origin_raw: filters["origin_raw"] = origin_raw
    if cutoff_external_id: filters["cutoff_external_id"] = cutoff_external_id
    return get_attributions(db, filters, limit, offset)


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4.7: Workbook Import (Integral)
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/workbook-import/preview")
def workbook_import_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    sheets_data = _extract_xlsx_all_sheets(content)
    result = workbook_preview(db, filename, sheets_data)
    return result


@router.post("/workbook-import/commit")
def workbook_import_commit(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    content = file.file.read()
    sheets_data = _extract_xlsx_all_sheets(content)
    result = workbook_commit(db, sheets_data, filename)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# FASE 4.5: Template download
# ═══════════════════════════════════════════════════════════════════════════

import os as _os

_TEMPLATE_DIR = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))), "static")
_TEMPLATE_NAME = "Plantilla_AFILIATOR_Carga_Historica_Esquemas_Manual.xlsx"


@router.get("/templates")
def list_templates():
    templates = []
    if _os.path.isdir(_TEMPLATE_DIR):
        for f in _os.listdir(_TEMPLATE_DIR):
            if f.endswith(".xlsx"):
                templates.append({
                    "name": f,
                    "url": f"/scout-liq/templates/{f}",
                    "sheets": _get_sheet_names(_os.path.join(_TEMPLATE_DIR, f)),
                })
    return {"templates": templates}


@router.get("/templates/historical-import")
def download_historical_template():
    path = _os.path.join(_TEMPLATE_DIR, _TEMPLATE_NAME)
    if not _os.path.exists(path):
        raise HTTPException(status_code=404, detail="Plantilla no encontrada. Ejecuta: python scripts/generate_template.py")
    return Response(
        content=open(path, "rb").read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={_TEMPLATE_NAME}"},
    )


@router.get("/templates/{template_name}")
def download_template_by_name(template_name: str):
    path = _os.path.join(_TEMPLATE_DIR, template_name)
    if not _os.path.exists(path):
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    return Response(
        content=open(path, "rb").read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={template_name}"},
    )


@router.post("/templates/xlsx-sheets")
def get_xlsx_sheets(file: UploadFile = File(...)):
    content = file.file.read()
    sheets = _get_xlsx_sheet_names(content)
    # Also get row counts and types
    sheet_info = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        for name in sheets:
            ws = wb[name]
            row_count = ws.max_row - 1 if ws.max_row else 0  # subtract header
            import_type, _ = classify_sheet(name)
            sheet_info.append({
                "name": name,
                "import_type": import_type,
                "import_type_label": get_sheet_type_label(import_type),
                "row_count": row_count,
            })
        wb.close()
    except Exception:
        sheet_info = [{"name": s, "import_type": "unknown", "import_type_label": "Desconocido", "row_count": 0} for s in sheets]

    return {
        "filename": file.filename,
        "sheets": sheets,
        "sheet_info": sheet_info,
    }


def _get_sheet_names(path: str) -> List[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def _get_xlsx_sheet_names(content: bytes) -> List[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════════════════════
# Util: XLSX extraction
# ═══════════════════════════════════════════════════════════════════════════

def _extract_xlsx_rows(content: bytes, sheet: Optional[str] = None, skip_header: bool = True):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = None
    detected_sheet = ""

    if sheet:
        ws = wb[sheet]
        detected_sheet = sheet
    else:
        ws = wb.active
        detected_sheet = ws.title if ws else ""

    if not ws:
        wb.close()
        raise HTTPException(status_code=400, detail="No se pudo encontrar la hoja especificada")

    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, []) if skip_header else []
    headers = [str(h).strip() if h else "" for h in headers_row]
    rows = []
    row_num = 1 if not skip_header else 2
    for row in rows_iter:
        d = {}
        for j, cell in enumerate(row):
            key = headers[j] if j < len(headers) else f"col_{j}"
            val = str(cell).strip() if cell is not None else ""
            if val:
                d[key] = val
        row_num += 1
        # Skip rows that start with "OBLIGATORIO" (template row 2)
        if skip_header and row_num == 3:
            # Check if this is the "OBLIGATORIO" marker row
            vals = [str(c).strip() if c else "" for c in row]
            if all(v == "OBLIGATORIO" or not v for v in vals):
                continue
        # Skip completely empty rows
        if any(d.values()):
            rows.append(d)
    wb.close()
    return rows, detected_sheet


# ═══════════════════════════════════════════════════════════
# OPERATION VIEW — Grilla unificada de afiliaciones
# ═══════════════════════════════════════════════════════════

@router.get("/operation/summary")
def operation_summary(
    week_iso: Optional[str] = None,
    hire_date_from: Optional[str] = None,
    hire_date_to: Optional[str] = None,
    scout_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    origin: Optional[str] = None,
    only_manual_review: bool = False,
    only_paid: bool = False,
    only_without_driver: bool = False,
    only_without_scout: bool = False,
    db: Session = Depends(get_db),
):
    filters = {k: v for k, v in {
        "week_iso": week_iso, "hire_date_from": hire_date_from,
        "hire_date_to": hire_date_to, "scout_id": scout_id,
        "supervisor_id": supervisor_id, "origin": origin,
        "only_manual_review": only_manual_review, "only_paid": only_paid,
        "only_without_driver": only_without_driver,
        "only_without_scout": only_without_scout,
    }.items() if v}
    return get_operation_summary(db, **filters)


@router.get("/operation/filters")
def operation_filters(db: Session = Depends(get_db)):
    return get_operation_filters(db)


@router.get("/operation/affiliations")
def operation_affiliations(
    week_iso: Optional[str] = None,
    hire_date_from: Optional[str] = None,
    hire_date_to: Optional[str] = None,
    scout_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    origin: Optional[str] = None,
    alert_level: Optional[str] = None,
    only_manual_review: bool = False,
    only_paid: bool = False,
    only_without_driver: bool = False,
    only_without_scout: bool = False,
    driver_id: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    return get_affiliations(
        db=db,
        week_iso=week_iso,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        scout_id=scout_id,
        supervisor_id=supervisor_id,
        origin=origin,
        alert_level=alert_level,
        only_manual_review=only_manual_review,
        only_paid=only_paid,
        only_without_driver=only_without_driver,
        only_without_scout=only_without_scout,
        driver_id_filter=driver_id,
        limit=limit,
        offset=offset,
    )


@router.get("/operation/affiliations/{row_id}")
def operation_affiliation_detail(row_id: int, db: Session = Depends(get_db)):
    detail = get_affiliation_detail(db, row_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Afiliacion no encontrada")
    return detail


@router.get("/operation/export")
def operation_export(
    week_iso: Optional[str] = None,
    hire_date_from: Optional[str] = None,
    hire_date_to: Optional[str] = None,
    scout_id: Optional[int] = None,
    origin: Optional[str] = None,
    alert_level: Optional[str] = None,
    only_manual_review: bool = False,
    only_without_driver: bool = False,
    only_without_scout: bool = False,
    db: Session = Depends(get_db),
):
    filters = {
        k: v for k, v in {
            "week_iso": week_iso,
            "hire_date_from": hire_date_from,
            "hire_date_to": hire_date_to,
            "scout_id": scout_id,
            "origin": origin,
            "alert_level": alert_level,
            "only_manual_review": only_manual_review,
            "only_without_driver": only_without_driver,
            "only_without_scout": only_without_scout,
        }.items() if v
    }
    csv_data = export_affiliations_csv(db, filters)
    return Response(
        content=csv_data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=afiliaciones.csv"},
    )


# ═══════════════════════════════════════════════════════════════════════════
# CANONICAL OPERATION — Fuente maestra module_ct_cabinet_drivers
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/operation/canonical", response_model=CanonicalOperationSnapshotResponse)
def operation_canonical_snapshot(
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    origin: Optional[str] = Query(None),
    scout_id: Optional[int] = Query(None),
    attribution_status: Optional[str] = Query(None),
    payment_status: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    scheme_type: Optional[str] = Query(None, description="cabinet | fleet | custom"),
    db: Session = Depends(get_db),
):
    """
    Canonical operation snapshot desde module_ct_cabinet_drivers.
    Incluye todos los drivers (con y sin scout), viajes reales desde
    trips_2025/trips_2026, pago como overlay (cutoff + historical),
    y metadata de frescura.
    """
    return get_canonical_operation_snapshot(
        db,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        origin=origin,
        scout_id=scout_id,
        attribution_status=attribution_status,
        payment_status=payment_status,
        limit=limit,
        offset=offset,
        scheme_type=scheme_type,
    )


@router.get("/operation/diagnostic", response_model=OperationDiagnosticResponse)
def operation_diagnostic(
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    origin: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Diagnostico liviano: total drivers, con/sin scout, data lag, conflictos. NO calcula trips."""
    return get_operation_diagnostic(
        db,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        origin=origin,
    )


# ═══════════════════════════════════════════════════════════
# COHORTES — Modelo temporal ISO week
# ═══════════════════════════════════════════════════════════

@router.get("/operation/cohorts", response_model=CohortListResponse)
def operation_cohorts(
    readiness: Optional[str] = Query(None, description="Filtrar por readiness_status: open, mature, locked, paid"),
    db: Session = Depends(get_db),
):
    """Lista todas las cohortes ISO detectadas desde la fuente, con metadata temporal y readiness_status."""
    items = get_iso_cohorts(db, readiness_filter=readiness)
    return CohortListResponse(total=len(items), cohorts=items)


@router.get("/operation/cohorts/diagnostic", response_model=CohortDiagnosticResponse)
def operation_cohorts_diagnostic(
    db: Session = Depends(get_db),
):
    """Diagnostico de cohortes: conteos por estado de madurez, cohortes liquidables."""
    return get_cohort_diagnostic(db)


# ═══════════════════════════════════════════════════════════
# PAYMENT SCHEME RESOLVER
# ═══════════════════════════════════════════════════════════

@router.get("/payment-schemes/resolve", response_model=ResolvedPaymentScheme)
def resolve_payment_scheme(
    cohort_iso_week: str,
    scheme_type: str,
    db: Session = Depends(get_db),
):
    """
    Resuelve la version de esquema de pago aplicable para una cohorte ISO.

    - cohort_iso_week: ej. '2026-W18'
    - scheme_type: cabinet | fleet | custom
    """
    try:
        return resolve_payment_scheme_for_cohort(
            db,
            cohort_iso_week=cohort_iso_week,
            scheme_type=scheme_type,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ═══════════════════════════════════════════════════════════
# PAYMENT SCHEME ADMIN — CRUD completo
# ═══════════════════════════════════════════════════════════

@router.get("/payment-schemes", response_model=List[PaymentSchemeListItem])
def api_list_payment_schemes(db: Session = Depends(get_db)):
    """Lista todos los esquemas de pago con su version activa."""
    return list_payment_schemes(db)


@router.get("/payment-schemes/history")
def api_get_payment_schemes_history(
    scheme_type: Optional[str] = Query(None, description="Filtrar por tipo: cabinet, fleet, custom"),
    db: Session = Depends(get_db),
):
    """Historico de versiones por scheme_type."""
    return get_payment_schemes_history(db, scheme_type=scheme_type)


@router.get("/payment-schemes/{scheme_id}", response_model=PaymentSchemeDetail)
def api_get_payment_scheme(scheme_id: int, db: Session = Depends(get_db)):
    """Detalle de un esquema con todas sus versiones y tiers."""
    try:
        return get_payment_scheme_detail(db, scheme_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/payment-schemes", response_model=SchemeCreatedResponse)
def api_create_payment_scheme(
    body: CreatePaymentSchemeRequest,
    db: Session = Depends(get_db),
):
    """Crea un nuevo esquema de pago (cabinet, fleet, custom)."""
    try:
        return create_payment_scheme(db, body.name, body.scheme_type, body.description)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payment-schemes/{scheme_id}/versions", response_model=VersionCreatedResponse)
def api_create_payment_scheme_version(
    scheme_id: int,
    body: CreateVersionRequest,
    db: Session = Depends(get_db),
):
    """Crea una nueva version draft de un esquema."""
    try:
        return create_payment_scheme_version(
            db,
            scheme_id=scheme_id,
            version_name=body.version_name,
            valid_from_cohort_iso_week=body.valid_from_cohort_iso_week,
            maturity_days=body.maturity_days,
            min_activated=body.min_activated,
            activation_rule=body.activation_rule,
            quality_rule=body.quality_rule,
            formula_type=body.formula_type,
            currency=body.currency,
            tiers=[t.model_dump() for t in body.tiers],
            volume_rule=body.volume_rule,
            min_volume_count=body.min_volume_count,
            pays_on_rule=body.pays_on_rule,
            payout_formula_type=body.payout_formula_type,
            counts_volume_rule=body.counts_volume_rule,
            counts_quality_rule=body.counts_quality_rule,
            maturity_window_days=body.maturity_window_days,
            fixed_payout_amount=body.fixed_payout_amount,
            minimum_enabled=body.minimum_enabled,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payment-scheme-versions/{version_id}/activate", response_model=VersionActivatedResponse)
def api_activate_payment_scheme_version(
    version_id: int,
    db: Session = Depends(get_db),
):
    """
    Activa una version draft. Automaticamente cierra la version activa anterior
    estableciendo su valid_to_cohort_iso_week a la semana anterior al valid_from de la nueva.
    """
    try:
        return activate_payment_scheme_version(db, version_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/payment-scheme-versions/{version_id}/archive")
def api_archive_payment_scheme_version(
    version_id: int,
    db: Session = Depends(get_db),
):
    """Archiva una version draft. No se puede archivar versiones activas o usadas por cortes."""
    try:
        return archive_payment_scheme_version(db, version_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ═══════════════════════════════════════════════════════════
# MANUAL OVERRIDES — Acciones manuales auditables
# ═══════════════════════════════════════════════════════════

@router.get("/manual-overrides", response_model=List[ManualOverrideResponse])
def api_list_manual_overrides(
    driver_id: Optional[str] = None,
    override_type: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    return list_manual_overrides(db, driver_id=driver_id, override_type=override_type, status=status)


@router.get("/drivers/{driver_id}/manual-overrides", response_model=List[ManualOverrideResponse])
def api_get_driver_overrides(driver_id: str, db: Session = Depends(get_db)):
    return get_driver_overrides(db, driver_id)


@router.post("/manual-overrides", response_model=ManualOverrideResponse)
def api_create_manual_override(
    body: CreateManualOverrideRequest,
    db: Session = Depends(get_db),
):
    try:
        return create_manual_override(
            db,
            driver_id=body.driver_id,
            override_type=body.override_type,
            reason=body.reason,
            cohort_iso_week=body.cohort_iso_week,
            scout_id=body.scout_id,
            scout_id_before=body.scout_id_before,
            amount=body.amount,
            currency=body.currency,
            notes=body.notes,
            created_by=body.created_by,
            auto_approve=True,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/manual-overrides/{override_id}/approve", response_model=ManualOverrideResponse)
def api_approve_manual_override(
    override_id: int,
    approved_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        return approve_manual_override(db, override_id, approved_by)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/manual-overrides/{override_id}/apply", response_model=ManualOverrideResponse)
def api_apply_manual_override(override_id: int, db: Session = Depends(get_db)):
    try:
        return apply_manual_override(db, override_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/manual-overrides/{override_id}/reject", response_model=ManualOverrideResponse)
def api_reject_manual_override(override_id: int, db: Session = Depends(get_db)):
    try:
        return reject_manual_override(db, override_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ═══════════════════════════════════════════════════════════
# DASHBOARD — KPIs ejecutivos
# ═══════════════════════════════════════════════════════════

@router.get("/dashboard/overview")
def dashboard_overview(
    week_iso: Optional[str] = None,
    scout_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    origin: Optional[str] = None,
    hire_date_from: Optional[str] = None,
    hire_date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    f = {k: v for k, v in {
        "week_iso": week_iso, "scout_id": scout_id,
        "supervisor_id": supervisor_id, "origin": origin,
        "hire_date_from": hire_date_from, "hire_date_to": hire_date_to,
    }.items() if v}
    return get_dashboard_overview(db, **f)


@router.get("/dashboard/by-scout")
def dashboard_by_scout(
    week_iso: Optional[str] = None,
    scout_id: Optional[int] = None,
    origin: Optional[str] = None,
    db: Session = Depends(get_db),
):
    f = {k: v for k, v in {
        "week_iso": week_iso, "scout_id": scout_id, "origin": origin,
    }.items() if v}
    return get_dashboard_by_scout(db, **f)


@router.get("/dashboard/by-week")
def dashboard_by_week(
    scout_id: Optional[int] = None,
    origin: Optional[str] = None,
    db: Session = Depends(get_db),
):
    f = {k: v for k, v in {
        "scout_id": scout_id, "origin": origin,
    }.items() if v}
    return get_dashboard_by_week(db, **f)


@router.get("/dashboard/quality-funnel")
def dashboard_quality_funnel(
    week_iso: Optional[str] = None,
    db: Session = Depends(get_db),
):
    f = {"week_iso": week_iso} if week_iso else {}
    return get_dashboard_quality_funnel(db, **f)


@router.get("/dashboard/alerts")
def dashboard_alerts(
    week_iso: Optional[str] = None,
    scout_id: Optional[int] = None,
    origin: Optional[str] = None,
    db: Session = Depends(get_db),
):
    f = {k: v for k, v in {
        "week_iso": week_iso, "scout_id": scout_id, "origin": origin,
    }.items() if v}
    return get_dashboard_alerts(db, **f)


@router.get("/dashboard/trend")
def dashboard_trend(db: Session = Depends(get_db)):
    return get_cutoff_trend(db)


# ═══════════════════════════════════════════════════════════
# HEALTH & FRESHNESS — Salud de datos operativos
# ═══════════════════════════════════════════════════════════

@router.get("/health/summary")
def health_summary(
    db: Session = Depends(get_db),
    mode: Optional[str] = Query(None, description="lite (default, rapido) | full (incluye detalle cohorts)"),
):
    """Resumen ejecutivo de salud de datos. Default: lite (sin cohorts pesadas, sin trips)."""
    try:
        if mode == "full":
            _logger.info("[GET /health/summary] mode=full")
            return get_health_summary(db)
        return get_health_summary_lite(db)
    except Exception as e:
        _logger.error(f"[GET /health/summary] mode={mode}: {e}\n{traceback.format_exc()}")
        return {
            "global_status": "UNKNOWN",
            "evaluated_at": date.today().isoformat(),
            "sections": {
                "source": {"status": "UNKNOWN", "reason_text": "No disponible", "data_lag_days": None, "last_data_date": None},
                "scouts": {"status": "UNKNOWN", "reason_text": "No disponible", "coverage_pct": None},
                "cohorts": {"status": "UNKNOWN", "reason_text": "No disponible", "warning_count": 0, "blocked_count": 0},
                "jobs": {"status": "UNKNOWN", "reason_text": "No disponible"},
            },
            "alerts": [],
            "status": "ERROR",
            "error_code": "HEALTH_SUMMARY_FAILED",
            "message": str(e),
        }


@router.get("/health/cohorts")
def health_cohorts(
    weeks_limit: int = Query(4, ge=1, le=52, description="Ultimas N semanas a evaluar. Max 12. Default 4 para rendimiento."),
    status: Optional[str] = Query(None, description="Filtrar por status: OK | WARNING | BLOCKED"),
    db: Session = Depends(get_db),
):
    """Salud por cohorte: drivers, scouts, activaciones, maduracion, flags."""
    try:
        if weeks_limit > 12:
            weeks_limit = 12
        return get_cohort_health(db, weeks_limit=weeks_limit, status_filter=status)
    except Exception as e:
        _logger.error(f"[GET /health/cohorts] weeks_limit={weeks_limit} status={status}: {e}\n{traceback.format_exc()}")
        return {
            "cohorts": [],
            "total_cohorts_visible": 0,
            "global_status": "UNKNOWN",
            "global_reason": f"No se pudo evaluar: {e}",
            "warning_count": 0,
            "blocked_count": 0,
            "status": "ERROR",
            "error_code": "COHORTS_FAILED",
            "message": str(e),
        }


@router.get("/health/sources")
def health_sources(db: Session = Depends(get_db)):
    """Salud de la fuente operativa (module_ct_cabinet_drivers)."""
    try:
        return get_source_health(db)
    except Exception as e:
        _logger.error(f"[GET /health/sources] {e}\n{traceback.format_exc()}")
        return {
            "status": "UNKNOWN",
            "reason_text": f"No se pudo evaluar: {e}",
            "last_data_date": None,
            "data_lag_days": None,
            "evaluated_at": date.today().isoformat(),
            "metrics": {},
            "error_code": "SOURCES_FAILED",
            "message": str(e),
        }


@router.get("/health/jobs")
def health_jobs(db: Session = Depends(get_db)):
    """Estado inferido de jobs/crons/procesos batch."""
    try:
        return get_jobs_health(db)
    except Exception as e:
        _logger.error(f"[GET /health/jobs] {e}\n{traceback.format_exc()}")
        return {
            "jobs": [],
            "global_status": "UNKNOWN",
            "global_reason": f"No se pudo evaluar: {e}",
            "inferred_only": True,
            "note": "Error al consultar infraestructura de jobs",
            "error_code": "JOBS_FAILED",
            "message": str(e),
        }


@router.get("/health/score")
def health_score(
    db: Session = Depends(get_db),
    mode: Optional[str] = Query(None, description="lite (default, rapido) | full (incluye detalle cohorts)"),
):
    """Health score global 0-100 con breakdown. Default: lite (sin cohorts, sin trips)."""
    try:
        if mode == "full":
            _logger.info("[GET /health/score] mode=full")
            return compute_health_score(db)
        return compute_health_score_lite(db)
    except Exception as e:
        _logger.error(f"[GET /health/score] mode={mode}: {e}\n{traceback.format_exc()}")
        return {
            "score": None,
            "status": "UNKNOWN",
            "reason_text": f"No se pudo calcular: {e}",
            "breakdown": {},
            "max_score": 100,
            "evaluated_at": date.today().isoformat(),
            "error_code": "SCORE_FAILED",
            "message": str(e),
        }


@router.get("/health/registry")
def health_registry(db: Session = Depends(get_db)):
    """Lista el registro de refresh de fuentes y procesos."""
    try:
        return get_registry(db)
    except Exception as e:
        _logger.error(f"[GET /health/registry] {e}\n{traceback.format_exc()}")
        return {
            "entries": [],
            "status": "ERROR",
            "error_code": "REGISTRY_FAILED",
            "message": str(e),
        }


@router.get("/health/events")
def health_events(
    status: Optional[str] = Query(None, description="open | resolved | all"),
    severity: Optional[str] = Query(None, description="BLOCKED | WARNING | INFO"),
    limit: int = Query(50, ge=1, le=500, description="Max eventos a retornar"),
    db: Session = Depends(get_db),
):
    """Lista eventos de salud detectados."""
    try:
        return get_events(db, status=status, severity=severity, limit=limit)
    except Exception as e:
        _logger.error(f"[GET /health/events] status={status} severity={severity} limit={limit}: {e}\n{traceback.format_exc()}")
        return {
            "events": [],
            "status": "ERROR",
            "error_code": "EVENTS_FAILED",
            "message": str(e),
        }


@router.post("/health/registry/refresh")
def health_registry_refresh(db: Session = Depends(get_db)):
    """Ejecuta ciclo completo: snapshot registry + detectar eventos + resolver recuperados + score."""
    try:
        return full_refresh_cycle(db)
    except Exception as e:
        _logger.error(f"[POST /health/registry/refresh] {e}\n{traceback.format_exc()}")
        return {
            "score": {"score": None, "status": "UNKNOWN"},
            "registry": {"entries": [], "total": 0},
            "events_detected": {"new_events": 0, "events": []},
            "events_resolved": {"resolved_count": 0},
            "cycle_completed_at": datetime.utcnow().isoformat(),
            "status": "ERROR",
            "error_code": "REFRESH_CYCLE_FAILED",
            "message": str(e),
        }


# ═══════════════════════════════════════════════════════════
# RECONCILIATION — Conciliacion operacional/financiera
# ═══════════════════════════════════════════════════════════

@router.get("/reconciliation/export")
def reconciliation_export(
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    scheme_type: Optional[str] = Query(None, description="cabinet | fleet | custom"),
    pay_until_date: Optional[date] = Query(None),
    only_matured: bool = Query(False),
    db: Session = Depends(get_db),
):
    """Exporta CSV de conciliacion con el estado esperado por el sistema."""
    csv_content = export_reconciliation_csv(
        db,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        scheme_type=scheme_type,
        pay_until_date=pay_until_date,
        only_matured=only_matured,
    )
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=reconciliation_export.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@router.post("/reconciliation/compare-upload", response_model=ReconciliationCompareResponse)
def reconciliation_compare_upload(
    file: UploadFile = File(...),
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    scheme_type: Optional[str] = Query(None, description="cabinet | fleet | custom"),
    db: Session = Depends(get_db),
):
    """Sube un CSV de pagos reales y lo compara contra el estado esperado del sistema."""
    try:
        content = file.file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            file.file.seek(0)
            content = file.file.read().decode("latin-1")
        except Exception:
            raise HTTPException(status_code=400, detail="No se pudo decodificar el archivo. Use CSV UTF-8.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {e}")

    result = compare_upload(
        db,
        csv_content=content,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        scheme_type=scheme_type,
    )

    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])

    return ReconciliationCompareResponse(**result)


# ═══════════════════════════════════════════════════════════
# UNIFIED LOAD — Carga unificada plana por licencia
# ═══════════════════════════════════════════════════════════

@router.get("/unified-load/template")
def unified_load_template():
    """Descarga la plantilla simple CSV para carga unificada."""
    import csv as _csv
    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow([
        "licencia", "scout", "supervisor",
        "pagado", "monto_pagado", "fecha_pago", "observacion",
        "driver_id", "nombre_conductor", "origen",
        "tipo_scout", "motivo_pago", "cohorte_iso",
    ])
    writer.writerow([
        "Q12345678", "Scout Alpha", "Juan Perez",
        "SI", "150.00", "2026-01-15", "Pago enero",
        "", "Carlos Garcia", "cabinet",
        "destajo", "Conversion 40%", "2026-W03",
    ])
    writer.writerow([
        "Q87654321", "Scout Beta", "Maria Lopez",
        "NO", "0", "2026-01-15", "Solo atribucion",
        "", "Ana Martinez", "fleet",
        "", "", "",
    ])
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_unificada.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@router.post("/unified-load/preview-stream")
def unified_load_preview_stream(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Streaming preview: devuelve NDJSON con eventos progresivos."""
    content = file.file.read()
    filename = file.filename or "archivo.csv"
    parse_metadata: dict = {}

    def generate():
        yield json.dumps({"type": "started", "filename": filename}) + "\n"

        if filename.lower().endswith(".xlsx"):
            rows, errors = _parse_rows_from_xlsx(content)
            columns = []
        else:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
            rows, errors, parse_metadata = _parse_rows_from_csv(text)
            columns = parse_metadata.get("columns_detected", [])

        if parse_metadata.get("structural_error"):
            yield json.dumps({"type": "structural_error", **parse_metadata}) + "\n"
            yield json.dumps({"type": "summary", "total_rows": 0, "valid_rows": 0,
                              "error_rows": 0, "done": True}) + "\n"
            return

        if not rows:
            yield json.dumps({"type": "error", "message": "Archivo vacio o sin filas de datos"}) + "\n"
            yield json.dumps({"type": "summary", "total_rows": 0, "valid_rows": 0,
                              "error_rows": 0, "done": True}) + "\n"
            return

        yield json.dumps({"type": "file_parsed", "rows": len(rows),
                          "columns": columns}) + "\n"

        for event in unified_preview_stream(db, rows):
            yield json.dumps(event, default=str) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson",
                             headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


@router.post("/unified-load/preview", response_model=UnifiedLoadPreviewResponse)
def unified_load_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Sube CSV/XLSX y devuelve preview con acciones deducidas."""
    content = file.file.read()
    parse_metadata: dict = {}

    if file.filename and file.filename.lower().endswith(".xlsx"):
        rows, errors = _parse_rows_from_xlsx(content)
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        rows, errors, parse_metadata = _parse_rows_from_csv(text)

    # Structural error: return preview with metadata, NO HTTPException
    if parse_metadata.get("structural_error"):
        result = {
            "total_rows": 0, "valid_rows": 0, "error_rows": 0,
            "drivers_found": 0, "drivers_not_found": 0,
            "scouts_to_create": 0, "supervisors_to_create": 0,
            "assignments_to_create": 0, "assignments_to_change": 0,
            "assignments_already_exist": 0,
            "payments_to_create": 0, "already_paid": 0, "amount_mismatch": 0,
            "warnings": errors, "lines": [],
            "parse_metadata": parse_metadata,
        }
        return UnifiedLoadPreviewResponse(**result)

    if not rows and errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))

    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacio o sin filas de datos")

    result = unified_preview(db, rows)
    result["parse_metadata"] = parse_metadata

    if errors:
        result["warnings"] = result.get("warnings", []) + errors

    return UnifiedLoadPreviewResponse(**result)


@router.post("/unified-load/apply", response_model=UnifiedLoadApplyResponse)
def unified_load_apply(
    file: UploadFile = File(...),
    applied_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Aplica solo lo validado en preview."""
    content = file.file.read()

    if file.filename and file.filename.lower().endswith(".xlsx"):
        rows, errors = _parse_rows_from_xlsx(content)
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        rows, errors, _ = _parse_rows_from_csv(text)

    if not rows and errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))

    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacio o sin filas de datos")

    result = unified_apply(db, rows, applied_by=applied_by)
    return UnifiedLoadApplyResponse(**result)


@router.post("/unified-load/apply-stream")
async def unified_load_apply_stream(
    request: Request,
    applied_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Streaming apply: recibe apply_plan JSON del preview y lo ejecuta."""
    body = await request.json()
    plan = body.get("apply_plan", [])
    if not plan:
        raise HTTPException(status_code=400, detail="apply_plan vacio o ausente")

    def generate():
        for event in unified_apply_stream(db, plan, applied_by=applied_by):
            yield json.dumps(event, default=str) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson",
                             headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


@router.post("/unified-load/report/preview")
def unified_load_preview_report(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Descarga reporte de auditoria del preview en formato CSV."""
    content = file.file.read()
    parse_metadata: dict = {}

    if file.filename and file.filename.lower().endswith(".xlsx"):
        rows, errors = _parse_rows_from_xlsx(content)
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        rows, errors, parse_metadata = _parse_rows_from_csv(text)

    if parse_metadata.get("structural_error"):
        raise HTTPException(status_code=400, detail="Error estructural en archivo: " + str(parse_metadata))

    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacio o sin filas de datos")

    csv_content = generate_preview_audit_csv(db, rows)
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=reporte_preview_carga.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@router.post("/unified-load/report/apply")
async def unified_load_apply_report(
    request: Request,
    db: Session = Depends(get_db),
):
    """Descarga reporte de auditoria del apply combinando preview + apply en formato CSV."""
    body = await request.json()
    preview_lines = body.get("preview_lines", [])
    apply_lines = body.get("apply_lines", [])

    if not preview_lines:
        raise HTTPException(status_code=400, detail="preview_lines vacio o ausente")

    csv_content = generate_apply_audit_csv(preview_lines, apply_lines)
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=reporte_final_carga.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@router.post("/unified-load/report/full-audit")
async def unified_load_full_audit_report(
    request: Request,
    db: Session = Depends(get_db),
):
    """Descarga auditoria COMPLETA: TODAS las filas del input + columnas de auditoria.
    Ninguna fila se omite. Incluye columnas originales del archivo y columnas de diagnostico."""
    body = await request.json()
    preview_lines = body.get("preview_lines", [])
    apply_lines = body.get("apply_lines", [])
    file_name = body.get("file_name", "")
    preview_result = body.get("preview_result", {})
    apply_summary = body.get("apply_summary", {})

    if not preview_lines:
        raise HTTPException(status_code=400, detail="preview_lines vacio o ausente")

    # Generar BOM manualmente para compatibilidad Excel
    csv_content = "\ufeff" + generate_full_audit_csv(
        preview_lines, apply_lines, file_name
    )
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=apply_full_audit_report.csv",
            "Content-Type": "text/csv; charset=utf-8",
        },
    )


@router.post("/unified-load/report/summary")
async def unified_load_summary_report(
    request: Request,
    db: Session = Depends(get_db),
):
    """Descarga resumen de auditoria en CSV independiente."""
    body = await request.json()
    preview_result = body.get("preview_result", {})
    apply_summary = body.get("apply_summary", {})
    preview_lines = body.get("preview_lines", [])
    apply_lines = body.get("apply_lines", [])
    file_name = body.get("file_name", "")

    csv_content = "\ufeff" + generate_summary_csv(
        preview_result, apply_summary,
        len(preview_lines), len(apply_lines),
        file_name,
    )
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=apply_summary.csv",
            "Content-Type": "text/csv; charset=utf-8",
        },
    )


@router.get("/unified-load/preview-result/{preview_id}")
def get_preview_result(preview_id: str):
    """Recupera el resultado completo del preview (apply_plan + lines) por ID."""
    data = _get_preview(preview_id)
    if not data:
        raise HTTPException(status_code=404, detail="Preview expirado o no encontrado")
    return data


@router.get("/unified-load/rescue-csv/{preview_id}")
def get_rescue_csv(preview_id: str):
    """Descarga CSV con solo las filas pendientes de correccion (errores + no encontrados)."""
    data = _get_preview(preview_id)
    if not data:
        raise HTTPException(status_code=404, detail="Preview expirado o no encontrado")
    
    lines = data.get("lines", [])
    rescues = [l for l in lines if l.get("status") in ("error", "warning") 
               or "driver_not_found" in l.get("deduced_actions", [])
               or l.get("preview_status") == "error"]
    
    import io as _io, csv as _csv
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["licencia", "scout", "supervisor", "origen", "suggested_fix", "errors"])
    for l in rescues:
        src = l.get("source_row", "")
        lic = l.get("licencia", "")
        scout = l.get("scout", "")
        sup = l.get("supervisor", "")
        origen = l.get("origen", "")
        fix = l.get("suggested_fix", "")
        errs = "; ".join(l.get("errors", []))
        w.writerow([lic, scout, sup, origen, fix, errs])
    
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=pendientes_correccion.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )
