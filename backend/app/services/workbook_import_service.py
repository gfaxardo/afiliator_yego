"""
Workbook Import Service - Importacion integral de workbook completo.
Procesa hojas en orden: scouts → supervisors → schemes → attributions → payments.
Usa contexto temporal entre hojas para resolver referencias cruzadas.
"""

import io
import json
import time
import logging
import unicodedata
from collections import Counter
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Optional, Any, Set, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import text

from app.models.scout_liq import (
    Scout, DriverAssignment, HistoricalImportBatch, HistoricalImportLine,
    PaidHistory, HistoricalAttribution, SchemeVersion, ManualPayment,
)
from app.services.historical_import_service import (
    build_scout_cache, build_license_cache, build_driver_id_cache,
    _extract_field, _classify_standard_row, parse_decimal_safe,
    resolve_scout_cached, normalize_text, is_explicitly_not_paid,
)
from app.services.attribution_import_service import _classify_attribution_row
from app.services.sheet_validator import SHEET_TYPE_MAP, classify_sheet

logger = logging.getLogger("scout_liq")


def normalize_name(val: Optional[str]) -> str:
    """Normalize for comparison - must match build_scout_cache key format (upper)."""
    if not val:
        return ""
    s = str(val).strip().upper()
    return s


class WorkbookContext:
    """Temporary context built from workbook sheets before commit."""

    def __init__(self, db: Session):
        self.db = db
        # Scout name normalized → (raw_name, status, scout_id_or_None)
        self.scout_map: Dict[str, Dict] = {}
        # Supervisor name normalized → (raw_name, scout_name_raw, source_sheet, frequency)
        self.supervisor_candidates: Dict[str, Dict] = {}
        # scout_norm → supervisor_norm
        self.scout_supervisor_links: Dict[str, str] = {}
        # supervisor conflicts: scout_norm → list of (supervisor_norm, source, freq)
        self.supervisor_conflicts: Dict[str, List] = {}
        # DB scout cache
        self.db_scout_cache = build_scout_cache(db)
        # Reverse cache: id → name
        self.db_scout_names: Dict[int, str] = {}
        for s in db.query(Scout).all():
            self.db_scout_names[s.id] = s.scout_name

    def register_scout_from_sheet(self, raw_name: str, supervisor_raw: Optional[str] = None):
        norm = normalize_name(raw_name)
        if not norm:
            return
        if norm not in self.scout_map:
            db_id = self.db_scout_cache.get(normalize_name(raw_name))
            self.scout_map[norm] = {
                "raw_name": raw_name.strip(),
                "exists_in_db": db_id is not None,
                "db_id": db_id,
                "will_be_created": db_id is None,
            }
        if supervisor_raw and supervisor_raw.strip():
            sup_norm = normalize_name(supervisor_raw)
            if sup_norm and sup_norm != norm:
                if sup_norm not in self.supervisor_candidates:
                    self.supervisor_candidates[sup_norm] = {
                        "raw_name": supervisor_raw.strip(),
                        "scouts": Counter(),
                        "sources": set(),
                    }
                self.supervisor_candidates[sup_norm]["scouts"][norm] += 1
                self.supervisor_candidates[sup_norm]["sources"].add("02_SCOUTS")
                # Link if not already linked or same
                if norm not in self.scout_supervisor_links:
                    self.scout_supervisor_links[norm] = sup_norm

    def register_supervisor_from_payment(self, scout_raw: str, supervisor_raw: str):
        scout_norm = normalize_name(scout_raw)
        sup_norm = normalize_name(supervisor_raw)
        if not scout_norm or not sup_norm:
            return
        if sup_norm not in self.supervisor_candidates:
            self.supervisor_candidates[sup_norm] = {
                "raw_name": supervisor_raw.strip(),
                "scouts": Counter(),
                "sources": set(),
            }
        self.supervisor_candidates[sup_norm]["scouts"][scout_norm] += 1
        self.supervisor_candidates[sup_norm]["sources"].add("01_PAGOS_HISTORICOS")
        # Link if not already linked or same source
        existing = self.scout_supervisor_links.get(scout_norm)
        if not existing or existing == sup_norm:
            self.scout_supervisor_links[scout_norm] = sup_norm

    def resolve_scout_for_preview(self, raw_name: str) -> Tuple[Optional[int], str]:
        """Returns (scout_id_or_None, resolution_method)."""
        norm = normalize_name(raw_name)
        if not norm:
            return None, "no_name"
        # Check DB first
        db_id = self.db_scout_cache.get(norm)
        if db_id:
            return db_id, "db_resolved"
        # Check workbook context
        ctx = self.scout_map.get(norm)
        if ctx:
            return ctx.get("db_id"), "workbook_resolved"
        # Try partial match in DB
        for cached_norm, sid in self.db_scout_cache.items():
            if norm in cached_norm or cached_norm in norm:
                return sid, "db_partial_match"
        return None, "not_found"

    def get_supervisor_candidates(self) -> List[dict]:
        """Return supervisor candidates with status."""
        result = []
        for sup_norm, info in self.supervisor_candidates.items():
            # Check if exists in DB
            db_id = self.db_scout_cache.get(sup_norm)
            # Check conflicts
            has_conflict = sup_norm in self.supervisor_conflicts
            scouts = dict(info["scouts"].most_common(10))
            result.append({
                "supervisor_name": info["raw_name"],
                "supervisor_norm": sup_norm,
                "exists_in_db": db_id is not None,
                "db_id": db_id,
                "scouts_count": len(scouts),
                "top_scouts": list(scouts.keys())[:5],
                "frequency_total": sum(scouts.values()),
                "sources": list(info["sources"]),
                "has_conflict": has_conflict,
                "status": "ready" if db_id else "needs_create",
            })
        return sorted(result, key=lambda x: -x["frequency_total"])


def _extract_xlsx_all_sheets(content: bytes) -> Dict[str, List[dict]]:
    """Extract all sheets from an XLSX into a dict of sheet_name → rows."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, [])
        headers = [str(h).strip() if h else "" for h in headers_row]
        rows = []
        for row in rows_iter:
            d = {}
            for j, cell in enumerate(row):
                key = headers[j] if j < len(headers) else f"col_{j}"
                val = str(cell).strip() if cell is not None else ""
                if val:
                    d[key] = val
            if any(d.values()):
                rows.append(d)
        sheets_data[sheet_name] = rows
    wb.close()
    return sheets_data


def workbook_preview(db: Session, source_file: str, sheets_data: Dict[str, List[dict]]) -> Dict[str, Any]:
    """Preview all sheets of a workbook in dependency order."""
    t0 = time.time()

    ctx = WorkbookContext(db)
    result = {
        "source_file": source_file,
        "detected_sheets": [],
        "ignored_sheets": [],
        "scouts": None,
        "supervisors": None,
        "schemes": None,
        "attributions": None,
        "payments": None,
        "global": {},
        "errors": [],
    }

    # Phase 1: detect and classify sheets
    for sheet_name, rows in sheets_data.items():
        import_type, _ = classify_sheet(sheet_name)
        if import_type == "reference_only":
            result["ignored_sheets"].append({"name": sheet_name, "reason": "reference_only"})
        else:
            result["detected_sheets"].append({
                "name": sheet_name, "import_type": import_type, "rows": len(rows)
            })

    # Phase 2: build context from scouts sheet
    scouts_rows = sheets_data.get("02_SCOUTS", [])
    for row in scouts_rows:
        scout_name = str(row.get("scout_name", row.get("SCOUT", ""))).strip()
        supervisor = str(row.get("supervisor_name_raw", row.get("SUPERVISOR", ""))).strip()
        if scout_name:
            ctx.register_scout_from_sheet(scout_name, supervisor)

    # Phase 2b: register supervisors from payment sheet
    payments_rows = sheets_data.get("01_PAGOS_HISTORICOS", [])
    for row in payments_rows:
        scout = str(row.get("scout_name_raw", row.get("SCOUT", ""))).strip()
        sup = str(row.get("supervisor_name_raw", row.get("SUPERVISOR", ""))).strip()
        if scout and sup:
            ctx.register_supervisor_from_payment(scout, sup)
        # Also register scout from payments
        if scout:
            ctx.register_scout_from_sheet(scout)

    # Phase 3: preview scouts
    result["scouts"] = _preview_scouts_section(ctx, scouts_rows)

    # Phase 4: preview supervisors
    result["supervisors"] = _build_supervisors_section(ctx)

    # Phase 5: preview attributions
    attr_rows = sheets_data.get("06_ATRIBUCIONES_HISTORICAS", [])
    result["attributions"] = _preview_attributions_section(db, ctx, attr_rows, source_file, "06_ATRIBUCIONES_HISTORICAS")

    # Phase 6: preview payments (with workbook context so scouts from same workbook resolve)
    result["payments"] = _preview_payments_section(db, ctx, payments_rows, source_file, "01_PAGOS_HISTORICAS")

    # Phase 7: schemes
    schemes_rows = sheets_data.get("03_ESQUEMAS", [])
    result["schemes"] = _preview_schemes_section(schemes_rows)

    # Phase 8: manual payments
    manual_rows = sheets_data.get("04_PAGOS_MANUALES", [])
    result["manual_payments"] = {
        "total_rows": len(manual_rows),
        "note": "Procesamiento manual - usar tab Pagos Manuales"
    }

    # Phase 9: supervisor bonuses
    bonus_rows = sheets_data.get("05_SUPERVISORES_BONOS", [])
    result["supervisor_bonus"] = {
        "total_rows": len(bonus_rows),
        "note": "Procesamiento manual - usar tab Sup & Bonos"
    }

    # Global summary
    result["global"] = {
        "total_rows": sum(s["rows"] for s in result["detected_sheets"]),
        "scouts_ready": result["scouts"].get("will_create", 0) + result["scouts"].get("will_update", 0),
        "supervisors_detected": len(result["supervisors"].get("candidates", [])),
        "supervisor_conflicts": result["supervisors"].get("conflicts", 0),
        "attribution_ready": result["attributions"].get("ready", 0),
        "payment_ready": result["payments"].get("payment_financial", {}).get("ready", 0),
        "payment_not_applicable": result["payments"].get("payment_financial", {}).get("not_applicable", 0),
        "blocking_ready": result["payments"].get("payment_blocking", {}).get("ready", 0),
        "amount_ready": result["payments"].get("payment_financial", {}).get("amount_ready", 0),
        "elapsed_ms": int((time.time() - t0) * 1000),
    }

    logger.info(
        "[SCOUT_LIQ_WORKBOOK] preview_done file=%s sheets=%d ignored=%d scouts=%d supervisors=%d attr_ready=%d pay_ready=%d pay_na=%d amount=%s elapsed_ms=%d",
        source_file,
        len(result["detected_sheets"]), len(result["ignored_sheets"]),
        result["global"]["scouts_ready"],
        result["global"]["supervisors_detected"],
        result["global"]["attribution_ready"],
        result["global"]["payment_ready"],
        result["global"]["payment_not_applicable"],
        result["global"]["amount_ready"],
        result["global"]["elapsed_ms"],
    )

    return result


def _preview_scouts_section(ctx: WorkbookContext, rows: List[dict]) -> dict:
    result = {"total_rows": len(rows), "will_create": 0, "will_update": 0,
              "duplicates": 0, "manual_review": 0, "rejected": 0, "lines": []}
    seen = set()
    for i, row in enumerate(rows):
        scout_name = str(row.get("scout_name", row.get("SCOUT", ""))).strip()
        if not scout_name:
            result["rejected"] += 1
            continue
        norm = normalize_name(scout_name)
        if norm in seen:
            result["duplicates"] += 1
            continue
        seen.add(norm)
        db_id = ctx.db_scout_cache.get(norm)
        if db_id:
            result["will_update"] += 1
        else:
            result["will_create"] += 1
        result["lines"].append({
            "source_row": i + 2,
            "scout_name": scout_name,
            "exists_in_db": db_id is not None,
            "db_id": db_id,
            "action": "update" if db_id else "create",
        })
    return result


def _build_supervisors_section(ctx: WorkbookContext) -> dict:
    candidates = ctx.get_supervisor_candidates()
    conflicts = 0
    for c in candidates:
        if c["has_conflict"]:
            conflicts += 1
    return {
        "candidates": candidates,
        "total_detected": len(candidates),
        "ready_to_link": sum(1 for c in candidates if c["exists_in_db"]),
        "needs_create": sum(1 for c in candidates if not c["exists_in_db"]),
        "conflicts": conflicts,
        "scout_supervisor_links": [
            {"scout": sn, "supervisor": ctx.supervisor_candidates.get(sp, {}).get("raw_name", sp)}
            for sn, sp in ctx.scout_supervisor_links.items()
        ],
    }


def _preview_payments_section(db: Session, ctx: WorkbookContext,
                                rows: List[dict], source_file: str, sheet: str) -> dict:
    """Preview payments using workbook context for scout resolution."""
    result = {
        "total_rows": len(rows),
        "attribution": {"total": 0, "ready": 0, "manual_review": 0, "rejected": 0},
        "payment": {"total": 0, "ready": 0, "not_applicable": 0, "manual_review": 0,
                     "duplicates": 0, "amount_ready": Decimal("0")},
        "payment_financial": {
            "ready": 0, "not_applicable": 0, "manual_review": 0,
            "amount_ready": Decimal("0"),
        },
        "payment_blocking": {
            "ready": 0, "manual_review": 0, "duplicates": 0,
            "amount_ready": Decimal("0"),
        },
        "rejected_only": 0,
        "lines": [],
    }

    scout_cache = ctx.db_scout_cache
    all_licenses = set()
    all_driver_ids = set()
    for row in rows:
        lic = _extract_field(row, "driver_license_raw", "LICENCIA", "licencia", "Brevete")
        if lic: all_licenses.add(lic.strip())
        did = _extract_field(row, "driver_id_resolved")
        if did: all_driver_ids.add(did.strip())

    license_cache = build_license_cache(db, list(all_licenses)) if all_licenses else {}
    driver_id_set = build_driver_id_cache(db, list(all_driver_ids)) if all_driver_ids else set()

    for i, row in enumerate(rows):
        line = _classify_standard_row_with_context(db, row, source_file, sheet, i + 2,
                                                     ctx, scout_cache, license_cache, driver_id_set)
        result["lines"].append(line)

        attr_s = line.get("attribution_status", "")
        fin_s = line.get("payment_financial_status", "")
        blk_s = line.get("payment_blocking_status", "")
        amt = line.get("amount_paid") or 0
        amt_d = Decimal(str(amt))

        if attr_s and attr_s != "attribution_rejected_missing_scout_and_driver":
            result["attribution"]["total"] += 1
            if attr_s == "attribution_ready":
                result["attribution"]["ready"] += 1
            elif attr_s == "attribution_manual_review":
                result["attribution"]["manual_review"] += 1

        # Financial payment metrics
        if fin_s:
            if fin_s == "payment_financial_ready":
                result["payment_financial"]["ready"] += 1
                result["payment_financial"]["amount_ready"] += amt_d
            elif "not_applicable" in fin_s:
                result["payment_financial"]["not_applicable"] += 1
            elif "manual_review" in fin_s:
                result["payment_financial"]["manual_review"] += 1

        # Blocking payment metrics
        if blk_s:
            if blk_s == "payment_blocking_ready":
                result["payment_blocking"]["ready"] += 1
                result["payment_blocking"]["amount_ready"] += amt_d
            elif "manual_review" in blk_s:
                result["payment_blocking"]["manual_review"] += 1
            elif "duplicate" in blk_s:
                result["payment_blocking"]["duplicates"] += 1

        # Legacy compat: payment.ready = blocking_ready
        pay_s = line.get("payment_status", "")
        if pay_s:
            result["payment"]["total"] += 1
            if pay_s == "payment_blocking_ready":
                result["payment"]["ready"] += 1
                result["payment"]["amount_ready"] += amt_d
            elif "not_applicable" in pay_s:
                result["payment"]["not_applicable"] += 1
            elif "manual_review" in pay_s:
                result["payment"]["manual_review"] += 1
            elif "duplicate" in pay_s:
                result["payment"]["duplicates"] += 1

        if line.get("final_status") == "rejected":
            result["rejected_only"] += 1

    return result


def _classify_standard_row_with_context(db: Session, row: dict, source_file: str,
                                         sheet: str, row_num: int,
                                         ctx: WorkbookContext,
                                         scout_cache: Dict[str, int],
                                         license_cache: Dict[str, str],
                                         driver_id_set: Set[str]) -> dict:
    """Same as _classify_standard_row but uses WorkbookContext for scout resolution."""
    scout_name_raw = _extract_field(row, "scout_name_raw", "scout_name")
    driver_license_raw = _extract_field(row, "driver_license_raw")
    driver_id_resolved = _extract_field(row, "driver_id_resolved")
    supervisor_raw = _extract_field(row, "supervisor_name_raw")
    estado_pago = normalize_text(_extract_field(row, "estado_pago", "estado") or "")
    amount_raw = _extract_field(row, "amount_paid", "amount") or "0"
    payment_rule = _extract_field(row, "payment_rule")
    cutoff_external_id = _extract_field(row, "cutoff_external_id", "CORTE_ID", "CORTE")
    fecha_pago_raw = _extract_field(row, "fecha_pago", "FECHA_PAGO")
    currency = _extract_field(row, "currency") or "PEN"

    base = {
        "source_sheet": sheet,
        "source_row": row_num,
        "scout_name_raw": scout_name_raw,
        "driver_license_raw": driver_license_raw,
        "driver_id_resolved": driver_id_resolved,
        "supervisor_raw": supervisor_raw,
        "corte_id_raw": cutoff_external_id,
        "fecha_pago_raw": fecha_pago_raw,
        "payment_rule_raw": payment_rule,
        "amount_paid_raw": amount_raw,
        "estado_pago_raw": estado_pago,
        "currency": currency,
    }

    # Resolve scout: try workbook context first, then DB cache
    scout_id, scout_method = ctx.resolve_scout_for_preview(scout_name_raw) if scout_name_raw else (None, "no_name")
    base["scout_id_resolved"] = scout_id
    base["scout_resolution_method"] = scout_method

    # Resolve driver
    did = driver_id_resolved
    if did:
        base["driver_id_resolved"] = did if did in driver_id_set else None
    elif driver_license_raw:
        base["driver_id_resolved"] = license_cache.get(driver_license_raw.strip())

    # Resolve supervisor
    sup_id, _ = ctx.resolve_scout_for_preview(supervisor_raw) if supervisor_raw else (None, "")
    base["supervisor_id_resolved"] = sup_id

    # Parse amount
    amt = parse_decimal_safe(amount_raw)
    base["amount_paid"] = float(amt) if amt else None
    has_amount = amt is not None and float(amt) > 0

    # ── Attribution ──
    has_scout_data = bool(scout_name_raw)
    has_driver_data = bool(driver_license_raw or driver_id_resolved)

    if not has_scout_data and not has_driver_data:
        attr_s = "attribution_rejected_missing_scout_and_driver"
        attr_r = "sin scout ni licencia"
    else:
        reasons = []
        if not scout_id:
            reasons.append(f"manual_review_no_scout_match (resolved_by={scout_method})")
        if not base["driver_id_resolved"]:
            reasons.append("manual_review_no_driver_match")
        if reasons:
            attr_s = "attribution_manual_review"
            attr_r = "; ".join(reasons)
        else:
            attr_s = "attribution_ready"
            attr_r = None

    base["attribution_status"] = attr_s
    base["attribution_reason"] = attr_r

    # ── CAPA B: Payment Financial ──
    fin_s = None
    fin_r = None
    if not has_amount:
        fin_s = "payment_financial_not_applicable_no_amount"
        fin_r = "monto = 0"
    elif not scout_id:
        fin_s = "payment_financial_manual_review_no_scout"
        fin_r = "scout no resuelto"
    else:
        fin_s = "payment_financial_ready"
        fin_r = None

    base["payment_financial_status"] = fin_s
    base["payment_financial_reason"] = fin_r

    # ── CAPA C: Payment Blocking ──
    blk_s = None
    blk_r = None
    blocks = False
    if not has_amount:
        blk_s = "payment_blocking_not_applicable_no_amount"
        blk_r = "monto = 0"
    elif is_explicitly_not_paid(estado_pago):
        blk_s = "payment_blocking_not_applicable_bad_status"
        blk_r = f"estado explic. negativo: {estado_pago}"
    elif not scout_id:
        blk_s = "payment_blocking_manual_review_no_scout"
        blk_r = "scout no resuelto"
    elif not base["driver_id_resolved"]:
        blk_s = "payment_blocking_manual_review_no_driver"
        blk_r = "driver no resuelto"
    else:
        blk_s = "payment_blocking_ready"
        blk_r = None
        blocks = True

    base["payment_blocking_status"] = blk_s
    base["payment_blocking_reason"] = blk_r
    base["blocks_future_payment"] = blocks

    # Legacy compat
    base["payment_status"] = blk_s or fin_s
    base["payment_reason"] = blk_r or fin_r

    # ── Final ──
    if attr_s == "attribution_rejected_missing_scout_and_driver":
        base["final_status"] = "rejected"
    elif attr_s == "attribution_ready":
        if blk_s == "payment_blocking_ready":
            base["final_status"] = "attribution_and_blocking_ready"
        elif fin_s == "payment_financial_ready":
            base["final_status"] = "attribution_and_financial_ready"
        else:
            base["final_status"] = "attribution_only_ready"
    elif attr_s == "attribution_manual_review":
        base["final_status"] = "manual_review"
    else:
        base["final_status"] = fin_s or "unknown"

    return base


def _preview_attributions_section(db: Session, ctx: WorkbookContext,
                                    rows: List[dict], source_file: str, sheet: str) -> dict:
    result = {"total_rows": len(rows), "ready": 0, "manual_review": 0,
              "conflicts": 0, "duplicates": 0, "rejected": 0, "lines": []}

    license_cache = {}
    driver_id_set = set()
    all_lic = set()
    all_did = set()
    for row in rows:
        lic = str(row.get("driver_license_raw", "")).strip()
        if lic: all_lic.add(lic)
        did = str(row.get("driver_id_resolved", "")).strip()
        if did: all_did.add(did)
    if all_lic:
        license_cache = build_license_cache(db, list(all_lic))
    if all_did:
        driver_id_set = build_driver_id_cache(db, list(all_did))

    for i, row in enumerate(rows):
        line = _classify_attribution_row_with_context(row, source_file, sheet, i + 2, ctx,
                                                       license_cache, driver_id_set)
        s = line.get("import_status", "pending")
        if s == "ready_to_import": result["ready"] += 1
        elif s == "manual_review": result["manual_review"] += 1
        elif s == "duplicate": result["duplicates"] += 1
        elif s == "rejected": result["rejected"] += 1
        result["lines"].append(line)

    return result


def _classify_attribution_row_with_context(row: dict, source_file: str, sheet: str,
                                            row_num: int, ctx: WorkbookContext,
                                            license_cache: Dict[str, str],
                                            driver_id_set: Set[str]) -> dict:
    scout_name_raw = str(row.get("scout_name_raw", row.get("SCOUT", ""))).strip()
    driver_license_raw = str(row.get("driver_license_raw", row.get("LICENCIA", ""))).strip()
    driver_id_raw = str(row.get("driver_id_resolved", "")).strip()
    origin_raw = str(row.get("origin_raw", "")).strip()

    scout_id, method = ctx.resolve_scout_for_preview(scout_name_raw)

    did = driver_id_raw if driver_id_raw and driver_id_raw in driver_id_set else None
    if not did and driver_license_raw:
        did = license_cache.get(driver_license_raw.strip())

    base = {
        "source_file": source_file, "source_sheet": sheet, "source_row": row_num,
        "scout_name_raw": scout_name_raw or None,
        "scout_id_resolved": scout_id,
        "driver_license_raw": driver_license_raw or None,
        "driver_id_resolved": did,
        "origin_raw": origin_raw or None,
        "scout_resolution_method": method,
    }

    if not scout_name_raw and not driver_license_raw:
        base["import_status"] = "rejected"
        base["import_reason"] = "sin scout ni licencia"
        return base

    reasons = []
    if not scout_id:
        reasons.append(f"manual_review_no_scout_match (resolved_by={method})")
    if not did:
        reasons.append("manual_review_no_driver_match")

    if reasons:
        base["import_status"] = "manual_review"
        base["import_reason"] = "; ".join(reasons)
    else:
        # Check conflict
        existing = ctx.db.query(DriverAssignment).filter(
            DriverAssignment.driver_id == did,
            DriverAssignment.status == "active",
        ).first()
        if existing and existing.scout_id != scout_id:
            base["import_status"] = "manual_review"
            base["import_reason"] = "manual_review_assignment_conflict"
        else:
            base["import_status"] = "ready_to_import"

    return base


def _preview_schemes_section(rows: List[dict]) -> dict:
    return {"total_rows": len(rows), "ready": len(rows), "rejected": 0}


def workbook_commit(db: Session, sheets_data: Dict[str, List[dict]],
                     source_file: str) -> Dict[str, Any]:
    """Commit all sheets in dependency order."""
    t0 = time.time()
    result = {
        "source_file": source_file,
        "scouts_created": 0, "scouts_updated": 0,
        "supervisors_created": 0, "scout_supervisor_links_created": 0,
        "supervisor_conflicts": 0,
        "schemes_created": 0,
        "historical_attributions_created": 0,
        "assignments_created": 0,
        "paid_history_created": 0,
        "manual_review_saved": 0,
        "conflicts": 0,
        "errors": [],
    }

    ctx = WorkbookContext(db)
    source_file = source_file or "workbook_import"

    # Step 1: Register context from 02_SCOUTS + 01_PAGOS_HISTORICOS
    for row in sheets_data.get("02_SCOUTS", []):
        sn = str(row.get("scout_name", row.get("SCOUT", ""))).strip()
        sp = str(row.get("supervisor_name_raw", row.get("SUPERVISOR", ""))).strip()
        if sn:
            ctx.register_scout_from_sheet(sn, sp)
    for row in sheets_data.get("01_PAGOS_HISTORICOS", []):
        sn = str(row.get("scout_name_raw", "")).strip()
        sp = str(row.get("supervisor_name_raw", "")).strip()
        if sn:
            ctx.register_scout_from_sheet(sn)
        if sn and sp:
            ctx.register_supervisor_from_payment(sn, sp)

    # Step 2: Create/update scouts
    seen = set()
    for row in sheets_data.get("02_SCOUTS", []):
        raw_name = str(row.get("scout_name", row.get("SCOUT", ""))).strip()
        if not raw_name:
            continue
        norm = normalize_name(raw_name)
        if norm in seen:
            continue
        seen.add(norm)
        existing_id = ctx.db_scout_cache.get(norm)
        if existing_id:
            scout = db.query(Scout).filter(Scout.id == existing_id).first()
            if scout:
                scout.scout_type = scout.scout_type or str(row.get("scout_type", "")).strip() or None
                scout.imported_from = "workbook_import"
                scout.source_sheet = "02_SCOUTS"
                scout.source_row = result["scouts_updated"] + 1
                result["scouts_updated"] += 1
        else:
            scout = Scout(
                scout_name=raw_name,
                scout_type=str(row.get("scout_type", "")).strip() or None,
                status="active",
                imported_from="workbook_import",
                source_sheet="02_SCOUTS",
                source_row=result["scouts_created"] + 1,
            )
            db.add(scout)
            db.flush()
            ctx.db_scout_cache[norm] = scout.id
            result["scouts_created"] += 1

    # Step 3: Link supervisors to scouts
    for scout_norm, sup_norm in ctx.scout_supervisor_links.items():
        scout_id = ctx.db_scout_cache.get(scout_norm)
        sup_id = ctx.db_scout_cache.get(sup_norm)
        if scout_id and sup_id:
            scout = db.query(Scout).filter(Scout.id == scout_id).first()
            if scout and not scout.supervisor_id:
                scout.supervisor_id = sup_id
                scout.supervisor_name_raw = ctx.supervisor_candidates.get(sup_norm, {}).get("raw_name", "")
                result["scout_supervisor_links_created"] += 1
        elif scout_id and not sup_id:
            result["supervisor_conflicts"] += 1

    # Step 4: Schemes
    for row in sheets_data.get("03_ESQUEMAS", []):
        sn = str(row.get("scheme_name", "")).strip()
        if sn:
            result["schemes_created"] += 1

    # Step 5: Attributions
    for sheet_key in ["06_ATRIBUCIONES_HISTORICAS", "01_PAGOS_HISTORICOS"]:
        for row in sheets_data.get(sheet_key, []):
            scout_raw = str(row.get("scout_name_raw", row.get("SCOUT", ""))).strip()
            lic_raw = str(row.get("driver_license_raw", row.get("LICENCIA", ""))).strip()
            did_raw = str(row.get("driver_id_resolved", "")).strip()
            if not scout_raw and not lic_raw:
                continue
            scout_id, _ = ctx.resolve_scout_for_preview(scout_raw)
            # Save historical attribution
            attr = HistoricalAttribution(
                source_file=source_file,
                source_sheet=sheet_key,
                source_row=result["historical_attributions_created"] + 1,
                scout_id_resolved=scout_id,
                scout_name_raw=scout_raw or None,
                driver_license_raw=lic_raw or None,
                driver_id_resolved=did_raw or None,
                import_status="imported" if scout_id else "manual_review",
                import_reason="workbook_commit" if scout_id else "manual_review_no_scout_match",
            )
            db.add(attr)
            result["historical_attributions_created"] += 1

    # Step 6: Payments (only for payment_ready rows - must match preview logic)
    # Build license cache for payment rows
    all_pay_licenses = set()
    all_pay_dids = set()
    for row in sheets_data.get("01_PAGOS_HISTORICOS", []):
        lic = str(row.get("driver_license_raw", "")).strip()
        if lic: all_pay_licenses.add(lic)
        did = str(row.get("driver_id_resolved", "")).strip()
        if did: all_pay_dids.add(did)
    pay_license_cache = build_license_cache(db, list(all_pay_licenses)) if all_pay_licenses else {}
    pay_driver_id_set = build_driver_id_cache(db, list(all_pay_dids)) if all_pay_dids else set()
    seen_hashes = set()

    for row in sheets_data.get("01_PAGOS_HISTORICOS", []):
        amount_raw = str(row.get("amount_paid", row.get("amount", "0"))).strip()
        amt = parse_decimal_safe(amount_raw)
        if not amt or float(amt) <= 0:
            continue
        estado_pago = normalize_text(str(row.get("estado_pago", "")).strip())
        if is_explicitly_not_paid(estado_pago):
            continue
        scout_raw = str(row.get("scout_name_raw", "")).strip()
        scout_id, _ = ctx.resolve_scout_for_preview(scout_raw)
        if not scout_id:
            continue
        # Resolve driver - MUST match preview logic
        lic_raw = str(row.get("driver_license_raw", "")).strip()
        did_raw = str(row.get("driver_id_resolved", "")).strip()
        did = None
        if did_raw and did_raw in pay_driver_id_set:
            did = did_raw
        elif lic_raw and lic_raw in pay_license_cache:
            did = pay_license_cache[lic_raw]
        has_driver = bool(did)

        # Check duplicate
        payment_rule = str(row.get("payment_rule", "")).strip() or None
        corte_id = str(row.get("cutoff_external_id", row.get("CORTE_ID", ""))).strip() or None
        import hashlib
        hash_val = hashlib.sha256(
            f"wb|01_PAGOS_HISTORICOS|{result['paid_history_created']}|{corte_id}|{lic_raw}|{scout_raw}|{amt}|{payment_rule}"
            .encode()).hexdigest()[:64]

        if hash_val in seen_hashes:
            continue
        seen_hashes.add(hash_val)

        ph = PaidHistory(
            cutoff_run_id=None,
            scout_id=scout_id,
            driver_id=did,
            driver_license_raw=lic_raw or None,
            scout_name_raw=scout_raw,
            amount_paid=amt,
            currency="PEN",
            paid_at=datetime.now(),
            import_source="historical_upload",
            source_file=source_file,
            source_sheet="01_PAGOS_HISTORICOS",
            source_row=result["paid_history_created"] + 1,
            payment_component="scout_driver_payment",
            payment_rule=payment_rule,
            cutoff_external_id=corte_id,
            unique_hash=hash_val,
            status="paid",
            resolution_status="resolved" if has_driver else "unresolved_driver",
            blocks_future_payment=has_driver,
            financial_record_status="recorded",
            original_payment_status_raw=estado_pago or None,
        )
        db.add(ph)
        result["paid_history_created"] += 1

    db.commit()
    elapsed = int((time.time() - t0) * 1000)
    logger.info(
        "[SCOUT_LIQ_WORKBOOK] commit_done file=%s scouts_created=%d scouts_updated=%d "
        "supervisor_links=%d attributions=%d paid_history=%d elapsed_ms=%d",
        source_file, result["scouts_created"], result["scouts_updated"],
        result["scout_supervisor_links_created"], result["historical_attributions_created"],
        result["paid_history_created"], elapsed,
    )

    return result
