"""
Payment Flow Service - Flujo completo de pagos del Liquidador Scouts Yego.

Estado del corte:
  draft -> reviewed -> approved -> paid
  draft -> cancelled
  reviewed -> cancelled
  approved -> cancelled (solo si no hay paid_history)

Reversas permitidas:
  reviewed -> draft
  approved -> reviewed (si no fue pagado)
  paid -> NO se revierte destructivamente

Regla de oro:
  - No pagar dos veces al mismo driver
  - No borrar paid_history
  - Toda linea debe tener motivo
  - Todo corte debe tener snapshot de esquema
"""
import csv as _csv
import io as _io
import json
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, List, Optional, Any

from sqlalchemy.orm import Session
from sqlalchemy import text

from app.models.scout_liq import (
    Scout, DriverAssignment, ConversionScheme, ConversionTier,
    CutoffRun, CutoffScoutSummary, CutoffDriverLine, PaidHistory,
    PaymentScheme, PaymentSchemeVersion, PaymentSchemeTier,
)
from app.adapters.source_adapter import compute_trip_counts_batch
from app.services.cutoff_engine import (
    calculate_cutoff,
    get_cutoff_summary,
    get_cutoff_lines,
    check_already_paid,
    _parse_rule,
    _get_trip_count_for_window,
    _compute_lifecycle,
)


# ═══════════════════════════════════════════════════════════════════════════
# VALIDACION DE ESTADOS Y TRANSICIONES
# ═══════════════════════════════════════════════════════════════════════════

VALID_STATUSES = {"draft", "calculated", "reviewed", "approved", "paid", "cancelled"}

ALLOWED_TRANSITIONS = {
    "draft": {"reviewed", "calculated", "cancelled"},
    "calculated": {"reviewed", "cancelled"},
    "reviewed": {"approved", "cancelled", "draft"},
    "approved": {"paid", "cancelled", "reviewed"},
    "paid": set(),
    "cancelled": set(),
}

ALLOWED_UNDO = {
    "reviewed": "draft",
    "approved": "reviewed",
}


def _validate_transition(current_status: str, target_status: str):
    if current_status not in VALID_STATUSES:
        raise ValueError(f"Estado '{current_status}' no reconocido")
    allowed = ALLOWED_TRANSITIONS.get(current_status, set())
    if target_status not in allowed:
        raise ValueError(
            f"No se permite transicion '{current_status}' -> '{target_status}'. "
            f"Permitidas: {allowed}"
        )


def _validate_undo(current_status: str) -> str:
    if current_status not in ALLOWED_UNDO:
        raise ValueError(
            f"No se puede deshacer estado '{current_status}'. "
            f"Solo permitido para: {list(ALLOWED_UNDO.keys())}"
        )
    return ALLOWED_UNDO[current_status]


# ═══════════════════════════════════════════════════════════════════════════
# A. CREAR BORRADOR DE PAGO
# ═══════════════════════════════════════════════════════════════════════════

def create_payment_draft(
    db: Session,
    hire_date_from: date,
    hire_date_to: date,
    scheme_id: int,
    origin: Optional[str] = None,
    country: Optional[str] = None,
    city: Optional[str] = None,
    scout_type: Optional[str] = None,
    notes: Optional[str] = None,
    created_by: Optional[str] = None,
) -> Dict[str, Any]:
    scheme = db.query(ConversionScheme).filter(ConversionScheme.id == scheme_id).first()
    if not scheme:
        raise ValueError(f"Scheme {scheme_id} no encontrado")

    tiers = db.query(ConversionTier).filter(
        ConversionTier.scheme_id == scheme_id,
        ConversionTier.active == True,
    ).order_by(ConversionTier.min_conversion_rate).all()

    if not tiers:
        raise ValueError(f"Scheme {scheme_id} no tiene tramos activos")

    config_snapshot = {
        "scheme_id": scheme.id,
        "scheme_name": scheme.scheme_name,
        "origin": scheme.origin,
        "scout_type": scheme.scout_type,
        "min_affiliations": scheme.min_affiliations,
        "min_activated": scheme.min_affiliations,
        "tiers": [
            {
                "min_conversion_rate": float(t.min_conversion_rate),
                "payout_amount": float(t.payment_per_converted_driver),
                "payment_per_converted_driver": float(t.payment_per_converted_driver),
                "currency": t.currency,
                "sort_order": 0,
            }
            for t in tiers
        ],
        "conversion_metric": "5plus_0_7",
        "activation_rule": "1V7D",
        "quality_rule": "5V7D",
        "volume_rule": "1V7D",
        "pays_on_rule": "ACTIVATED_BASE",
        "payout_formula_type": "ACTIVATED_X_TIER",
        "formula_type": "ACTIVATED_X_TIER",
        "currency": "PEN",
        "frozen_at": datetime.now().isoformat(),
    }

    cutoff_name = f"Pago {scheme.scheme_name} {hire_date_from} a {hire_date_to}"
    if scout_type:
        cutoff_name += f" ({scout_type})"

    run = CutoffRun(
        cutoff_name=cutoff_name,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        origin_filter=origin,
        country_filter=country,
        city_filter=city,
        scout_type_filter=scout_type,
        status="draft",
        config_snapshot=json.dumps(config_snapshot),
        created_by=created_by,
        notes=notes,
        quality_data_contract_status="ok",
        conversion_metric_code="5plus_0_7",
        conversion_metric_status="pending",
        source_mapping_snapshot=json.dumps({
            "driver_id": "module_ct_cabinet_drivers.driver_id",
            "hire_date": "module_ct_cabinet_drivers.hire_date",
            "origin": "module_ct_cabinet_drivers.origen",
            "conversion_rule": "5v7d_rate = converted_5v7d / activated_drivers",
            "payment_rule": "total_payable = activated_drivers x tier_amount",
        }),
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    result = calculate_cutoff(db, run.id)

    return {
        "cutoff_run_id": run.id,
        "cutoff_name": run.cutoff_name,
        "status": run.status,
        "hire_date_from": str(run.hire_date_from),
        "hire_date_to": str(run.hire_date_to),
        "origin_filter": run.origin_filter,
        "scheme_id": scheme_id,
        "scheme_name": scheme.scheme_name,
        "notes": notes,
        "calculation": result,
    }


# ═══════════════════════════════════════════════════════════════════════════
# B. RECALCULAR BORRADOR
# ═══════════════════════════════════════════════════════════════════════════

def recalculate_draft(db: Session, cutoff_run_id: int) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")
    if run.status not in ("draft", "calculated"):
        raise ValueError(f"No se puede recalcular en estado '{run.status}'. Solo permitido en draft/calculated")

    return calculate_cutoff(db, cutoff_run_id)


# ═══════════════════════════════════════════════════════════════════════════
# C. REVISAR CORTE  (draft -> reviewed)
# ═══════════════════════════════════════════════════════════════════════════

def review_cutoff_flow(db: Session, cutoff_run_id: int) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")

    allowed_from = {"draft", "calculated"}
    if run.status not in allowed_from:
        raise ValueError(f"No se puede revisar en estado '{run.status}'. Permitido desde: {allowed_from}")

    run.status = "reviewed"
    db.commit()
    return {"status": "reviewed", "cutoff_run_id": cutoff_run_id}


# ═══════════════════════════════════════════════════════════════════════════
# D. APROBAR CORTE  (reviewed -> approved)
# ═══════════════════════════════════════════════════════════════════════════

def approve_cutoff_flow(
    db: Session,
    cutoff_run_id: int,
    approved_by: Optional[str] = None,
) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")
    if run.status not in ("reviewed",):
        raise ValueError(f"No se puede aprobar en estado '{run.status}'. Solo desde 'reviewed'")

    run.status = "approved"
    run.approved_by = approved_by
    run.approved_at = datetime.now()

    summaries = db.query(CutoffScoutSummary).filter(
        CutoffScoutSummary.cutoff_run_id == cutoff_run_id,
    ).all()
    for s in summaries:
        s.amount_approved = s.amount_calculated

    db.commit()
    return {"status": "approved", "cutoff_run_id": cutoff_run_id}


# ═══════════════════════════════════════════════════════════════════════════
# E. MARCAR COMO PAGADO  (approved -> paid)
# ═══════════════════════════════════════════════════════════════════════════

def mark_cutoff_paid_flow(db: Session, cutoff_run_id: int, paid_by: Optional[str] = None) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")
    if run.status != "approved":
        raise ValueError(f"No se puede pagar en estado '{run.status}'. Solo desde 'approved'")
    if run.status == "paid":
        raise ValueError(f"Cutoff {cutoff_run_id} ya fue pagado. No se puede pagar dos veces.")

    config = json.loads(run.config_snapshot) if run.config_snapshot else {}
    scheme_id = config.get("scheme_id")
    scheme_name = config.get("scheme_name", "")
    scheme_type = config.get("scheme_type", "")

    paid_driver_count = 0
    total_paid_amount = Decimal("0")
    blocked_duplicates = 0

    summaries = db.query(CutoffScoutSummary).filter(
        CutoffScoutSummary.cutoff_run_id == cutoff_run_id,
    ).all()

    for s in summaries:
        if s.status == "blocked":
            continue

        lines = db.query(CutoffDriverLine).filter(
            CutoffDriverLine.cutoff_run_id == cutoff_run_id,
            CutoffDriverLine.scout_id == s.scout_id,
            CutoffDriverLine.payout_eligible_flag == True,
        ).all()

        for l in lines:
            already = check_already_paid(db, l.driver_id)
            if already:
                l.line_status = "blocked_already_paid"
                l.payment_status = "blocked"
                l.blocked_reason = "ya pagado en corte anterior (detectado al marcar pago)"
                l.payout_eligible_flag = False
                blocked_duplicates += 1
                continue

            ph = PaidHistory(
                cutoff_run_id=cutoff_run_id,
                scout_id=s.scout_id,
                driver_id=l.driver_id,
                origin=l.origin,
                payment_rule=l.payment_rule,
                amount_paid=l.calculated_amount or s.payout_per_activated,
                currency="PEN",
                paid_at=datetime.now(),
                import_source="cutoff_engine",
                payment_component="scout_driver_payment",
                payment_scheme_id=scheme_id,
                payment_scheme_name=scheme_name,
                payment_scheme_type=scheme_type or "quality_conversion",
                cutoff_window_from=run.hire_date_from,
                cutoff_window_to=run.hire_date_to,
                status="paid",
                blocks_future_payment=True,
                reason=f"Pago desde corte #{cutoff_run_id}",
                paid_by=paid_by,
            )
            db.add(ph)
            l.line_status = "paid"
            l.payment_status = "paid"
            paid_driver_count += 1
            total_paid_amount += ph.amount_paid or Decimal("0")

        s.status = "paid"
        s.total_payable = s.amount_calculated
        s.amount_approved = s.amount_calculated

    run.status = "paid"
    run.paid_at = datetime.now()
    db.commit()

    return {
        "status": "paid",
        "cutoff_run_id": cutoff_run_id,
        "drivers_paid": paid_driver_count,
        "total_paid_amount": float(total_paid_amount),
        "blocked_duplicates_at_payment": blocked_duplicates,
    }


# ═══════════════════════════════════════════════════════════════════════════
# F. CANCELAR CORTE
# ═══════════════════════════════════════════════════════════════════════════

def cancel_cutoff(db: Session, cutoff_run_id: int, reason: str) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")
    if run.status == "cancelled":
        raise ValueError(f"Cutoff {cutoff_run_id} ya esta cancelado")

    cancellable = {"draft", "calculated", "reviewed"}
    if run.status not in cancellable:
        if run.status == "approved":
            paid_count = db.query(PaidHistory).filter(
                PaidHistory.cutoff_run_id == cutoff_run_id,
            ).count()
            if paid_count > 0:
                raise ValueError(
                    f"No se puede cancelar approved con {paid_count} pagos ya realizados. "
                    f"Revierta los pagos primero."
                )
        else:
            raise ValueError(f"No se puede cancelar en estado '{run.status}'. Estados cancelables: {cancellable} o approved sin pagos")

    run.status = "cancelled"
    run.cancelled_at = datetime.now()
    run.cancelled_reason = reason
    db.commit()

    return {"status": "cancelled", "cutoff_run_id": cutoff_run_id, "reason": reason}


# ═══════════════════════════════════════════════════════════════════════════
# G. DESHACER ESTADO (UNDO)
# ═══════════════════════════════════════════════════════════════════════════

def undo_cutoff_status(db: Session, cutoff_run_id: int) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")

    if run.status == "paid":
        raise ValueError("NO se puede deshacer un corte pagado. El historial de pago es inmutable.")
    if run.status == "cancelled":
        raise ValueError("NO se puede deshacer un corte cancelado.")

    new_status = _validate_undo(run.status)
    run.status = new_status
    db.commit()

    return {"status": new_status, "cutoff_run_id": cutoff_run_id, "previous_status": run.status}


# ═══════════════════════════════════════════════════════════════════════════
# H. REPORTE DE CORTE
# ═══════════════════════════════════════════════════════════════════════════

def get_cutoff_report(db: Session, cutoff_run_id: int) -> Dict[str, Any]:
    run = db.query(CutoffRun).filter(CutoffRun.id == cutoff_run_id).first()
    if not run:
        raise ValueError(f"Cutoff run {cutoff_run_id} no encontrado")

    config = json.loads(run.config_snapshot) if run.config_snapshot else {}

    summaries = get_cutoff_summary(db, cutoff_run_id)
    lines = get_cutoff_lines(db, cutoff_run_id)

    total_calculated = sum(s["amount_calculated"] for s in summaries)
    total_approved = sum(s.get("amount_approved", 0) for s in summaries)

    paid_rows = db.query(PaidHistory).filter(
        PaidHistory.cutoff_run_id == cutoff_run_id,
    ).all()

    blocked_lines = [l for l in lines if l.get("line_status", "").startswith("blocked")]
    duplicate_blocked = [l for l in lines if l.get("already_paid")]

    return {
        "metadata": {
            "id": run.id,
            "cutoff_name": run.cutoff_name,
            "status": run.status,
            "hire_date_from": str(run.hire_date_from) if run.hire_date_from else None,
            "hire_date_to": str(run.hire_date_to) if run.hire_date_to else None,
            "origin_filter": run.origin_filter,
            "country_filter": run.country_filter,
            "city_filter": run.city_filter,
            "scout_type_filter": run.scout_type_filter,
            "notes": run.notes,
            "cohort_iso_week": run.cohort_iso_week,
            "cohort_from": str(run.cohort_from) if run.cohort_from else None,
            "cohort_to": str(run.cohort_to) if run.cohort_to else None,
            "created_by": run.created_by,
            "created_at": str(run.created_at) if run.created_at else None,
            "approved_by": run.approved_by,
            "approved_at": str(run.approved_at) if run.approved_at else None,
            "paid_at": str(run.paid_at) if run.paid_at else None,
            "cancelled_at": str(run.cancelled_at) if run.cancelled_at else None,
            "cancelled_reason": run.cancelled_reason,
        },
        "config_snapshot": config,
        "scout_summaries": summaries,
        "totals": {
            "scouts_evaluated": len(summaries),
            "drivers_total": len(lines),
            "drivers_payable": len([l for l in lines if l.get("payout_eligible_flag")]),
            "drivers_blocked": len(blocked_lines),
            "drivers_already_paid": len(duplicate_blocked),
            "drivers_paid_in_cutoff": len(paid_rows),
            "amount_calculated_total": total_calculated,
            "amount_approved_total": total_approved,
            "amount_paid_total": sum(float(r.amount_paid or 0) for r in paid_rows),
        },
        "paid_history": [
            {
                "id": r.id,
                "driver_id": r.driver_id,
                "scout_id": r.scout_id,
                "amount_paid": float(r.amount_paid) if r.amount_paid else 0,
                "paid_at": str(r.paid_at) if r.paid_at else None,
                "blocks_future_payment": r.blocks_future_payment,
            }
            for r in paid_rows
        ],
        "driver_lines": lines,
    }


# ═══════════════════════════════════════════════════════════════════════════
# I. REPORTE POR SCOUT
# ═══════════════════════════════════════════════════════════════════════════

def get_scout_payment_report(
    db: Session,
    scout_id: int,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> Dict[str, Any]:
    scout = db.query(Scout).filter(Scout.id == scout_id).first()
    if not scout:
        raise ValueError(f"Scout {scout_id} no encontrado")

    assignments = db.query(DriverAssignment).filter(
        DriverAssignment.scout_id == scout_id,
        DriverAssignment.status == "active",
    )
    if date_from:
        assignments = assignments.filter(DriverAssignment.hire_date >= date_from)
    if date_to:
        assignments = assignments.filter(DriverAssignment.hire_date <= date_to)
    assignments = assignments.all()

    driver_ids = [a.driver_id for a in assignments]

    paid_q = db.query(PaidHistory).filter(PaidHistory.scout_id == scout_id)
    if date_from:
        paid_q = paid_q.filter(PaidHistory.paid_at >= date_from)
    if date_to:
        paid_q = paid_q.filter(PaidHistory.paid_at <= date_to)
    paid_rows = paid_q.all()

    paid_driver_ids = set()
    total_paid_amount = Decimal("0")
    for p in paid_rows:
        if p.driver_id:
            paid_driver_ids.add(p.driver_id)
        total_paid_amount += p.amount_paid or Decimal("0")

    trip_counts = {}
    if driver_ids:
        trip_counts = compute_trip_counts_batch(db, driver_ids)

    converted = 0
    not_converted = 0
    for did in driver_ids:
        tc = trip_counts.get(did, {"trips_0_7_count": 0})
        trips_7d = tc.get("trips_0_7_count", 0) or 0
        if trips_7d >= 5:
            converted += 1
        elif trips_7d >= 1:
            not_converted += 1

    total_aff = len(driver_ids)
    conversion_rate = (Decimal(str(converted)) / Decimal(str(total_aff)) * 100) if total_aff > 0 else Decimal("0")

    return {
        "scout_id": scout_id,
        "scout_name": scout.scout_name,
        "date_from": str(date_from) if date_from else None,
        "date_to": str(date_to) if date_to else None,
        "drivers_total": total_aff,
        "drivers_converted_5v7d": converted,
        "drivers_not_converted": not_converted - converted if not_converted > converted else not_converted,
        "drivers_blocked_already_paid": len(paid_driver_ids & set(driver_ids)),
        "drivers_paid_in_window": len(paid_rows),
        "conversion_rate_5v7d_pct": float(round(conversion_rate, 2)),
        "total_paid_amount": float(total_paid_amount),
        "paid_history": [
            {
                "id": p.id,
                "driver_id": p.driver_id,
                "amount_paid": float(p.amount_paid) if p.amount_paid else 0,
                "paid_at": str(p.paid_at) if p.paid_at else None,
                "payment_rule": p.payment_rule,
                "blocks_future_payment": p.blocks_future_payment,
            }
            for p in paid_rows
        ],
        "driver_trips": [
            {
                "driver_id": did,
                "trips_7d": (trip_counts.get(did, {}) or {}).get("trips_0_7_count", 0) or 0,
                "trips_14d": ((trip_counts.get(did, {}) or {}).get("trips_0_7_count", 0) or 0) +
                             ((trip_counts.get(did, {}) or {}).get("trips_8_14_count", 0) or 0),
                "converted_5v7d": ((trip_counts.get(did, {}) or {}).get("trips_0_7_count", 0) or 0) >= 5,
                "already_paid": did in paid_driver_ids,
            }
            for did in driver_ids[:500]
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════
# J. REPORTE POR COHORTE
# ═══════════════════════════════════════════════════════════════════════════

def get_cohort_payment_report(db: Session, cohort_key: str) -> Dict[str, Any]:
    from app.services.cohort_service import iso_week_dates, cohort_maturity, get_iso_cohorts

    cohorts = get_iso_cohorts(db)
    cohort = next((c for c in cohorts if c["cohort_iso_week"] == cohort_key), None)

    if not cohort:
        from datetime import date as dt_date
        try:
            year, week = cohort_key.split("-W")
            monday, sunday = iso_week_dates(int(year), int(week))
            cohort = {
                "cohort_iso_week": cohort_key,
                "cohort_from": monday.isoformat(),
                "cohort_to": sunday.isoformat(),
                "readiness_status": "open",
            }
        except (ValueError, AttributeError):
            raise ValueError(f"Cohorte '{cohort_key}' no encontrada ni parseable")

    cohort_from = datetime.strptime(cohort["cohort_from"], "%Y-%m-%d").date() if isinstance(cohort["cohort_from"], str) else cohort["cohort_from"]
    cohort_to = datetime.strptime(cohort["cohort_to"], "%Y-%m-%d").date() if isinstance(cohort["cohort_to"], str) else cohort["cohort_to"]

    assignments = db.query(DriverAssignment).filter(
        DriverAssignment.status == "active",
        DriverAssignment.hire_date >= cohort_from,
        DriverAssignment.hire_date <= cohort_to,
    ).all()

    driver_ids = [a.driver_id for a in assignments]
    with_scout = len(set(a.driver_id for a in assignments if a.scout_id))
    without_scout = len(set(driver_ids)) - with_scout

    trip_counts = {}
    if driver_ids:
        trip_counts = compute_trip_counts_batch(db, driver_ids)

    converted = 0
    for did in driver_ids:
        tc = trip_counts.get(did, {})
        trips_7d = tc.get("trips_0_7_count", 0) or 0
        if trips_7d >= 5:
            converted += 1

    paid_rows = db.query(PaidHistory).filter(
        PaidHistory.cutoff_window_from >= cohort_from,
        PaidHistory.cutoff_window_to <= cohort_to,
    ).all()

    paid_driver_ids = set(p.driver_id for p in paid_rows if p.driver_id)
    blocked_count = len(paid_driver_ids & set(driver_ids))
    total_paid = sum(float(p.amount_paid or 0) for p in paid_rows)

    return {
        "cohort_key": cohort_key,
        "cohort_from": str(cohort_from),
        "cohort_to": str(cohort_to),
        "readiness_status": cohort.get("readiness_status", "unknown"),
        "drivers_total": len(set(driver_ids)),
        "drivers_with_scout": with_scout,
        "drivers_without_scout": without_scout,
        "drivers_converted_5v7d": converted,
        "drivers_payable": converted - blocked_count if converted > blocked_count else 0,
        "drivers_blocked_already_paid": blocked_count,
        "drivers_paid_in_window": len(paid_rows),
        "total_paid_amount": total_paid,
    }


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT CSV
# ═══════════════════════════════════════════════════════════════════════════

def export_cutoff_csv_full(db: Session, cutoff_run_id: int) -> str:
    report = get_cutoff_report(db, cutoff_run_id)

    buf = _io.StringIO()
    writer = _csv.writer(buf)

    writer.writerow(["=== REPORTE DE PAGO - CORTE #" + str(cutoff_run_id) + " ==="])
    writer.writerow([])

    meta = report["metadata"]
    writer.writerow(["Metadata:"])
    for k, v in meta.items():
        writer.writerow([k, str(v)])
    writer.writerow([])

    writer.writerow(["=== RESUMEN POR SCOUT ==="])
    writer.writerow([
        "scout_id", "scout_name", "origin", "afiliados", "activados",
        "convertidos_5v7d", "no_convertidos", "conversion_rate",
        "tramo", "pago_por_convertido", "total_calculado", "total_aprobado", "estado", "bloqueo"
    ])

    for s in report["scout_summaries"]:
        writer.writerow([
            s.get("scout_id", ""),
            s.get("scout_name", ""),
            s.get("origin", ""),
            s.get("total_affiliations", 0),
            s.get("total_activated", 0),
            s.get("drivers_5plus_0_7", 0),
            s.get("not_converted", 0),
            f"{s.get('conversion_rate', 0):.4f}",
            s.get("tier_reached", ""),
            f"{s.get('payout_per_activated', 0):.2f}",
            f"{s.get('amount_calculated', 0):.2f}",
            f"{s.get('amount_approved', 0):.2f}",
            s.get("status", ""),
            s.get("blocked_reason", ""),
        ])

    writer.writerow([])
    writer.writerow(["=== DETALLE POR DRIVER ==="])
    writer.writerow([
        "scout_id", "driver_id", "hire_date", "origin", "trips_7d", "trips_14d",
        "converted_5v7d", "converted_5v14d", "lifecycle", "line_status",
        "payment_status", "reason", "paid_amount", "already_paid", "eligible"
    ])

    for l in report["driver_lines"]:
        writer.writerow([
            l.get("scout_id", ""),
            l.get("driver_id", ""),
            l.get("hire_date", ""),
            l.get("origin", ""),
            l.get("trips_0_7_count", 0),
            l.get("trips_0_14_count", 0),
            l.get("is_converted_5trips_7d", False),
            l.get("is_converted_5trips_14d", False),
            l.get("driver_lifecycle_status", ""),
            l.get("line_status", ""),
            l.get("payment_status", ""),
            l.get("blocked_reason", ""),
            f"{l.get('calculated_amount', 0) or 0:.2f}",
            l.get("already_paid", False),
            l.get("eligible", False),
        ])

    writer.writerow([])
    totals = report["totals"]
    writer.writerow(["=== TOTALES ==="])
    for k, v in totals.items():
        writer.writerow([k, str(v)])

    writer.writerow([])
    writer.writerow(["=== CONFIGURACION (SNAPSHOT) ==="])
    cfg = report["config_snapshot"]
    writer.writerow(["scheme_id", cfg.get("scheme_id", "")])
    writer.writerow(["scheme_name", cfg.get("scheme_name", "")])
    writer.writerow(["min_affiliations", cfg.get("min_affiliations", "")])
    writer.writerow(["quality_rule", cfg.get("quality_rule", "")])
    writer.writerow(["activation_rule", cfg.get("activation_rule", "")])
    writer.writerow(["pays_on_rule", cfg.get("pays_on_rule", "")])
    writer.writerow(["formula_type", cfg.get("formula_type", "")])
    writer.writerow(["frozen_at", cfg.get("frozen_at", "")])
    writer.writerow([])
    writer.writerow(["=== TRAMOS (TIERS) ==="])
    writer.writerow(["min_conversion_rate", "payout_amount", "currency"])
    for t in cfg.get("tiers", []):
        writer.writerow([
            t.get("min_conversion_rate", ""),
            t.get("payout_amount", ""),
            t.get("currency", ""),
        ])

    writer.writerow([])
    writer.writerow(["=== PAGOS REALIZADOS ==="])
    writer.writerow(["payment_id", "driver_id", "scout_id", "amount", "paid_at", "blocks_future"])
    for p in report["paid_history"]:
        writer.writerow([
            p["id"], p["driver_id"], p["scout_id"],
            f"{p['amount_paid']:.2f}", p["paid_at"], p["blocks_future_payment"],
        ])

    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT XLSX
# ═══════════════════════════════════════════════════════════════════════════

def export_cutoff_xlsx(db: Session, cutoff_run_id: int) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

    report = get_cutoff_report(db, cutoff_run_id)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    # ── Sheet 1: Resumen por Scout ──
    ws1 = wb.active
    ws1.title = "Resumen Scout"
    headers1 = [
        "Scout ID", "Scout", "Origen", "Afiliados", "Activados",
        "Convertidos 5V/7D", "No Convertidos", "Conversion %",
        "Tramo", "Pago/Convertido", "Total Calc", "Total Aprob", "Estado", "Bloqueo"
    ]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))

    for r, s in enumerate(report["scout_summaries"], 2):
        ws1.cell(row=r, column=1, value=s.get("scout_id"))
        ws1.cell(row=r, column=2, value=s.get("scout_name"))
        ws1.cell(row=r, column=3, value=s.get("origin"))
        ws1.cell(row=r, column=4, value=s.get("total_affiliations", 0))
        ws1.cell(row=r, column=5, value=s.get("total_activated", 0))
        ws1.cell(row=r, column=6, value=s.get("drivers_5plus_0_7", 0))
        ws1.cell(row=r, column=7, value=s.get("not_converted", 0))
        ws1.cell(row=r, column=8, value=round(s.get("conversion_rate", 0) * 100, 2))
        ws1.cell(row=r, column=9, value=s.get("tier_reached"))
        ws1.cell(row=r, column=10, value=s.get("payout_per_activated", 0))
        ws1.cell(row=r, column=11, value=s.get("amount_calculated", 0))
        ws1.cell(row=r, column=12, value=s.get("amount_approved", 0))
        ws1.cell(row=r, column=13, value=s.get("status"))
        ws1.cell(row=r, column=14, value=s.get("blocked_reason"))

    # ── Sheet 2: Detalle Drivers ──
    ws2 = wb.create_sheet("Detalle Drivers")
    headers2 = [
        "Scout ID", "Driver ID", "Hire Date", "Origen", "Trips 7D", "Trips 14D",
        "Conv 5V/7D", "Conv 5V/14D", "Lifecycle", "Line Status",
        "Payment Status", "Motivo", "Paid Amount", "Already Paid", "Eligible"
    ]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    for r, l in enumerate(report["driver_lines"], 2):
        ws2.cell(row=r, column=1, value=l.get("scout_id"))
        ws2.cell(row=r, column=2, value=l.get("driver_id"))
        ws2.cell(row=r, column=3, value=l.get("hire_date"))
        ws2.cell(row=r, column=4, value=l.get("origin"))
        ws2.cell(row=r, column=5, value=l.get("trips_0_7_count", 0))
        ws2.cell(row=r, column=6, value=l.get("trips_0_14_count", 0))
        ws2.cell(row=r, column=7, value=l.get("is_converted_5trips_7d", False))
        ws2.cell(row=r, column=8, value=l.get("is_converted_5trips_14d", False))
        ws2.cell(row=r, column=9, value=l.get("driver_lifecycle_status"))
        ws2.cell(row=r, column=10, value=l.get("line_status"))
        ws2.cell(row=r, column=11, value=l.get("payment_status"))
        ws2.cell(row=r, column=12, value=l.get("blocked_reason"))
        ws2.cell(row=r, column=13, value=l.get("calculated_amount") or 0)
        ws2.cell(row=r, column=14, value=l.get("already_paid", False))
        ws2.cell(row=r, column=15, value=l.get("eligible", False))

    # ── Sheet 3: Config + Totales ──
    ws3 = wb.create_sheet("Config y Auditoria")
    r = 1
    ws3.cell(row=r, column=1, value="CAMPO").font = header_font
    ws3.cell(row=r, column=2, value="VALOR").font = header_font
    r += 1

    cfg_fields = [
        ("Cutoff ID", report["metadata"]["id"]),
        ("Nombre", report["metadata"]["cutoff_name"]),
        ("Estado", report["metadata"]["status"]),
        ("Ventana Desde", report["metadata"]["hire_date_from"]),
        ("Ventana Hasta", report["metadata"]["hire_date_to"]),
        ("Origen", report["metadata"]["origin_filter"]),
        ("Scheme ID", report["config_snapshot"].get("scheme_id")),
        ("Scheme Name", report["config_snapshot"].get("scheme_name")),
        ("Min Afiliaciones", report["config_snapshot"].get("min_affiliations")),
        ("Quality Rule", report["config_snapshot"].get("quality_rule")),
        ("Activation Rule", report["config_snapshot"].get("activation_rule")),
        ("Pays On Rule", report["config_snapshot"].get("pays_on_rule")),
        ("Formula Type", report["config_snapshot"].get("formula_type")),
        ("Frozen At", report["config_snapshot"].get("frozen_at")),
        ("Creado", report["metadata"]["created_at"]),
        ("Aprobado", report["metadata"]["approved_at"]),
        ("Pagado", report["metadata"]["paid_at"]),
        ("Cancelado", report["metadata"]["cancelled_at"]),
    ]
    for field, val in cfg_fields:
        r += 1
        ws3.cell(row=r, column=1, value=field)
        ws3.cell(row=r, column=2, value=str(val) if val is not None else "")

    r += 2
    ws3.cell(row=r, column=1, value="TOTALES").font = header_font
    r += 1
    for k, v in report["totals"].items():
        r += 1
        ws3.cell(row=r, column=1, value=k)
        ws3.cell(row=r, column=2, value=str(v))

    r += 2
    ws3.cell(row=r, column=1, value="TRAMOS").font = header_font
    r += 1
    for c, h in enumerate(["Min Conversion", "Payout", "Moneda"], 1):
        ws3.cell(row=r, column=c, value=h).font = header_font
    for t in report["config_snapshot"].get("tiers", []):
        r += 1
        ws3.cell(row=r, column=1, value=t.get("min_conversion_rate"))
        ws3.cell(row=r, column=2, value=t.get("payout_amount"))
        ws3.cell(row=r, column=3, value=t.get("currency", "PEN"))

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# REPORTE DE HISTORIAL DE PAGOS
# ═══════════════════════════════════════════════════════════════════════════

def get_payment_history_report(
    db: Session,
    scout_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    limit: int = 100,
    offset: int = 0,
) -> Dict[str, Any]:
    q = db.query(PaidHistory)
    if scout_id:
        q = q.filter(PaidHistory.scout_id == scout_id)
    if date_from:
        q = q.filter(PaidHistory.paid_at >= date_from)
    if date_to:
        q = q.filter(PaidHistory.paid_at <= date_to)

    total = q.count()
    rows = q.order_by(PaidHistory.created_at.desc()).offset(offset).limit(limit).all()

    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r.id,
                "cutoff_run_id": r.cutoff_run_id,
                "scout_id": r.scout_id,
                "driver_id": r.driver_id,
                "amount_paid": float(r.amount_paid) if r.amount_paid else 0,
                "currency": r.currency,
                "paid_at": str(r.paid_at) if r.paid_at else None,
                "import_source": r.import_source,
                "payment_rule": r.payment_rule,
                "blocks_future_payment": r.blocks_future_payment,
                "reason": r.reason,
                "status": r.status,
            }
            for r in rows
        ],
    }
